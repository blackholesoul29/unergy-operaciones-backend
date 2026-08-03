"""
Facturación de energía v2 — reproduce el Excel de facturación con datos de la
plataforma:  facturación = kWh(despacho) × tarifa_indexada,  donde
tarifa_indexada = round(tarifa_base_PPA × IPP_mes / IPP_base_PPA, 2)  (redondeo a
2 decimales ANTES de multiplicar, idéntico al Excel de la usuaria).

Fuentes:
  - kWh:  despacho_contrato_mensual (ingerido del archivo XM dspcttos_txf_MM.xlsx)
  - contrato XM → proyecto/PPA:  AsicSolicitud (codigo_sic_contrato, vigente)
  - tarifa base del mes + IPP base:  PPAContrato / PPATarifa
  - IPP del mes:  ipp_mensual
"""
import io
import logging
from calendar import monthrange
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import delete

from app.core.database import get_db
from app.api.v1.auth import get_current_user
from app.models import PPAContrato, PPATarifa, Proyecto, AsicSolicitud
from app.models.contratos import (
    DespachoContratoMensual, IppMensual, FacturaAgrupacion, FacturaOrden, FacturaEmitida,
)
from app.models.asic import EstadoSolicitudAsicEnum, TipoSolicitudAsicEnum
from app.services.facturacion_factura import construir_mensaje, contribuciones
from app.utils.gescon_vigencia import resolver_vigencias
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/facturacion", tags=["Facturación"])


def _norm_periodo(periodo: str) -> str:
    try:
        y, m = periodo.strip().split("-")
        return f"{int(y):04d}-{int(m):02d}"
    except Exception:
        raise HTTPException(422, "El período debe tener formato YYYY-MM")


# ── Ingesta del despacho XM ────────────────────────────────────────────────────
@router.post("/despacho")
async def cargar_despacho(
    periodo: str = Query(..., description="YYYY-MM del despacho"),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Sube el archivo de despachos de XM (xlsx, hoja con CONTRATO + DESP_HORA 01..24)
    y guarda la energía mensual por contrato (suma de las 24 horas de todo el mes).
    Reemplaza el despacho previo del período."""
    import openpyxl
    per = _norm_periodo(periodo)
    contenido = await archivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el Excel: {e}")
    ws = wb[wb.sheetnames[0]]

    filas = ws.iter_rows(values_only=True)
    try:
        header = list(next(filas))
    except StopIteration:
        raise HTTPException(400, "El archivo está vacío")
    # Localizar columnas por encabezado.
    def _find(names):
        for i, h in enumerate(header):
            if isinstance(h, str) and h.strip().upper() in names:
                return i
        return None
    i_con = _find({"CONTRATO"})
    i_ven = _find({"VENDEDOR"})
    i_com = _find({"COMPRADOR"})
    i_tipo = _find({"TIPO"})
    i_fec = _find({"FECHADOCUMENTO", "FECHA DOCUMENTO", "FECHA"})
    if i_con is None:
        raise HTTPException(400, "No encontré la columna CONTRATO en el archivo")
    horas = [i for i, h in enumerate(header)
             if isinstance(h, str) and (h.strip().upper().startswith("DESP_HORA")
                                        or h.strip().upper().replace(" ", "").startswith("H0")
                                        or h.strip().upper().replace(" ", "") in {f"H{n}" for n in range(24)})]
    if not horas:
        raise HTTPException(400, "No encontré columnas de horas (DESP_HORA 01..24)")

    agg: dict = {}
    for r in filas:
        if not r or i_con >= len(r) or r[i_con] is None:
            continue
        con = str(int(r[i_con])) if isinstance(r[i_con], float) else str(r[i_con]).strip()
        kwh = sum(r[i] for i in horas if i < len(r) and isinstance(r[i], (int, float)))
        d = agg.setdefault(con, {"kwh": 0.0, "vendedor": None, "comprador": None, "tipo": None, "fechas": set()})
        d["kwh"] += kwh
        if i_fec is not None and i_fec < len(r) and r[i_fec] is not None:
            fv = r[i_fec]
            fecha = None
            if hasattr(fv, "date"):          # datetime
                fecha = fv.date()
            elif hasattr(fv, "year"):        # date
                fecha = fv
            else:
                try:
                    fecha = date.fromisoformat(str(fv)[:10])
                except ValueError:
                    fecha = None
            if fecha is not None:
                d["fechas"].add(fecha)
        if i_ven is not None and i_ven < len(r) and r[i_ven]:
            d["vendedor"] = str(r[i_ven]).strip()
        if i_com is not None and i_com < len(r) and r[i_com]:
            d["comprador"] = str(r[i_com]).strip()
        if i_tipo is not None and i_tipo < len(r) and r[i_tipo]:
            d["tipo"] = str(r[i_tipo]).strip()

    db.execute(delete(DespachoContratoMensual).where(DespachoContratoMensual.periodo == per))
    for con, d in agg.items():
        fechas = d.get("fechas") or set()
        db.add(DespachoContratoMensual(
            periodo=per, codigo_sic_contrato=con, kwh=round(d["kwh"], 4),
            vendedor=d["vendedor"], comprador=d["comprador"], tipo=d["tipo"],
            dias=(len(fechas) or None),
            fecha_min=(min(fechas) if fechas else None),
            fecha_max=(max(fechas) if fechas else None),
            archivo=archivo.filename,
        ))
    db.commit()
    total = sum(d["kwh"] for d in agg.values())
    return {"periodo": per, "contratos": len(agg), "kwh_total": round(total, 2), "archivo": archivo.filename}


@router.get("/despacho")
def listar_despacho(periodo: str = Query(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    per = _norm_periodo(periodo)
    rows = (
        db.query(DespachoContratoMensual)
        .filter(DespachoContratoMensual.periodo == per)
        .order_by(DespachoContratoMensual.codigo_sic_contrato).all()
    )
    return {
        "periodo": per,
        "kwh_total": round(sum(float(r.kwh) for r in rows), 2),
        "contratos": [
            {"contrato": r.codigo_sic_contrato, "vendedor": r.vendedor, "comprador": r.comprador,
             "tipo": r.tipo, "kwh": float(r.kwh)} for r in rows
        ],
        "archivo": rows[0].archivo if rows else None,
    }


# ── Cálculo de facturación ─────────────────────────────────────────────────────
def _facturacion_periodo(db: Session, per: str) -> dict:
    y, m = per.split("-")
    año, mes = int(y), int(m)

    despacho = db.query(DespachoContratoMensual).filter(DespachoContratoMensual.periodo == per).all()
    ipp_row = db.query(IppMensual).filter(IppMensual.año == año, IppMensual.mes == mes).first()
    ipp = float(ipp_row.valor) if ipp_row else None

    # contrato XM → solicitud asic VIGENTE al fin del período (resolver_vigencias
    # corre sobre el universo publicado, igual que la vista de asic).
    last_day = date(año, mes, monthrange(año, mes)[1])
    universo = (
        db.query(AsicSolicitud)
        .filter(
            AsicSolicitud.estado_solicitud == EstadoSolicitudAsicEnum.publicado,
            AsicSolicitud.tipo_solicitud != TipoSolicitudAsicEnum.desistimiento,
        )
        .order_by(
            AsicSolicitud.fecha_inicio.asc().nullsfirst(),
            AsicSolicitud.fecha_solicitud.asc().nullsfirst(),
            AsicSolicitud.created_at.asc(),
        ).all()
    )
    vig = resolver_vigencias(universo, hasta=last_day)
    sol: dict = {}
    for s in universo:
        v = vig.get(s.id)
        if v is not None and v.vigente:
            c = (s.codigo_sic_contrato or "").strip()
            if c:
                sol[c] = s
    ppas = {p.id: p for p in db.query(PPAContrato).all()}
    tarifas: dict = {}
    for t in db.query(PPATarifa).filter(PPATarifa.año == año, PPATarifa.mes == mes).all():
        tarifas[t.contrato_id] = float(t.tarifa) if t.tarifa is not None else None
    proy_nombre = {pid: nom for pid, nom in db.query(Proyecto.id, Proyecto.nombre_comercial).all()}
    # Agrupación manual de facturas: codigo_sic_contrato → (nombre, porcentaje). El
    # porcentaje se lee acá porque `contribuciones` lo necesita para partir el
    # contrato; cargarlo solo como nombre reventaba el reparto.
    agrup = {
        a.codigo_sic_contrato: (a.nombre, float(a.porcentaje) if a.porcentaje is not None else None)
        for a in db.query(FacturaAgrupacion).all()
    }

    lineas = []
    for d in despacho:
        c = d.codigo_sic_contrato
        kwh = float(d.kwh)
        s = sol.get(c)
        pid = s.contrato_ppa_id if s else None
        proyid = s.proyecto_id if s else None
        ppa = ppas.get(pid) if pid else None
        base = tarifas.get(pid) if pid else None
        ipp_base = float(ppa.valor_indexacion_base) if (ppa and ppa.valor_indexacion_base) else None

        estado = "ok"
        tarifa_idx = fact = None
        if not pid:
            estado = "sin_ppa"
        elif base is None:
            estado = "sin_tarifa"
        elif not ipp_base:
            estado = "sin_ipp_base"
        elif ipp is None:
            estado = "sin_ipp_mes"
        else:
            tarifa_idx = round(base * ipp / ipp_base, 2)   # redondeo idéntico al Excel
            fact = round(kwh * tarifa_idx, 2)
        ppa_nom = (ppa.nombre_interno or ppa.numero_codigo_contrato) if ppa else None
        ag = agrup.get(c)
        lineas.append({
            "contrato": c,
            "comprador": d.comprador,
            "proyecto_id": proyid,
            "proyecto": proy_nombre.get(proyid) if proyid else None,
            "ppa": ppa_nom,
            # Nombre de la factura: agrupación manual del CONTRATO si existe, si no el PPA.
            "factura": (ag[0] if ag and ag[0] else ppa_nom),
            # Datos del contrato para el mensaje que se pega en la factura.
            "numero_contrato": ppa.numero_codigo_contrato if ppa else None,
            "periodo_ipp_base": ppa.periodo_indexacion_base if ppa else None,
            # Días efectivos del despacho (para el mensaje; mes parcial).
            "dias": d.dias,
            "fecha_min": d.fecha_min,
            "fecha_max": d.fecha_max,
            "kwh": kwh,
            "tarifa_base": base,
            "ipp_base": ipp_base,
            "ipp_mes": ipp,
            "tarifa_indexada": tarifa_idx,
            "facturacion": fact,
            "estado": estado,
        })

    def _suma(rows):
        return round(sum(r["facturacion"] or 0 for r in rows), 2)
    facturables = [l for l in lineas if l["estado"] == "ok"]

    # Agrupación por comprador (código SIC).
    por_sic: dict = {}
    for l in facturables:
        k = l["comprador"] or "—"
        g = por_sic.setdefault(k, {"comprador": k, "kwh": 0.0, "facturacion": 0.0, "contratos": 0})
        g["kwh"] += l["kwh"]; g["facturacion"] += l["facturacion"]; g["contratos"] += 1
    for g in por_sic.values():
        g["kwh"] = round(g["kwh"], 2); g["facturacion"] = round(g["facturacion"], 2)

    # Agrupación por FACTURA — replica la hoja "Facturación" del Excel. Una fila por
    # factura (agrupación manual si existe, si no el PPA). Un contrato con agrupación
    # PARCIAL (porcentaje) reparte su kWh/valor: el % va a la factura nombrada y el
    # resto (100-%) queda en el PPA. Misma tarifa en ambas partes.
    por_factura: dict = {}
    for l in facturables:
        a = agrup.get(l["contrato"])           # (nombre, porcentaje) | None
        ppa_nom = l["ppa"] or "—"
        contribs = contribuciones(a, ppa_nom)
        for (fname, fr, pct) in contribs:
            es_custom = bool(a and a[0] and fname == a[0])
            g = por_factura.setdefault(fname, {
                "factura": fname, "ppa": ppa_nom, "comprador": l["comprador"], "contratos": 0,
                "kwh": 0.0, "tarifa_base": l["tarifa_base"], "ipp_base": l["ipp_base"],
                "tarifa_indexada": l["tarifa_indexada"], "facturacion": 0.0,
                "personalizada": False, "sin_ppa": False, "_tarifas": set(), "proyectos": [],
                "periodo_ipp_base": l["periodo_ipp_base"],
                "_numeros": [], "_sic": [], "_dias": set(), "_fmin": None, "_fmax": None,
            })
            if l.get("dias"):
                g["_dias"].add(l["dias"])
            if l.get("fecha_min"):
                g["_fmin"] = l["fecha_min"] if g["_fmin"] is None else min(g["_fmin"], l["fecha_min"])
            if l.get("fecha_max"):
                g["_fmax"] = l["fecha_max"] if g["_fmax"] is None else max(g["_fmax"], l["fecha_max"])
            if l["numero_contrato"] and l["numero_contrato"] not in g["_numeros"]:
                g["_numeros"].append(l["numero_contrato"])
            if l["contrato"] and l["contrato"] not in g["_sic"]:
                g["_sic"].append(l["contrato"])
            kwh_p = round(l["kwh"] * fr, 2); fact_p = round(l["facturacion"] * fr, 2)
            g["kwh"] += kwh_p; g["facturacion"] += fact_p; g["contratos"] += 1
            if es_custom:
                g["personalizada"] = True
            if l["tarifa_indexada"] is not None:
                g["_tarifas"].add(l["tarifa_indexada"])
            g["proyectos"].append({
                "proyecto_id": l["proyecto_id"], "proyecto": l["proyecto"],
                "contrato": l["contrato"], "ppa": l["ppa"], "tarifa_indexada": l["tarifa_indexada"],
                "kwh": kwh_p, "facturacion": fact_p, "porcentaje": pct,
                "asignada": es_custom,
            })
    # Contratos SIN PPA (ej. UNGC / bolsa): XM SÍ los factura, así que se muestran
    # como factura agrupada por comprador para que la energía total coincida con XM.
    # No tienen tarifa PPA → la facturación queda pendiente ($0); solo cuenta la energía.
    for l in lineas:
        if l["estado"] != "sin_ppa":
            continue
        fname = f"{l['comprador'] or 'Sin PPA'} (sin PPA)"
        g = por_factura.setdefault(fname, {
            "factura": fname, "ppa": None, "comprador": l["comprador"], "contratos": 0,
            "kwh": 0.0, "tarifa_base": None, "ipp_base": None,
            "tarifa_indexada": None, "facturacion": 0.0,
            "personalizada": False, "sin_ppa": True, "_tarifas": set(), "proyectos": [],
            "periodo_ipp_base": None,
            "_numeros": [], "_sic": [], "_dias": set(), "_fmin": None, "_fmax": None,
        })
        if l.get("dias"):
            g["_dias"].add(l["dias"])
        if l.get("fecha_min"):
            g["_fmin"] = l["fecha_min"] if g["_fmin"] is None else min(g["_fmin"], l["fecha_min"])
        if l.get("fecha_max"):
            g["_fmax"] = l["fecha_max"] if g["_fmax"] is None else max(g["_fmax"], l["fecha_max"])
        if l["contrato"] and l["contrato"] not in g["_sic"]:
            g["_sic"].append(l["contrato"])
        g["kwh"] += l["kwh"]; g["contratos"] += 1
        g["proyectos"].append({
            "proyecto_id": l["proyecto_id"], "proyecto": l["proyecto"],
            "contrato": l["contrato"], "ppa": None, "tarifa_indexada": None,
            "kwh": l["kwh"], "facturacion": 0.0, "porcentaje": None, "asignada": False,
        })

    # Orden manual (fijo) y marca de emitida (del período).
    orden_manual = {o.nombre: o.orden for o in db.query(FacturaOrden).all()}
    emitidas = {
        e.nombre: {"por": e.emitida_por, "at": e.emitida_at.isoformat() if e.emitida_at else None}
        for e in db.query(FacturaEmitida).filter(FacturaEmitida.periodo == per).all()
    }

    for g in por_factura.values():
        g["kwh"] = round(g["kwh"], 2); g["facturacion"] = round(g["facturacion"], 2)
        # Si la factura mezcla contratos de PPAs con tarifas distintas, no hay una
        # sola tarifa: se marca "mixta" (la UI muestra 'varía').
        g["tarifa_mixta"] = len(g["_tarifas"]) > 1
        if g["tarifa_mixta"]:
            g["tarifa_indexada"] = None
        g.pop("_tarifas", None)

        em = emitidas.get(g["factura"])
        g["emitida"] = em is not None
        g["emitida_por"] = em["por"] if em else None
        g["emitida_at"] = em["at"] if em else None
        g["orden"] = orden_manual.get(g["factura"])
        g["numeros_contrato"] = g.pop("_numeros")
        g["contratos_sic"] = g.pop("_sic")
        # Días efectivos del despacho (máx entre los contratos; comparten el archivo).
        _dias = g.pop("_dias", set())
        g["dias"] = max(_dias) if _dias else None
        g["fecha_min"] = g.pop("_fmin", None)
        g["fecha_max"] = g.pop("_fmax", None)
        g["fecha_min"] = g["fecha_min"].isoformat() if g["fecha_min"] else None
        g["fecha_max"] = g["fecha_max"].isoformat() if g["fecha_max"] else None
        # El mensaje se arma acá (no en el frontend) para que el formato viva en un
        # solo lugar, fijado por test. La tarifa base del grupo puede ser mixta; en
        # ese caso el mensaje igual sale y la UI avisa que revise.
        g["mensaje"] = construir_mensaje(
            numeros_contrato=g["numeros_contrato"],
            periodo=per,
            kwh=g["kwh"],
            contratos_sic=g["contratos_sic"],
            tarifa_base=g["tarifa_base"],
            ipp_base=g["ipp_base"],
            periodo_ipp_base=g["periodo_ipp_base"],
            ipp_mes=ipp,
            dias=g["dias"],
            fecha_min=g["fecha_min"],
            fecha_max=g["fecha_max"],
        )

    # Las que tienen orden manual van primero (por ese orden); el resto por valor.
    facturas = sorted(
        por_factura.values(),
        key=lambda x: (x["orden"] is None, x["orden"] if x["orden"] is not None else 0,
                       -x["facturacion"]),
    )

    return {
        "periodo": per,
        "ipp_mes": ipp,
        "resumen": {
            "contratos": len(lineas),
            "facturables": len(facturables),
            "sin_ppa": sum(1 for l in lineas if l["estado"] == "sin_ppa"),
            "kwh_total": round(sum(l["kwh"] for l in facturables), 2),
            "facturacion_total": _suma(facturables),
            "emitidas": sum(1 for g in por_factura.values() if g["emitida"]),
            "facturas": len(por_factura),
        },
        "lineas": lineas,
        "por_codigo_sic": sorted(por_sic.values(), key=lambda x: -x["facturacion"]),
        "por_factura": facturas,
    }


@router.get("")
def facturacion(periodo: str = Query(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    return _facturacion_periodo(db, _norm_periodo(periodo))


# ── Agrupación manual de facturas (dividir un PPA en sub-facturas) ─────────────
class AgrupacionIn(BaseModel):
    codigo_sic_contrato: str
    nombre: str | None = None   # None/"" = quitar la asignación (vuelve al PPA)
    # % del contrato que va a esta factura; el resto queda en el PPA. None = todo.
    # Ej. Uruaco 78596 → 22.8066% a "Terpel 1 Suno", 77.1934% en Terpel 1.
    porcentaje: float | None = None


def _agrup_out(a: FacturaAgrupacion) -> dict:
    return {
        "codigo_sic_contrato": a.codigo_sic_contrato,
        "nombre": a.nombre,
        "porcentaje": float(a.porcentaje) if a.porcentaje is not None else None,
    }


@router.get("/agrupaciones")
def listar_agrupaciones(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return [_agrup_out(a)
            for a in db.query(FacturaAgrupacion).order_by(FacturaAgrupacion.nombre).all()]


@router.put("/agrupaciones")
def guardar_agrupaciones(rows: list[AgrupacionIn], db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Asigna CONTRATOS (código SIC) a una factura con nombre (upsert por contrato).
    nombre vacío quita la asignación (el contrato vuelve a agrupar por su PPA)."""
    for r in rows:
        c = (r.codigo_sic_contrato or "").strip()
        if not c:
            continue
        pct = r.porcentaje
        if pct is not None and not 0 < pct <= 100:
            raise HTTPException(422, f"El porcentaje de {c} debe estar entre 0 y 100 (llegó {pct})")
        obj = db.query(FacturaAgrupacion).filter(FacturaAgrupacion.codigo_sic_contrato == c).first()
        nombre = (r.nombre or "").strip()
        if not nombre:
            if obj:
                db.delete(obj)
            continue
        if obj is None:
            db.add(FacturaAgrupacion(codigo_sic_contrato=c, nombre=nombre, porcentaje=pct))
        else:
            obj.nombre = nombre
            obj.porcentaje = pct
    db.commit()
    return [_agrup_out(a)
            for a in db.query(FacturaAgrupacion).order_by(FacturaAgrupacion.nombre).all()]


# ── Orden manual de las facturas (fijo, aplica cada mes) ───────────────────────
class OrdenIn(BaseModel):
    nombres: list[str]   # en el orden deseado; las que no vengan quedan al final


@router.put("/orden")
def guardar_orden(data: OrdenIn, db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Fija el orden de las facturas. Se reemplaza completo: llega la lista ordenada
    y se reescribe, así no quedan huecos ni posiciones viejas."""
    db.execute(delete(FacturaOrden))
    for i, nombre in enumerate(data.nombres):
        n = (nombre or "").strip()
        if n:
            db.add(FacturaOrden(nombre=n, orden=i))
    db.commit()
    return {"guardadas": len([n for n in data.nombres if (n or "").strip()])}


@router.delete("/orden")
def limpiar_orden(db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Vuelve al orden por valor (descendente)."""
    db.execute(delete(FacturaOrden))
    db.commit()
    return {"ok": True}


# ── Marca de "ya se facturó" (por período) ─────────────────────────────────────
class EmitidaIn(BaseModel):
    nombre: str
    periodo: str
    emitida: bool


@router.put("/emitida")
def marcar_emitida(data: EmitidaIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    per = _norm_periodo(data.periodo)
    nombre = (data.nombre or "").strip()
    if not nombre:
        raise HTTPException(422, "Falta el nombre de la factura")
    obj = (
        db.query(FacturaEmitida)
        .filter(FacturaEmitida.nombre == nombre, FacturaEmitida.periodo == per)
        .first()
    )
    if data.emitida and obj is None:
        db.add(FacturaEmitida(
            nombre=nombre, periodo=per,
            emitida_por=getattr(user, "nombre", None) or getattr(user, "email", None),
        ))
    elif not data.emitida and obj is not None:
        db.delete(obj)
    db.commit()
    return {"nombre": nombre, "periodo": per, "emitida": data.emitida}

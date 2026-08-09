"""
API de procesamiento de facturas Starlink.

Endpoints:
  POST /starlink/procesar-pdf          → parsea el PDF y devuelve ítems + agrupado
  POST /starlink/excel                 → genera y descarga el Excel con dos hojas
  GET  /starlink/periodos              → lista períodos con datos guardados
  GET  /starlink/factura/{periodo}     → datos guardados de un período
  PUT  /starlink/factura/{periodo}     → guardar (crear/sobreescribir) un período
  DELETE /starlink/factura/{periodo}   → eliminar datos de un período
"""
from __future__ import annotations

import io
import json
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.v1.auth import get_current_user
from app.core.database import get_db
from app.models.starlink import StarlinkFactura, StarlinkMapeoSitio, StarlinkFacturaLinea
from app.models.proyectos import Proyecto
from app.services.starlink_resolver import resolver_lineas
from app.services.starlink_parser import parsear_pdf, ResultadoStarlink

router = APIRouter(prefix="/starlink", tags=["Starlink"])


def _regenerar_lineas(db: Session, fac: StarlinkFactura) -> None:
    """(Re)genera starlink_factura_linea para una factura desde su agrupado_json,
    resolviendo cada sitio contra el catálogo starlink_mapeo_sitio (match por nombre)."""
    agrupado = json.loads(fac.agrupado_json)
    mapeos = [
        {"patron": m.patron, "proyecto_id": m.proyecto_id}
        for m in db.query(StarlinkMapeoSitio).filter(StarlinkMapeoSitio.activo.is_(True)).all()
    ]
    db.query(StarlinkFacturaLinea).filter(StarlinkFacturaLinea.factura_id == fac.id).delete()
    for ln in resolver_lineas(agrupado, mapeos):
        db.add(StarlinkFacturaLinea(
            factura_id=fac.id,
            proyecto_id=ln["proyecto_id"],
            descripcion=ln["descripcion"],
            sin_iva=ln["sin_iva"],
            iva=ln["iva"],
            monto_total=ln["monto_total"],
        ))


# ── POST /starlink/procesar-pdf ───────────────────────────────────────────────

@router.post("/procesar-pdf")
async def procesar_pdf(
    file: UploadFile = File(...),
    _=Depends(get_current_user),
):
    """Recibe un PDF de factura Starlink y devuelve los ítems parseados."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "El archivo debe ser un PDF.")

    contenido = await file.read()
    if len(contenido) > 20 * 1024 * 1024:  # 20 MB max
        raise HTTPException(413, "Archivo demasiado grande (máx. 20 MB).")

    try:
        resultado = parsear_pdf(contenido)
    except Exception as e:
        raise HTTPException(422, f"Error al parsear el PDF: {e}")

    if not resultado["items"]:
        raise HTTPException(422, "No se encontraron ítems en el PDF. Verifica que sea una factura Starlink válida.")

    return resultado


# ── GET /starlink/periodos ────────────────────────────────────────────────────

@router.get("/periodos")
def listar_periodos(
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Devuelve la lista de períodos con datos guardados, ordenados desc."""
    rows = db.query(StarlinkFactura.periodo).order_by(StarlinkFactura.periodo.desc()).all()
    return [r.periodo for r in rows]


# ── GET /starlink/factura/{periodo} ──────────────────────────────────────────

@router.get("/factura/{periodo}")
def obtener_factura(
    periodo: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Devuelve los datos guardados de un período específico."""
    fac = db.query(StarlinkFactura).filter(StarlinkFactura.periodo == periodo).first()
    if not fac:
        raise HTTPException(404, f"No hay datos para el período {periodo}.")

    lineas_rows = (
        db.query(StarlinkFacturaLinea)
        .filter(StarlinkFacturaLinea.factura_id == fac.id)
        .all()
    )
    pids = {l.proyecto_id for l in lineas_rows if l.proyecto_id is not None}
    nombres = {}
    tipos = {}
    codigos = {}
    if pids:
        proys = db.query(Proyecto.id, Proyecto.nombre_comercial, Proyecto.tipo_proyecto, Proyecto.codigo_tsf).filter(Proyecto.id.in_(pids)).all()
        nombres = {pid: nombre for pid, nombre, _tipo, _codigo in proys}
        tipos = {pid: _tipo for pid, _nombre, _tipo, _codigo in proys}
        codigos = {pid: _codigo for pid, _nombre, _tipo, _codigo in proys}
    lineas = [
        {
            "descripcion":      l.descripcion,
            "proyecto_id":      l.proyecto_id,
            "nombre_comercial": nombres.get(l.proyecto_id),
            "codigo_tsf":       codigos.get(l.proyecto_id),
            "tipo_proyecto":    tipos.get(l.proyecto_id),
            "sin_iva":          float(l.sin_iva),
            "iva":              float(l.iva),
            "monto_total":      float(l.monto_total),
        }
        for l in lineas_rows
    ]

    return {
        "periodo":        fac.periodo,
        "items":          json.loads(fac.items_json),
        "agrupado":       json.loads(fac.agrupado_json),
        "cargos_totales": float(fac.cargos_totales) if fac.cargos_totales else None,
        "suma_items":     float(fac.suma_items),
        "updated_at":     fac.updated_at.isoformat() if fac.updated_at else None,
        "lineas":         lineas,
    }


# ── PUT /starlink/factura/{periodo} ──────────────────────────────────────────

@router.put("/factura/{periodo}")
def guardar_factura(
    periodo: str,
    payload: dict,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Crea o sobreescribe los datos de un período."""
    import re
    if not re.match(r"^\d{4}-\d{2}$", periodo):
        raise HTTPException(400, "Período debe tener formato YYYY-MM.")

    items    = payload.get("items", [])
    agrupado = payload.get("agrupado", [])
    if not items:
        raise HTTPException(400, "Sin ítems para guardar.")

    fac = db.query(StarlinkFactura).filter(StarlinkFactura.periodo == periodo).first()
    if fac:
        fac.items_json     = json.dumps(items,    ensure_ascii=False)
        fac.agrupado_json  = json.dumps(agrupado, ensure_ascii=False)
        fac.cargos_totales = payload.get("cargos_totales")
        fac.suma_items     = payload.get("suma_items", 0)
    else:
        fac = StarlinkFactura(
            periodo        = periodo,
            items_json     = json.dumps(items,    ensure_ascii=False),
            agrupado_json  = json.dumps(agrupado, ensure_ascii=False),
            cargos_totales = payload.get("cargos_totales"),
            suma_items     = payload.get("suma_items", 0),
        )
        db.add(fac)

    db.flush()          # asegura fac.id antes de generar las líneas
    _regenerar_lineas(db, fac)
    db.commit()
    db.refresh(fac)
    return {"ok": True, "periodo": fac.periodo}


# ── DELETE /starlink/factura/{periodo} ────────────────────────────────────────

@router.delete("/factura/{periodo}")
def eliminar_factura(
    periodo: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Elimina los datos de un período."""
    fac = db.query(StarlinkFactura).filter(StarlinkFactura.periodo == periodo).first()
    if not fac:
        raise HTTPException(404, f"No hay datos para el período {periodo}.")
    db.delete(fac)
    db.commit()
    return {"ok": True}


# ── GET /starlink/mapeo ───────────────────────────────────────────────────────

@router.get("/mapeo")
def listar_mapeo(db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Catálogo de mapeos sitio→proyecto, con el nombre comercial resuelto."""
    filas = db.query(StarlinkMapeoSitio).order_by(StarlinkMapeoSitio.patron).all()
    pids = {m.proyecto_id for m in filas if m.proyecto_id is not None}
    nombres = {}
    if pids:
        nombres = {
            pid: nombre
            for pid, nombre in db.query(Proyecto.id, Proyecto.nombre_comercial)
                                 .filter(Proyecto.id.in_(pids)).all()
        }
    return [
        {
            "id": m.id, "patron": m.patron, "proyecto_id": m.proyecto_id,
            "nombre_comercial": nombres.get(m.proyecto_id), "activo": m.activo,
        }
        for m in filas
    ]


# ── PUT /starlink/mapeo ───────────────────────────────────────────────────────

@router.put("/mapeo")
def upsert_mapeo(payload: dict, db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Crea o actualiza un mapeo sitio→proyecto (clave: patron) y reprocesa TODAS
    las facturas guardadas para reflejar el cambio en las líneas."""
    patron = (payload.get("patron") or "").strip()
    if not patron:
        raise HTTPException(400, "patron es obligatorio.")
    proyecto_id = payload.get("proyecto_id")

    m = db.query(StarlinkMapeoSitio).filter(StarlinkMapeoSitio.patron == patron).first()
    if m:
        m.proyecto_id = proyecto_id
        m.activo = payload.get("activo", True)
    else:
        m = StarlinkMapeoSitio(patron=patron, proyecto_id=proyecto_id,
                               activo=payload.get("activo", True))
        db.add(m)
    db.flush()

    for fac in db.query(StarlinkFactura).all():
        _regenerar_lineas(db, fac)
    db.commit()
    db.refresh(m)
    return {"ok": True, "id": m.id, "patron": m.patron}


# ── POST /starlink/excel ──────────────────────────────────────────────────────

@router.post("/excel")
async def generar_excel(
    payload: dict,
    _=Depends(get_current_user),
):
    """Recibe los datos procesados y devuelve un archivo .xlsx."""
    items    = payload.get("items", [])
    agrupado = payload.get("agrupado", [])
    if not items:
        raise HTTPException(400, "Sin datos para generar Excel.")

    wb = _construir_excel(items, agrupado)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=starlink_factura.xlsx"},
    )


# ── Excel builder ─────────────────────────────────────────────────────────────

def _construir_excel(items: list[dict], agrupado: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Colores
    HEADER_COLOR = "1F4E79"   # azul oscuro
    ROW_ALT      = "BDD7EE"   # azul claro
    TOTAL_COLOR  = "0D2137"   # azul muy oscuro
    WHITE        = "FFFFFF"
    HEADER_FONT  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    TOTAL_FONT   = Font(bold=True, color=WHITE, name="Calibri", size=10)
    NORMAL_FONT  = Font(name="Calibri", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left", vertical="center")

    thin_side = Side(style="thin", color="B8CCE4")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def money_fmt(ws, row, cols):
        for c in cols:
            ws.cell(row, c).number_format = '#,##0.00'

    wb = Workbook()

    # ── Hoja 1: Detalle ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Detalle"

    headers1 = ["Tipo", "Descripción", "Precio unitario", "Cantidad",
                "Total impuestos", "Sin IVA", "IVA", "Monto total"]
    money_cols1 = [3, 5, 6, 7, 8]   # columnas con formato dinero

    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(1, ci, h)
        cell.fill   = fill(HEADER_COLOR)
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border

    totals1 = {c: 0.0 for c in money_cols1}
    for ri, item in enumerate(items, 2):
        vals = [
            item["tipo"], item["descripcion"], item["precio_unitario"],
            item["cantidad"], item["total_impuestos"],
            item["sin_iva"], item["iva"], item["monto_total"],
        ]
        row_fill = fill(ROW_ALT) if ri % 2 == 0 else fill(WHITE)
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(ri, ci, v)
            cell.fill   = row_fill
            cell.font   = NORMAL_FONT
            cell.border = thin_border
            cell.alignment = CENTER if ci != 2 else LEFT
        money_fmt(ws1, ri, money_cols1)
        for c in money_cols1:
            totals1[c] += ws1.cell(ri, c).value or 0

    # Fila TOTAL hoja 1
    tr = len(items) + 2
    ws1.cell(tr, 1, "TOTAL").fill = fill(TOTAL_COLOR)
    ws1.cell(tr, 1).font = TOTAL_FONT
    ws1.cell(tr, 1).alignment = CENTER
    ws1.cell(tr, 2, "").fill = fill(TOTAL_COLOR)
    ws1.cell(tr, 3, "").fill = fill(TOTAL_COLOR)
    ws1.cell(tr, 4, "").fill = fill(TOTAL_COLOR)
    for c in money_cols1:
        cell = ws1.cell(tr, c, round(totals1[c], 2))
        cell.fill      = fill(TOTAL_COLOR)
        cell.font      = TOTAL_FONT
        cell.alignment = CENTER
        cell.number_format = '#,##0.00'
        cell.border    = thin_border
    for ci in range(1, len(headers1) + 1):
        ws1.cell(tr, ci).border = thin_border

    # Anchos automáticos hoja 1
    _auto_width(ws1, headers1)

    # ── Hoja 2: Agrupado ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Agrupado")
    headers2 = ["Descripción", "Cantidad total", "Precio unit. promedio",
                "Sin IVA", "IVA", "Monto total"]
    money_cols2 = [3, 4, 5, 6]

    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(1, ci, h)
        cell.fill   = fill(HEADER_COLOR)
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border

    totals2 = {c: 0.0 for c in money_cols2}
    for ri, item in enumerate(agrupado, 2):
        vals = [
            item["descripcion"], item["cantidad_total"],
            item["precio_unitario_promedio"],
            item["sin_iva"], item["iva"], item["monto_total"],
        ]
        row_fill = fill(ROW_ALT) if ri % 2 == 0 else fill(WHITE)
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(ri, ci, v)
            cell.fill   = row_fill
            cell.font   = NORMAL_FONT
            cell.border = thin_border
            cell.alignment = CENTER if ci != 1 else LEFT
        money_fmt(ws2, ri, money_cols2)
        for c in money_cols2:
            totals2[c] += ws2.cell(ri, c).value or 0

    # Fila TOTAL hoja 2
    tr2 = len(agrupado) + 2
    for ci in range(1, len(headers2) + 1):
        cell = ws2.cell(tr2, ci)
        cell.fill   = fill(TOTAL_COLOR)
        cell.font   = TOTAL_FONT
        cell.border = thin_border
        cell.alignment = CENTER
    ws2.cell(tr2, 1, "TOTAL")
    for c in money_cols2:
        cell = ws2.cell(tr2, c, round(totals2[c], 2))
        cell.number_format = '#,##0.00'

    _auto_width(ws2, headers2)

    return wb


def _auto_width(ws, headers):
    from openpyxl.utils import get_column_letter
    for ci, h in enumerate(headers, 1):
        max_len = len(h)
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

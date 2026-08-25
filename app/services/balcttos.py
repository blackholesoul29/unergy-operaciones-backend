"""Parser del BalCttos de XM (Balance de Contratos): extrae el NETO DE COMPRAS EN BOLSA.

El BalCttos (`BalCttos_txN_MM.xlsx`, bajado del FTP de XM, ~3 días de atraso) trae, por
día y concepto, 24 columnas horarias (HORA 01..24). El concepto 'NETO DE COMPRAS EN BOLSA'
es el neto REAL de compras en bolsa (ya cruzado hora a hora por XM, incluyendo los déficits
de los mínimos horarios de NEU/NITRO/Terpel), que es lo que genera la garantía. Somos el
VENDEDOR UNGG. Un neto positivo = compramos de más en bolsa (déficit).

AISLADO: funciones puras separadas de la lectura del xlsx. No toca modelos, tablas ni
endpoints. Salida en MWh.

Estructura de la hoja (columnas 0-based): 0 FechaDocumento, 1 CONCEPTO, 2 MERCADO,
3 CÓDIGO CONTRATO, 4 COMPRADOR, 5 VENDEDOR, 6 TIPO DE DESPACHO, 7 TIPO ASIGNA,
8..31 HORA 01..HORA 24 (energía en kWh).
"""
from __future__ import annotations

import unicodedata

CONCEPTO_NETO_COMPRAS = "neto de compras en bolsa"
_COL_HORA_INI = 8
_COL_HORA_FIN = 32  # exclusivo -> 24 columnas


def _norm(texto) -> str:
    """minúsculas, sin acentos, sin espacios extremos."""
    s = unicodedata.normalize("NFKD", str(texto or ""))
    return "".join(c for c in s if not unicodedata.combining(c)).strip().lower()


def neto_compras_bolsa(filas: list[dict]) -> dict:
    """filas = [{'concepto', 'fecha', 'horas': [24 valores kWh]}] ->
    {'total_mwh': float, 'por_dia': {fecha 'YYYY-MM-DD': mwh}}.

    Filtra concepto == 'NETO DE COMPRAS EN BOLSA' y suma las 24 horas (kWh -> MWh).
    """
    por_dia: dict[str, float] = {}
    for f in filas:
        if _norm(f.get("concepto")) != CONCEPTO_NETO_COMPRAS:
            continue
        kwh = sum(h for h in (f.get("horas") or []) if isinstance(h, (int, float)))
        dia = str(f.get("fecha") or "")[:10]
        por_dia[dia] = por_dia.get(dia, 0.0) + kwh / 1000.0
    return {"total_mwh": sum(por_dia.values()), "por_dia": por_dia}


def proyectar_neto_mwh(neto_mwh: float, dias_con_dato: int, dias_objetivo: int) -> float:
    """Proyecta el neto de compras en bolsa a `dias_objetivo` usando la tasa diaria REAL
    observada en el BalCttos (`neto_mwh` / `dias_con_dato`).

    Es el ancla de la proyección: en vez de proyectar generación bruta (que da el signo y
    la magnitud mal), se extrapola la compra neta real por día. `dias_objetivo` = días de la
    ventana a estimar (resto del mes = días que faltan; mes siguiente = días del mes).
    Devuelve 0 si aún no hay días con dato.
    """
    if dias_con_dato <= 0:
        return 0.0
    tasa_diaria = neto_mwh / dias_con_dato
    return tasa_diaria * dias_objetivo


def _filas_de_worksheet(ws) -> list[dict]:
    """Convierte la hoja del BalCttos en filas {fecha, concepto, horas}. Salta el header."""
    filas = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or r is None:
            continue
        filas.append({
            "fecha": r[0] if len(r) > 0 else None,
            "concepto": r[1] if len(r) > 1 else None,
            "horas": list(r[_COL_HORA_INI:_COL_HORA_FIN]),
        })
    return filas


def neto_compras_bolsa_de_bytes(contenido: bytes) -> dict:
    """Lee el xlsx del BalCttos desde bytes y devuelve el neto de compras en bolsa."""
    import io

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return neto_compras_bolsa(_filas_de_worksheet(ws))


def neto_compras_bolsa_de_archivo(path: str) -> dict:
    """Igual que `_de_bytes` pero desde una ruta de archivo."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return neto_compras_bolsa(_filas_de_worksheet(ws))

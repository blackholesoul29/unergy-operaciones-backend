"""Costo regulatorio del mes desde la hoja 'Facturas XM' del `Cruce facturas M YYYY txf.xlsx`.

AISLADO: no toca modelos, tablas ni endpoints. Dos capas: `extraer_facturas_xm` lee la
hoja (openpyxl); `costo_regulatorio_de_facturas` calcula sobre estructuras simples.

Regla (calibrada contra el archivo real de garantías de XM, 2026-08): el Valor Garantía de
XM = Exposición Energía + Cargo por Confiabilidad + Servicios CND-SIC-FAZNI. Por eso, del
Cruce facturas GENERADOR se suman Cargo por confiabilidad + Servicios (administración SIC,
despacho y coordinación CND) + FAZNI, y se EXCLUYEN:
  - "Energía en bolsa" (es la compra/exposición, va aparte),
  - "Arranque y parada" (XM no lo mete en el Valor Garantía),
  - el IVA (las columnas de XM son cargos base sin IVA),
  - las facturas COMERCIALIZADOR completas,
  - los subtotales ('Valor total', 'Total servicios ...').

Para volver a meter el IVA, sacar `_es_iva` del filtro (una línea).
"""
from __future__ import annotations

import unicodedata

# Conceptos que NO entran a la garantía: "energia en bolsa" es la compra/exposición;
# "arranque y parada" XM no lo incluye en el Valor Garantía.
_CONCEPTOS_EXCLUIDOS = {"energia en bolsa", "arranque y parada"}
# Tipo de factura que se excluye por completo.
_TIPO_EXCLUIDO = "comercializador"

NOMBRE_HOJA = "Facturas XM"


def _norm(texto) -> str:
    """minúsculas, sin acentos, sin espacios extremos."""
    s = unicodedata.normalize("NFKD", str(texto or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


def _es_subtotal(concepto_norm: str) -> bool:
    """Las filas 'Valor total' y 'Total servicios ...' son subtotales, no conceptos."""
    return concepto_norm.startswith("valor total") or concepto_norm.startswith("total ")


def _es_iva(concepto_norm: str) -> bool:
    """La línea de IVA (p. ej. '+ i.v.a. (19%)'); XM no la mete en el Valor Garantía."""
    return "i.v.a" in concepto_norm or concepto_norm.startswith("iva")


def costo_regulatorio_de_facturas(facturas: list[dict]) -> float:
    """facturas = [{'asic','tipo','lineas':[(concepto, monto), ...]}] -> total regulatorio."""
    total = 0.0
    for f in facturas:
        if _norm(f.get("tipo")) == _TIPO_EXCLUIDO:
            continue
        for concepto, monto in f.get("lineas", []):
            cn = _norm(concepto)
            if _es_subtotal(cn) or cn in _CONCEPTOS_EXCLUIDOS or _es_iva(cn):
                continue
            try:
                total += float(monto)
            except (TypeError, ValueError):
                continue
    return total


def extraer_facturas_xm(ws) -> list[dict]:
    """Lee una worksheet 'Facturas XM' -> [{'asic','tipo','lineas':[(concepto, monto)]}].

    Detecta cada factura por el encabezado 'Factura ASICxxxx - TIPO' (col A). Las filas
    siguientes con concepto en A y monto numérico en E (columna 'total') son líneas;
    ignora la fila header 'campo'.
    """
    facturas: list[dict] = []
    actual: dict | None = None
    for fila in ws.iter_rows(min_row=1, values_only=True):
        a = fila[0] if len(fila) > 0 else None
        total = fila[4] if len(fila) > 4 else None
        if a is None:
            continue
        texto = str(a).strip()
        if texto.lower().startswith("factura "):
            # 'Factura ASIC125059 - COMERCIALIZADOR'
            resto = texto[len("factura "):].strip()
            asic, _, tipo = resto.partition("-")
            actual = {"asic": asic.strip(), "tipo": tipo.strip(), "lineas": []}
            facturas.append(actual)
            continue
        if actual is None or texto.lower() == "campo":
            continue
        if isinstance(total, (int, float)):
            actual["lineas"].append((texto, float(total)))
    return facturas


def _costo_de_workbook(wb) -> float:
    ws = wb[NOMBRE_HOJA] if NOMBRE_HOJA in wb.sheetnames else wb[wb.sheetnames[0]]
    return costo_regulatorio_de_facturas(extraer_facturas_xm(ws))


def costo_regulatorio_de_archivo(path: str) -> float:
    """Abre el xlsx de una ruta y devuelve su costo regulatorio."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return _costo_de_workbook(wb)


def costo_regulatorio_de_bytes(contenido: bytes) -> float:
    """Igual que `_de_archivo` pero desde bytes (p. ej. un archivo bajado de Drive)."""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    return _costo_de_workbook(wb)

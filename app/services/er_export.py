"""Genera el Estado de Resultados en Excel: es lo que recibe el inversionista.

Se escribe pensando en quien lo abre, no en quien lo produce:

* **Todo va con fórmulas**, no con valores pegados. El inversionista puede
  auditar de dónde sale cada número y cambiar un supuesto para ver el efecto.
  Ojo: en un `.xlsx` las funciones se guardan SIEMPRE en inglés y Excel las
  muestra en el idioma del usuario -- escribir ``SUMA`` daría ``#¿NOMBRE?``. Se
  escribe ``SUM`` y un Excel en español lo lee como ``SUMA``.
* Arriba el resumen -- energía, ingreso, costos, utilidad y tarifa -- que es lo
  que se mira primero; el detalle debajo y el día a día al final.
* Sin cuadrícula, con los colores de la marca y los números en formato de moneda
  colombiana.

El archivo que genera la API de Liquidaciones trae las fórmulas sin evaluar y por
eso hace falta LibreOffice para leerlo; este trae fórmulas vivas y valores que
Excel calcula al abrir.
"""
from __future__ import annotations

import io
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Colores de marca.
MORADO = "2C2039"
LILA = "915BD8"
LIMA = "F6FF72"
CREMA = "FDFAF7"
GRIS = "6B6280"
BORDE = "E4DCF2"

MONEDA = '"$"#,##0;[Red]-"$"#,##0'
ENERGIA = '#,##0.00" kWh"'
TARIFA = '"$"#,##0.00'
PORCENTAJE = "0.0%"

# El logo se embebe si está disponible; sin él el encabezado sigue funcionando.
RUTA_LOGO = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                         "static", "logo_unergy.png")

_fina = Side(style="thin", color=BORDE)
_gruesa = Side(style="medium", color=MORADO)


def _titulo_seccion(ws, fila: int, texto: str, ancho: int = 4) -> int:
    """Franja de sección. Devuelve la fila siguiente."""
    for col in range(2, 2 + ancho):
        c = ws.cell(fila, col)
        c.fill = PatternFill("solid", fgColor=MORADO)
        c.border = Border(top=_fina, bottom=_fina)
    c = ws.cell(fila, 2, texto)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[fila].height = 20
    return fila + 1


def _encabezado(ws, nombre_proyecto: str, periodo: str, participe: str | None) -> int:
    """Banda superior con el logo y de quién es el documento."""
    for fila in range(1, 6):
        for col in range(1, 9):
            ws.cell(fila, col).fill = PatternFill("solid", fgColor=MORADO)

    if os.path.exists(RUTA_LOGO):
        try:
            img = XLImage(RUTA_LOGO)
            img.height, img.width = 34, 110
            ws.add_image(img, "B2")
        except Exception:
            # Un logo ilegible no puede impedir entregar el estado de resultados.
            pass

    c = ws.cell(2, 4, "ESTADO DE RESULTADOS")
    c.font = Font(bold=True, size=15, color="FFFFFF")
    c = ws.cell(3, 4, nombre_proyecto)
    c.font = Font(size=12, color=LIMA)
    c = ws.cell(4, 4, f"Período {periodo}" + (f"  ·  {participe}" if participe else ""))
    c.font = Font(size=9.5, color="D9D2E6")
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[5].height = 8
    return 7


def _resumen(ws, fila: int, ref: dict[str, str], participe: str | None,
             pct: float, energia_ref: str) -> int:
    """Lo primero que se mira: energía, ingresos, costos, utilidad y tarifa.

    Todo apunta con fórmulas a las celdas del detalle, así que cuadra por
    construcción y se puede auditar hacia abajo.
    """
    fila = _titulo_seccion(ws, fila, "RESUMEN DEL PERÍODO")
    inicio = fila

    filas = [
        ("Energía generada", energia_ref, ENERGIA, False),
        ("Ingresos brutos", ref["ingresos"], MONEDA, False),
        ("Costos de comercialización (XM)", f"-{ref['comercializacion']}", MONEDA, False),
        ("Costos operativos", f"-{ref['costos']}", MONEDA, False),
    ]
    for etiqueta, formula, fmt, _ in filas:
        ws.cell(fila, 2, etiqueta).font = Font(size=10)
        c = ws.cell(fila, 4, f"={formula}")
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
        c.border = Border(bottom=_fina)
        fila += 1

    # Utilidad = ingresos - comercialización - costos operativos.
    ws.cell(fila, 2, "UTILIDAD DEL PERÍODO").font = Font(bold=True, size=11, color=MORADO)
    util = ws.cell(fila, 4, f"=SUM(D{inicio + 1}:D{fila - 1})")
    util.font = Font(bold=True, size=11, color=MORADO)
    util.number_format = MONEDA
    util.alignment = Alignment(horizontal="right")
    util.fill = PatternFill("solid", fgColor=LIMA)
    ref["utilidad"] = f"D{fila}"
    fila += 1

    # Tarifa neta: lo que queda por cada kWh generado.
    ws.cell(fila, 2, "Tarifa neta").font = Font(size=10, italic=True)
    c = ws.cell(fila, 4, f'=IF({energia_ref}=0,"",{ref["utilidad"]}/{energia_ref})')
    c.number_format = TARIFA
    c.alignment = Alignment(horizontal="right")
    c.font = Font(italic=True)
    fila += 1

    if participe:
        ws.cell(fila, 2, "Participación").font = Font(size=10, italic=True)
        c = ws.cell(fila, 4, pct / 100)
        c.number_format = PORCENTAJE
        c.alignment = Alignment(horizontal="right")
        c.font = Font(italic=True)
        fila += 1
    return fila + 1


def _bloque(ws, fila: int, titulo: str, lineas, ref: dict[str, str],
            clave: str, etiqueta_total: str | None = None) -> int:
    """Un bloque de conceptos con su total por fórmula."""
    fila = _titulo_seccion(ws, fila, titulo)
    inicio = fila

    por_concepto: dict[str, float] = {}
    for l in lineas:
        por_concepto[l.concepto] = por_concepto.get(l.concepto, 0.0) + float(l.valor_cop or 0)

    if not por_concepto:
        ws.cell(fila, 2, "Sin movimientos en el período").font = Font(size=9.5, italic=True,
                                                                     color=GRIS)
        ref[clave] = "0"
        return fila + 2

    for concepto, valor in por_concepto.items():
        ws.cell(fila, 2, concepto).font = Font(size=10)
        c = ws.cell(fila, 4, round(abs(valor), 2))
        c.number_format = MONEDA
        c.alignment = Alignment(horizontal="right")
        c.border = Border(bottom=_fina)
        fila += 1

    ws.cell(fila, 2, etiqueta_total or f"Total {titulo.capitalize()}").font = Font(bold=True, size=10)
    total = ws.cell(fila, 4, f"=SUM(D{inicio}:D{fila - 1})")
    total.font = Font(bold=True, size=10)
    total.number_format = MONEDA
    total.alignment = Alignment(horizontal="right")
    total.border = Border(top=_fina, bottom=_gruesa)
    ref[clave] = f"D{fila}"
    return fila + 2


def _tabla_diaria(ws, fila: int, diario: list[dict[str, Any]],
                  comercializador: str | None) -> tuple[int, str]:
    """El día a día. Devuelve la fila libre y la celda del total de energía."""
    com = comercializador or "Comercializador"
    fila = _titulo_seccion(ws, fila, "DETALLE DIARIO", ancho=6)

    encabezados = ["Fecha", "Generación (kWh)", "Importación (kWh)",
                   f"Venta {com} (kWh)", f"Venta {com} ($)"]
    for i, texto in enumerate(encabezados):
        c = ws.cell(fila, 2 + i, texto)
        c.font = Font(bold=True, size=9, color=MORADO)
        c.fill = PatternFill("solid", fgColor=CREMA)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = Border(bottom=_fina)
    ws.row_dimensions[fila].height = 28
    fila += 1
    inicio = fila

    for d in diario:
        ws.cell(fila, 2, d["fecha"]).font = Font(size=9)
        for i, (clave, fmt) in enumerate([
            ("generacion_kwh", '#,##0.00'), ("importacion_kwh", '#,##0.00'),
            ("venta_kwh", '#,##0.00'), ("venta_cop", MONEDA),
        ]):
            c = ws.cell(fila, 3 + i, d.get(clave, 0.0))
            c.number_format = fmt
            c.font = Font(size=9)
        fila += 1

    ws.cell(fila, 2, "TOTAL").font = Font(bold=True, size=9.5)
    for i, fmt in enumerate(['#,##0.00', '#,##0.00', '#,##0.00', MONEDA]):
        col = get_column_letter(3 + i)
        c = ws.cell(fila, 3 + i,
                    f"=SUM({col}{inicio}:{col}{fila - 1})" if diario else 0)
        c.font = Font(bold=True, size=9.5)
        c.number_format = fmt
        c.border = Border(top=_gruesa)
    return fila + 2, f"C{fila}"


def generar_er_xlsx(panel, nombre_proyecto: str,
                    diario: list[dict[str, Any]] | None = None,
                    inversionista: str | None = None) -> bytes:
    """El ``.xlsx`` del período. Con ``inversionista``, solo la parte de esa persona."""
    lineas = list(panel.lineas)
    if inversionista:
        lineas = [l for l in lineas if l.inversionista_nombre == inversionista]
    diario = diario or []
    pct = next((float(l.porcentaje or 0) for l in lineas if l.porcentaje), 100.0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de resultados"
    ws.sheet_view.showGridLines = False

    fila = _encabezado(ws, nombre_proyecto, panel.periodo, inversionista)

    # El detalle se arma primero porque el resumen apunta a sus totales.
    ws_fila_resumen = fila
    fila += 9 if inversionista else 8   # espacio reservado para el resumen

    ref: dict[str, str] = {}
    fila = _bloque(ws, fila, "INGRESOS",
                   [l for l in lineas if l.grupo == "ingresos"], ref, "ingresos",
                   "Total ingresos brutos")
    fila = _bloque(ws, fila, "COMERCIALIZACIÓN (XM)",
                   [l for l in lineas if l.grupo == "comercializacion"], ref,
                   "comercializacion", "Total costos de comercialización")
    fila = _bloque(ws, fila, "COSTOS OPERATIVOS",
                   [l for l in lineas if l.grupo in ("costos", "facturas")], ref, "costos",
                   "Total costos operativos")
    fila, energia_ref = _tabla_diaria(ws, fila, diario,
                                      getattr(panel, "comercializador", None))

    _resumen(ws, ws_fila_resumen, ref, inversionista, pct, energia_ref)

    ws.cell(fila, 2, "Generado por Unergy · los valores son fórmulas y se recalculan "
                     "al abrir el archivo").font = Font(size=8, italic=True, color=GRIS)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 19
    ws.freeze_panes = "A7"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

"""
Servicio de ingesta de datos del MEM (XM).

Parsea archivos XLSX/CSV de generación ASIC y de precios de bolsa, y los
persiste en las tablas `mem_datos_asic` / `mem_precios_bolsa`. El cruce de la
generación con cada proyecto se hace por `Proyecto.codigo_asic`.

El parseo (`parse_tabular`) es una función pura sin dependencia de la BD para
poder testearlo de forma aislada.
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime

from sqlalchemy.orm import Session

from app.models.proyectos import Proyecto
from app.models.mem import MEMDatosASIC, MEMPrecioBolsa, MEMGesconEstado

logger = logging.getLogger(__name__)


def _norm_header(h: object) -> str:
    """Normaliza un encabezado: minúsculas, sin acentos ni espacios extra."""
    s = str(h or "").strip().lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")):
        s = s.replace(a, b)
    return s.replace(" ", "_").replace("-", "_")


# Sinónimos aceptados por columna lógica → nombre canónico.
_ALIASES = {
    "codigo_asic": {"codigo_asic", "codigo", "asic", "codigo_sic", "frt", "frontera"},
    "fecha": {"fecha", "date", "dia"},
    "hora": {"hora", "hour", "h"},
    "generacion_kwh": {"generacion_kwh", "generacion", "kwh", "energia_kwh", "energia"},
    "fuente": {"fuente", "source", "origen"},
    "precio_cop_kwh": {"precio_cop_kwh", "precio", "precio_bolsa", "precio_kwh", "mpo"},
}


def _canonical(header: str) -> str | None:
    for canon, names in _ALIASES.items():
        if header in names:
            return canon
    return None


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # ISO con tiempo
    return datetime.fromisoformat(s).date()


def _parse_float(value: object) -> float:
    """
    Parsea un número tolerando formato colombiano/XM (coma = separador DECIMAL,
    punto = separador de miles) y formato US (punto decimal, coma de miles).

    Los archivos de XM/ASIC vienen en formato colombiano: `230,5` = 230.5 kWh y
    `1.234,75` = 1234.75. Tratar la coma como separador de miles (el bug original)
    inflaba/deformaba la magnitud — la misma clase de error de unidad que ya costó
    dos incidentes (kWh↔MWh 1000×). El separador decimal es SIEMPRE el que aparece
    más a la derecha cuando ambos símbolos están presentes.
    """
    if isinstance(value, bool):
        raise ValueError("valor booleano no es numérico")
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        raise ValueError("valor numérico vacío")
    has_dot, has_comma = "." in s, "," in s
    if has_dot and has_comma:
        if s.rfind(",") > s.rfind("."):
            # 1.234,75 → coma decimal: el punto es de miles.
            s = s.replace(".", "").replace(",", ".")
        else:
            # 1,234.75 → punto decimal: la coma es de miles.
            s = s.replace(",", "")
    elif has_comma:
        # Solo coma. Una sola coma = separador decimal colombiano (230,5).
        # Varias comas = separador de miles estilo US (1,234,567).
        s = s.replace(",", ".") if s.count(",") == 1 else s.replace(",", "")
    return float(s)


def parse_tabular(file_content: bytes, filename: str | None = None) -> list[dict]:
    """
    Parsea un XLSX o CSV en una lista de filas (dicts con encabezados canónicos).

    Detecta el formato por la firma del archivo (ZIP → XLSX) o por la extensión.
    No toca la base de datos.
    """
    is_xlsx = (file_content[:2] == b"PK") or (filename or "").lower().endswith((".xlsx", ".xlsm"))

    if is_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_rows = [list(r) for r in rows_iter]
    else:
        text = file_content.decode("utf-8-sig", errors="replace")
        # Detecta el delimitador (coma o punto y coma).
        sample = text[:2048]
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        raw_rows = [r for r in csv.reader(io.StringIO(text), delimiter=delimiter)]

    if not raw_rows:
        return []

    header_row = raw_rows[0]
    col_map: dict[int, str] = {}
    for idx, h in enumerate(header_row):
        canon = _canonical(_norm_header(h))
        if canon:
            col_map[idx] = canon

    parsed: list[dict] = []
    for raw in raw_rows[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row: dict = {}
        for idx, canon in col_map.items():
            if idx < len(raw):
                row[canon] = raw[idx]
        if row:
            parsed.append(row)
    return parsed


class MEMIngestionService:
    def __init__(self, db: Session):
        self.db = db

    # ── ASIC ────────────────────────────────────────────────────────────────
    def ingest_asic_data(self, file_content: bytes, filename: str | None = None) -> dict:
        """Parsea generación horaria del ASIC y la persiste por proyecto."""
        rows = parse_tabular(file_content, filename)
        created = updated = failed = 0
        errors: list[str] = []

        # Cache de codigo_asic → proyecto_id para no consultar por fila.
        cache: dict[str, int | None] = {}

        for i, row in enumerate(rows, start=2):  # +2: fila 1 = encabezado
            try:
                codigo = str(row.get("codigo_asic", "")).strip()
                if not codigo:
                    raise ValueError("código ASIC vacío")
                if codigo not in cache:
                    p = self.db.query(Proyecto.id).filter(Proyecto.codigo_asic == codigo).first()
                    cache[codigo] = p[0] if p else None
                proyecto_id = cache[codigo]
                if proyecto_id is None:
                    raise ValueError(f"código ASIC desconocido '{codigo}'")

                fecha = _parse_date(row["fecha"])
                hora = int(_parse_float(row["hora"]))
                if not (0 <= hora <= 23):
                    raise ValueError(f"hora fuera de rango (0-23): {hora}")
                generacion = _parse_float(row["generacion_kwh"])
                fuente = (str(row["fuente"]).strip() if row.get("fuente") else "ASIC")

                existing = (
                    self.db.query(MEMDatosASIC)
                    .filter(
                        MEMDatosASIC.proyecto_id == proyecto_id,
                        MEMDatosASIC.fecha == fecha,
                        MEMDatosASIC.hora == hora,
                    )
                    .first()
                )
                if existing:
                    existing.generacion_kwh = generacion
                    existing.fuente = fuente
                    updated += 1
                else:
                    self.db.add(MEMDatosASIC(
                        proyecto_id=proyecto_id, fecha=fecha, hora=hora,
                        generacion_kwh=generacion, fuente=fuente,
                    ))
                    created += 1
            except (KeyError, ValueError, TypeError) as e:
                failed += 1
                errors.append(f"fila {i}: {e}")

        if created or updated:
            self.db.commit()

        return {
            "records_created": created,
            "records_updated": updated,
            "records_failed": failed,
            "rows_processed": len(rows),
            "errors": errors[:100],
        }

    # ── Precios de bolsa ──────────────────────────────────────────────────────
    def ingest_precio_bolsa(self, file_content: bytes, filename: str | None = None) -> dict:
        """Parsea precios horarios de bolsa y los persiste."""
        rows = parse_tabular(file_content, filename)
        created = updated = failed = 0
        errors: list[str] = []

        for i, row in enumerate(rows, start=2):
            try:
                fecha = _parse_date(row["fecha"])
                hora = int(_parse_float(row["hora"]))
                if not (0 <= hora <= 23):
                    raise ValueError(f"hora fuera de rango (0-23): {hora}")
                precio = _parse_float(row["precio_cop_kwh"])

                existing = (
                    self.db.query(MEMPrecioBolsa)
                    .filter(MEMPrecioBolsa.fecha == fecha, MEMPrecioBolsa.hora == hora)
                    .first()
                )
                if existing:
                    existing.precio_cop_kwh = precio
                    updated += 1
                else:
                    self.db.add(MEMPrecioBolsa(fecha=fecha, hora=hora, precio_cop_kwh=precio))
                    created += 1
            except (KeyError, ValueError, TypeError) as e:
                failed += 1
                errors.append(f"fila {i}: {e}")

        if created or updated:
            self.db.commit()

        return {
            "records_created": created,
            "records_updated": updated,
            "records_failed": failed,
            "rows_processed": len(rows),
            "errors": errors[:100],
        }

    # ── GESCON ────────────────────────────────────────────────────────────────
    def update_gescon_statuses(self, updates: list[dict] | None = None) -> dict:
        """
        Actualiza estados GESCON.

        Placeholder de la futura integración automática con XM/GESCON. Por ahora
        registra actualizaciones manuales: cada item es {proyecto_id, estado,
        observaciones?}.
        """
        updates = updates or []
        created = 0
        errors: list[str] = []
        for u in updates:
            try:
                self.db.add(MEMGesconEstado(
                    proyecto_id=int(u["proyecto_id"]),
                    estado=str(u["estado"]),
                    fecha_actualizacion=datetime.now(),
                    observaciones=u.get("observaciones"),
                ))
                created += 1
            except (KeyError, ValueError, TypeError) as e:
                errors.append(str(e))
        if created:
            self.db.commit()
        return {"records_created": created, "records_failed": len(errors), "errors": errors}

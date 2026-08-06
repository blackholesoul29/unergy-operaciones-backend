"""Generación del reporte CGM (extracción Quoia + Excel formato ASIC).

Puerto del script standalone `enviar_reporte_cgm.py` (repo ReporteCGM), pero
acotado a un conjunto específico de fronteras en vez de recorrer todo el
catálogo de Quoia -- el llamador decide qué frt_codes pedir.
"""
from __future__ import annotations

import calendar
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.services.mgs.gaia_client import GaiaClient
from app.services.reporte_energia import curvas as curvas_energia

HORAS = list(range(24))
HORAS_SOLARES = range(6, 18)  # 6am a 6pm (12 horas) -- ventana fija para Indisponibilidad

_MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def es_ultimo_dia_del_mes(fecha: date) -> bool:
    return fecha.day == calendar.monthrange(fecha.year, fecha.month)[1]


def dias_del_mes(fecha: date) -> list[str]:
    """Del día 1 al día `fecha` (inclusive, mismo mes) -- para el consolidado
    mensual que se adjunta cuando `fecha` es el último día del mes."""
    inicio = fecha.replace(day=1)
    return [(inicio + timedelta(days=i)).isoformat() for i in range((fecha - inicio).days + 1)]


def nombre_mes(fecha: date) -> str:
    return _MESES_ES[fecha.month - 1]


def titulo_hoja_mensual(fecha: date) -> str:
    return f"Consolidado {_MESES_ES[fecha.month - 1].capitalize()} {fecha.year}"

ESTADO_QUOIA = {
    "OK": "Exitoso",
    "WARNING": "Exitoso con novedades",
    "ERROR": "Fallo en validacion",
}
CATEGORIA = {1: "Frontera de generación", 2: "Frontera de generación - Consumo"}

COLUMNAS = (
    ["report date", "border frtcode", "border sic code", "border category", "meter", "state"]
    + [f"hour {h}" for h in HORAS]
    + ["total reported energy"]
)


def resolver_borders(gaia: GaiaClient, frt_codes: set[str]) -> dict[str, dict]:
    """Mapea frt_code (lowercase) -> {id, category, name} usando el listado de
    Quoia (get_all_borders, ya cacheado 1h) -- solo para los frt_codes pedidos."""
    wanted = {c.lower() for c in frt_codes}
    resultado: dict[str, dict] = {}
    for proyecto in gaia.get_all_borders():
        nombre = (proyecto.get("name") or "").strip()
        for key in ("frt_generation", "frt_consumption"):
            frt = proyecto.get(key)
            if not frt:
                continue
            frt_code = (frt.get("frt_code") or "").strip().lower()
            if frt_code in wanted:
                resultado[frt_code] = {
                    "id": frt.get("id"),
                    "category": frt.get("category"),
                    "name": nombre,
                }
    return resultado


def fetch_filas(gaia: GaiaClient, frt_code: str, border_meta: dict | None, fecha_str: str) -> list[dict]:
    """Filas main/backup (24h + total) para una frontera. border_meta viene de
    resolver_borders(); si es None (frt_code no encontrado en Quoia hoy),
    retorna filas en cero con estado "Sin reporte"."""
    nombre = border_meta.get("name", "") if border_meta else ""
    categoria = CATEGORIA.get(border_meta.get("category") if border_meta else None, "Frontera de generación")
    border_id = border_meta.get("id") if border_meta else None

    reporte = gaia.get_border_report_status(border_id, fecha_str) if border_id else None
    if reporte:
        estado = ESTADO_QUOIA.get(str(reporte.get("status", "")).upper(), "Sin reporte")
        main_curva = reporte.get("reported_data_main") or [0.0] * 24
        back_curva = reporte.get("reported_data_backup") or [0.0] * 24
    else:
        estado = "Sin reporte"
        main_curva = [0.0] * 24
        back_curva = [0.0] * 24

    filas = []
    for meter_label, curva in [("main", main_curva), ("backup", back_curva)]:
        fila = {
            "report date": fecha_str,
            "border frtcode": frt_code,
            "border sic code": nombre,
            "border category": categoria,
            "meter": meter_label,
            "state": estado,
        }
        for h in HORAS:
            fila[f"hour {h}"] = round(float(curva[h]), 3) if h < len(curva) else 0.0
        fila["total reported energy"] = round(sum(float(v) for v in curva), 3)
        filas.append(fila)
    return filas


def _tiene_dato(curva) -> bool:
    return curva is not None and curva.notna().any()


def _medidor_con_dato(curva_ppal, curva_resp):
    """Prefiere el medidor principal si tiene dato; el respaldo solo si el
    principal no tiene nada -- mismo criterio que _principal_o_respaldo() en
    clasificador.py, reimplementado acá para no importar un helper privado
    de otro módulo por una función de una línea."""
    return curva_ppal if _tiene_dato(curva_ppal) else curva_resp


def _horas_en_cero(curva) -> int:
    """Horas dentro de la ventana solar (HORAS_SOLARES) con lectura en 0 --
    sin recuperación activa (sería una llamada por hueco por día, hasta 90s
    cada una según curvas.curvas_de_frontera, y este cálculo corre TODOS los
    días para Clientes, no una vez al mes como el consolidado de OR). Una
    hora sin dato que la lectura pasiva no trajo se cuenta igual como en
    cero -- decisión explícita del usuario."""
    if curva is None:
        return len(HORAS_SOLARES)
    ventana = curva.reindex(HORAS_SOLARES).fillna(0)
    return int((ventana == 0).sum())


def columnas_resumen(etiqueta_columna: str) -> list[str]:
    return [
        etiqueta_columna, "Proyecto", "Total Generación (kWh)", "Total Consumo (kWh)",
        "Producción Específica (kWh/kWp)", "Indisponibilidad (%)", "Factor de Planta (%)",
    ]


def _calcular_resumen(
    gaia: GaiaClient,
    proyectos: dict[int, dict],
    filas_por_frt: dict[str, list[dict]],
    dias: list[str],
    etiqueta_columna: str,
    etiqueta_valor: str,
) -> list[dict]:
    """Una fila por proyecto, acumulado de `dias` -- Total Generación sale del
    mismo reported_data_main ya pedido para la Hoja 1 (no se vuelve a
    consultar Quoia para eso), filtrado a estos `dias` puntuales. Total
    Consumo e Indisponibilidad SÍ necesitan una llamada aparte por día -- la
    curva de MEDIDOR (no CGM), principal o respaldo el que tenga dato, sin
    recuperación activa (ver _horas_en_cero).

    Usada tanto para el Resumen Diario (Hoja 2, `dias=[fecha]`, todos los
    días) como para el Resumen Mensual (Hoja 3, `dias=dias_mes`, solo el
    último día del mes).

    `proyectos`: proyecto_id -> {
        "nombre": str,
        "frt_gen": str | None,        -- codigo_frontera de Generación
        "frt_con": str | None,        -- codigo_frontera de Consumo
        "capacidad_dc_kwp": float | None,      -- ProyectoInfoTecnica.capacidad_instalada_kwp
        "capacidad_efectiva_mw": float | None, -- Frontera.capacidad_efectiva_mw (Generación)
        "main_meter_gen": int | None, "backup_meter_gen": int | None,
        "main_meter_con": int | None, "backup_meter_con": int | None,
    }
    """
    mapa_nodo = curvas_energia.construir_mapa_medidor_nodo(gaia)
    dias_set = set(dias)
    n_dias = len(dias)
    horas_solares_total = len(HORAS_SOLARES) * n_dias

    filas_resumen = []
    for datos in proyectos.values():
        frt_gen = datos.get("frt_gen")
        frt_con = datos.get("frt_con")

        total_gen = sum(
            f["total reported energy"]
            for f in filas_por_frt.get(frt_gen, [])
            if f["meter"] == "main" and f["report date"] in dias_set
        ) if frt_gen else 0.0

        main_meter_gen, backup_meter_gen = datos.get("main_meter_gen"), datos.get("backup_meter_gen")
        horas_cero_total = None
        if frt_gen and (main_meter_gen or backup_meter_gen):
            horas_cero_total = 0
            for dia in dias:
                c = curvas_energia.curvas_de_frontera(
                    gaia, mapa_nodo, main_meter_gen, backup_meter_gen, dia, frt_gen, recuperar=False,
                )
                curva = _medidor_con_dato(c["curva_ppal"], c["curva_resp"])
                horas_cero_total += _horas_en_cero(curva)

        # Total Consumo -- medidor (variable iae, 'consumo_ppal'/'consumo_resp',
        # mismo criterio que ya usa clasificador_consumo.py para esta misma
        # frontera), no CGM -- principal si tiene dato, si no respaldo.
        main_meter_con, backup_meter_con = datos.get("main_meter_con"), datos.get("backup_meter_con")
        total_con = 0.0
        if frt_con and (main_meter_con or backup_meter_con):
            for dia in dias:
                c = curvas_energia.curvas_de_frontera(
                    gaia, mapa_nodo, main_meter_con, backup_meter_con, dia, frt_con, recuperar=False,
                )
                curva = _medidor_con_dato(c["consumo_ppal"], c["consumo_resp"])
                if curva is not None:
                    total_con += float(curva.fillna(0).sum())

        capacidad_dc = datos.get("capacidad_dc_kwp")
        capacidad_efectiva_kw = (
            float(datos["capacidad_efectiva_mw"]) * 1000 if datos.get("capacidad_efectiva_mw") else None
        )
        gen_max_teorica = capacidad_efectiva_kw * 24 * n_dias if capacidad_efectiva_kw else None

        filas_resumen.append({
            etiqueta_columna: etiqueta_valor,
            "Proyecto": datos["nombre"],
            "Total Generación (kWh)": round(total_gen, 3),
            "Total Consumo (kWh)": round(total_con, 3),
            "Producción Específica (kWh/kWp)": round(total_gen / capacidad_dc, 3) if capacidad_dc else None,
            "Indisponibilidad (%)": (
                round(horas_cero_total / horas_solares_total * 100, 2)
                if horas_cero_total is not None and horas_solares_total else None
            ),
            "Factor de Planta (%)": round(total_gen / gen_max_teorica * 100, 2) if gen_max_teorica else None,
        })
    return filas_resumen


def calcular_resumen_mensual(
    gaia: GaiaClient, proyectos: dict[int, dict], filas_por_frt: dict[str, list[dict]],
    dias_mes: list[str], mes_str: str,
) -> list[dict]:
    """Hoja 3 -- solo se adjunta el último día del mes, acumulado de todo el mes."""
    return _calcular_resumen(gaia, proyectos, filas_por_frt, dias_mes, "Mes", mes_str)


def calcular_resumen_diario(
    gaia: GaiaClient, proyectos: dict[int, dict], filas_por_frt: dict[str, list[dict]], fecha_str: str,
) -> list[dict]:
    """Hoja 2 -- todos los días, mismas variables que el Resumen Mensual pero
    de un solo día (n_dias=1)."""
    return _calcular_resumen(gaia, proyectos, filas_por_frt, [fecha_str], "Fecha", fecha_str)


def _estilo_encabezado(cell):
    cell.font = Font(bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = Side(style="thin", color="000000")
    cell.border = Border(left=borde, right=borde, top=borde, bottom=borde)


def _estilo_dato(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(size=9)
    borde = Side(style="thin", color="D0D0D0")
    cell.border = Border(left=borde, right=borde, top=borde, bottom=borde)


_ANCHOS_COLUMNA = {
    "report date": 12, "border frtcode": 14, "border sic code": 30,
    "border category": 30, "meter": 8, "state": 22,
    "total reported energy": 18,
}


def _escribir_hoja(ws, filas: list[dict], columnas: list[str] = COLUMNAS) -> None:
    for col_idx, nombre in enumerate(columnas, start=1):
        _estilo_encabezado(ws.cell(row=1, column=col_idx, value=nombre))

    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, col in enumerate(columnas, start=1):
            valor = fila.get(col)
            if isinstance(valor, float):
                valor = round(valor, 3)
            _estilo_dato(ws.cell(row=row_idx, column=col_idx, value=valor))

    for col_idx, nombre in enumerate(columnas, start=1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = _ANCHOS_COLUMNA.get(nombre, 8 if nombre.startswith("hour") else 12)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def generar_excel(filas: list[dict], titulo_hoja: str = "CGM Report") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = titulo_hoja[:31]  # límite de Excel para el nombre de hoja
    _escribir_hoja(ws, filas)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_excel_cliente(
    filas_acumuladas: list[dict], filas_resumen_diario: list[dict],
    filas_resumen_mensual: list[dict] | None = None,
) -> bytes:
    """Excel para destinatarios tipo Cliente -- dos hojas todos los días, una
    tercera solo el último día del mes:

    1. 'Diario acumulado' -- detalle horario acumulado desde el día 1 del
       mes hasta hoy (mismo formato/columnas que generar_excel).
    2. 'Resumen Diario' -- las mismas variables del Resumen Mensual, pero
       calculadas solo para el día del reporte.
    3. 'Resumen Mensual' -- SOLO si `filas_resumen_mensual` viene con datos
       (el llamador decide si es el último día del mes) -- las mismas
       variables acumuladas de todo el mes.
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Diario acumulado"
    _escribir_hoja(ws1, filas_acumuladas)

    ws2 = wb.create_sheet("Resumen Diario")
    _escribir_hoja(ws2, filas_resumen_diario, columnas=columnas_resumen("Fecha"))

    if filas_resumen_mensual:
        ws3 = wb.create_sheet("Resumen Mensual")
        _escribir_hoja(ws3, filas_resumen_mensual, columnas=columnas_resumen("Mes"))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

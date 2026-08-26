"""Genera el Excel en el formato manual que ya diligencia el equipo cada
mañana (hoja "datos") -- puerto de reporte_excel.py (repo Reporte-Energia),
leyendo directamente de las tablas ya guardadas en vez de un DataFrame en
memoria.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.fronteras import Frontera, TipoFronteraEnum
from app.models.reporte_energia import ReporteEnergiaGeneracion, ReporteEnergiaConsumo
from app.services.reporte_energia.utils import curva_respaldo_a_reportar

COLUMNAS = [
    "nombre_proyecto", "Hora", "Consumo_Principal", "Generación_Principal",
    "Consumo_Respaldo", "Generación_Respaldo",
]


def _respaldo_final(rep) -> list[float]:
    """Mismo dato que /enviar realmente manda como 'Backup' -- curva_respaldo_final
    congelada si ya existe (ver actualizar_respaldo_final), si no se calcula
    al vuelo con el mismo criterio (curva_respaldo_a_reportar), igual que
    hace _enviar_a_quoia()."""
    backup = getattr(rep, "curva_respaldo_final", None)
    if backup is None:
        backup, _ = curva_respaldo_a_reportar(rep)
    return backup


def generar_excel_dia(db: Session, fecha: date) -> bytes:
    gen_filas = db.execute(
        select(ReporteEnergiaGeneracion, Frontera)
        .join(Frontera, Frontera.id == ReporteEnergiaGeneracion.frontera_id)
        .where(ReporteEnergiaGeneracion.fecha == fecha)
    ).all()
    con_filas = db.execute(
        select(ReporteEnergiaConsumo, Frontera)
        .join(Frontera, Frontera.id == ReporteEnergiaConsumo.frontera_id)
        .where(ReporteEnergiaConsumo.fecha == fecha)
    ).all()

    con_por_proyecto = {}
    for rep, front in con_filas:
        if front.proyecto_id is not None:
            con_por_proyecto[front.proyecto_id] = rep

    wb = Workbook()
    ws = wb.active
    ws.title = "datos"
    for col_idx, nombre in enumerate(COLUMNAS, start=1):
        ws.cell(row=1, column=col_idx, value=nombre)

    fila_excel = 2
    for rep_gen, front_gen in gen_filas:
        curva_final = rep_gen.curva_final or [None] * 24
        reporte_ya_valido = rep_gen.medidor_usado == "cgm"
        # Mismo "Backup" que /enviar realmente manda a Quoia -- antes esto
        # era una fórmula =C*RAND() independiente, que ignoraba por completo
        # curva_respaldo_final (y por lo tanto el dato real del medidor
        # cuando aplicaba). None (celda en blanco) cuando el reporte ya es
        # válido en Quoia -- no se manda nada, tampoco Backup.
        respaldo_gen = None if reporte_ya_valido else _respaldo_final(rep_gen)

        rep_con = con_por_proyecto.get(front_gen.proyecto_id)
        consumo_ya_valido = bool(rep_con and rep_con.caso == "CGM")
        curva_con = (rep_con.curva_final if rep_con else None) or [None] * 24
        respaldo_con = None if (consumo_ya_valido or rep_con is None) else _respaldo_final(rep_con)

        for hora in range(24):
            valor_gen = None if reporte_ya_valido else curva_final[hora]
            if valor_gen is None and not reporte_ya_valido:
                valor_gen = 0.0
            valor_con = None if consumo_ya_valido else curva_con[hora]
            if valor_con is None and not consumo_ya_valido:
                valor_con = 0.0

            valor_resp_gen = None if respaldo_gen is None else respaldo_gen[hora]
            valor_resp_con = None if respaldo_con is None else respaldo_con[hora]

            ws.append([
                front_gen.nombre_frontera, hora,
                valor_con, valor_gen,
                valor_resp_con, valor_resp_gen,
            ])
            fila_excel += 1

    for col_idx, nombre in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(nombre) + 2, 10)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

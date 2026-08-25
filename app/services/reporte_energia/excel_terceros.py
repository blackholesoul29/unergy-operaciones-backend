"""Parsea el Excel que envía una empresa tercera para fronteras cuyo CGM lo
maneja esa empresa (FRONTERAS_TERCEROS, ver clasificador.py -- ej.
Cedillanos). Formato observado: una fila por (FECHA, ROLE), con columnas
CODIGO SIC | ROLE (Primary/Backup) | ENERGY TYPE | FECHA | HORA00..HORA23.

El CODIGO SIC del archivo se ignora deliberadamente -- puede traer el código
del tercero, no necesariamente el de nuestra frontera; qué frontera_id
recibe los datos lo decide la URL del endpoint, no el contenido del Excel.
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.reporte_energia import ReporteEnergiaGeneracion

_ENERGY_TYPE_OBJETIVO = "ENERGIAEXPORTADAACTIVA"
_ACENTOS = str.maketrans("ÁÉÍÓÚÑ", "AEIOUN")


def _normalizar(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip().upper().translate(_ACENTOS).replace(" ", "")


def parse_excel_terceros(contenido: bytes) -> dict[date, dict]:
    """Retorna {fecha: {"principal": [24 floats|None] | None, "respaldo": [24 floats|None] | None}},
    filtrado a ENERGY TYPE == 'ENERGIA EXPORTADA ACTIVA'. Lanza ValueError con
    un mensaje apto para mostrar al usuario si el archivo no tiene el formato
    esperado."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el Excel: {e}")
    ws = wb[wb.sheetnames[0]]

    filas = ws.iter_rows(values_only=True)
    try:
        header = list(next(filas))
    except StopIteration:
        raise ValueError("El archivo está vacío")
    header_norm = [_normalizar(h) for h in header]

    def _find(*names):
        for i, h in enumerate(header_norm):
            if h in names:
                return i
        return None

    i_role = _find("ROLE")
    i_tipo = _find("ENERGYTYPE")
    i_fecha = _find("FECHA")
    if i_role is None or i_tipo is None or i_fecha is None:
        raise ValueError("El archivo debe tener columnas ROLE, ENERGY TYPE y FECHA")

    horas_idx: dict[int, int] = {}
    for i, h in enumerate(header_norm):
        if h.startswith("HORA") and h[4:].isdigit():
            horas_idx[int(h[4:])] = i
    if len(horas_idx) < 24:
        raise ValueError(f"Encontré {len(horas_idx)} columnas HORA00..HORA23, se esperaban 24")

    resultado: dict[date, dict] = {}
    for r in filas:
        if not r or i_tipo >= len(r) or _normalizar(r[i_tipo]) != _ENERGY_TYPE_OBJETIVO:
            continue

        fv = r[i_fecha] if i_fecha < len(r) else None
        if hasattr(fv, "date"):
            fecha = fv.date()
        elif hasattr(fv, "year"):
            fecha = fv
        else:
            try:
                fecha = date.fromisoformat(str(fv)[:10])
            except (ValueError, TypeError):
                continue

        role = _normalizar(r[i_role]) if i_role < len(r) else ""
        curva = [
            float(r[horas_idx[h]]) if horas_idx[h] < len(r) and r[horas_idx[h]] is not None else None
            for h in range(24)
        ]

        dia = resultado.setdefault(fecha, {"principal": None, "respaldo": None})
        if role == "PRIMARY":
            dia["principal"] = curva
        elif role == "BACKUP":
            dia["respaldo"] = curva

    return resultado


def aplicar_excel_terceros(db: Session, frontera_id: int, contenido: bytes) -> list[date]:
    """Parsea `contenido` y aplica cada día a ReporteEnergiaGeneracion -- misma
    lógica para el upload manual (POST /cargar-excel-terceros) y la lectura
    automática de correo (excel_terceros_email.py). Retorna las fechas
    efectivamente cargadas (puede ser varias si el Excel trae más de un día).
    Lanza ValueError si el archivo no tiene el formato esperado.

    Si llegan las dos filas (Primary y Backup) se reportan ambas tal cual.
    Si solo llega una (falla real 2026-08-25: un día el Excel solo trajo
    Backup), esa se usa como curva_final y curva_respaldo_terceros queda en
    None -- _enviar_a_quoia() ya sabe estimar el respaldo con la fórmula ±1%
    cuando no hay dato real de terceros, así que no hace falta duplicar esa
    lógica acá."""
    por_fecha = parse_excel_terceros(contenido)

    fechas_cargadas: list[date] = []
    for fecha, datos in por_fecha.items():
        principal = datos["principal"]
        respaldo = datos["respaldo"]
        if principal is None and respaldo is None:
            continue  # ni Primary ni Backup para ese día -- nada que reportar

        curva_final = principal if principal is not None else respaldo
        curva_respaldo_terceros = respaldo if principal is not None else None

        rep = db.execute(
            select(ReporteEnergiaGeneracion).where(
                ReporteEnergiaGeneracion.frontera_id == frontera_id,
                ReporteEnergiaGeneracion.fecha == fecha,
            )
        ).scalar_one_or_none()
        if rep is None:
            rep = ReporteEnergiaGeneracion(frontera_id=frontera_id, fecha=fecha, caso=0)
            db.add(rep)

        rep.caso = 0
        rep.medidor_usado = "excel_terceros"
        rep.curva_final = curva_final
        rep.energia_final_kwh = round(sum(v for v in curva_final if v is not None), 4)
        rep.curva_respaldo_terceros = curva_respaldo_terceros
        rep.revisar_manualmente = False
        rep.editado_manualmente = True
        fechas_cargadas.append(fecha)

    return fechas_cargadas

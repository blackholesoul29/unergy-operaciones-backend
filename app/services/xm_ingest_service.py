"""Ingesta de generación XM (SinergoX) desde Excel hacia xm_generation_history.

El archivo de XM no tiene un encabezado 100% estable, así que la detección de
columnas es flexible (se normaliza el nombre del encabezado y se compara contra
varios alias por campo). Las fechas pueden venir como `datetime` (openpyxl ya las
convierte) o como texto en formato DD/MM/YYYY o YYYY-MM-DD.

Sigue la convención del repo: lectura con `openpyxl` (no pandas) y upsert masivo
con `postgresql.insert(...).on_conflict_do_update()` para reprocesar archivos sin
duplicar registros (clave única: proyecto_id, measurement_date, meter_id).
"""
from __future__ import annotations

import re
import unicodedata
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.models.generacion_xm import XMGeneracionHistorico
from app.models.proyectos import Proyecto


def _norm(s: Any) -> str:
    """Normaliza texto: sin tildes, minúsculas, solo alfanumérico con espacios simples."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


# Campo canónico → tokens normalizados que puede tener el encabezado en el Excel.
# El match es por subcadena sobre el encabezado normalizado.
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "proyecto_id":  ("proyecto id", "id proyecto", "project id"),
    "proyecto":     ("proyecto", "planta", "nombre", "frontera comercial", "project"),
    "meter_id":     ("medidor", "meter", "codigo sic", "frontera", "sic"),
    "fecha":        ("fecha", "date", "dia", "periodo"),
    "generacion":   ("generacion mwh", "generacion", "mwh", "energia", "kwh", "generation"),
}

# Factor de conversión a MWh (unidad de almacenamiento de la columna generation_mwh).
# El resto del dominio de generación del repo trabaja en kWh (kwh_real, kwh_p90,
# energia_activa_*_kwh), así que un archivo XM en kWh almacenado como MWh sin
# convertir corrompe los datos por 1000×. Detectamos la unidad del encabezado y la
# normalizamos; si el encabezado no la declara, fallamos fuerte en vez de adivinar.
_UNIT_TO_MWH: dict[str, Decimal] = {
    "gwh": Decimal("1000"),
    "mwh": Decimal("1"),
    "kwh": Decimal("0.001"),
}


def _detect_unit(header: str) -> str | None:
    """Detecta la unidad de energía declarada en el encabezado de la columna.

    Devuelve 'gwh' | 'mwh' | 'kwh', o None si el encabezado no declara unidad
    reconocible (ej. solo 'Generación' / 'Energía'). El orden importa: se prueba
    de mayor a menor para que 'mwh' no quede enmascarada por una coincidencia
    parcial. La comparación es sobre el encabezado normalizado (sin tildes, lower).
    """
    h = _norm(header)
    for unit in ("gwh", "mwh", "kwh"):
        if re.search(rf"\b{unit}\b", h):
            return unit
    return None


class XMIngestionService:
    def __init__(self, db: Session):
        self.db = db

    # ── lectura del Excel ──────────────────────────────────────────────────────
    def _read_rows(self, file: str | BinaryIO) -> tuple[list[dict], dict[str, str]]:
        """Lee la primera hoja y devuelve filas como dicts {campo_canonico: valor}.

        Devuelve (filas, mapeo_columnas). Lanza ValueError si no hay encabezado
        reconocible para fecha o generación (las columnas mínimas indispensables).
        """
        wb = load_workbook(file, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                raise ValueError("El archivo está vacío")

            # campo canónico → índice de columna
            col_map: dict[str, int] = {}
            for idx, raw_header in enumerate(header):
                norm_h = _norm(raw_header)
                if not norm_h:
                    continue
                for field, aliases in _COLUMN_ALIASES.items():
                    if field in col_map:
                        continue
                    if any(alias in norm_h for alias in aliases):
                        col_map[field] = idx
                        break

            missing = [c for c in ("fecha", "generacion") if c not in col_map]
            if missing:
                raise ValueError(
                    "No se encontraron columnas para: "
                    + ", ".join(missing)
                    + f". Encabezados detectados: {[h for h in header if h]}"
                )

            # La columna de generación DEBE declarar su unidad: almacenamos en MWh y
            # un valor en kWh interpretado como MWh es un error de 1000×. Si el
            # encabezado no la declara, fallamos fuerte (no adivinamos).
            gen_header = str(header[col_map["generacion"]])
            if _detect_unit(gen_header) is None:
                raise ValueError(
                    f"No se pudo determinar la unidad de la columna de generación "
                    f"('{gen_header}'). Etiquete el encabezado con kWh, MWh o GWh."
                )

            rows: list[dict] = []
            for raw in rows_iter:
                if raw is None or all(v is None for v in raw):
                    continue
                rows.append({
                    field: (raw[idx] if idx < len(raw) else None)
                    for field, idx in col_map.items()
                })
            return rows, {f: str(header[i]) for f, i in col_map.items()}
        finally:
            wb.close()

    # ── resolución de proyecto ──────────────────────────────────────────────────
    def build_proyecto_lookup(self) -> dict[str, int]:
        """Mapa nombre_normalizado → proyecto_id (comercial, alias y bitácora)."""
        lookup: dict[str, int] = {}
        proyectos = self.db.query(
            Proyecto.id, Proyecto.nombre_comercial,
            Proyecto.alias_monitoreo, Proyecto.nombre_bitacora,
        ).all()
        for p in proyectos:
            for name in (p.nombre_comercial, p.alias_monitoreo, p.nombre_bitacora):
                norm = _norm(name)
                if norm:
                    lookup.setdefault(norm, p.id)
        return lookup

    # ── parsing + validación (puro: no toca la BD) ───────────────────────────────
    @staticmethod
    def _parse_fecha(value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        # Probar formatos comunes; DD/MM/YYYY primero (formato XM colombiano).
        for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S",
                    "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_generacion(value: Any) -> Decimal | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        try:
            # Tolerar separador de miles "1.234,56" → "1234.56" y comas decimales.
            if isinstance(value, str):
                v = value.strip().replace(" ", "")
                if "," in v and "." in v:
                    v = v.replace(".", "").replace(",", ".")
                elif "," in v:
                    v = v.replace(",", ".")
                return Decimal(v)
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    def parse_rows(
        self, raw_rows: list[dict], proyecto_lookup: dict[str, int],
        unit_factor: Decimal = Decimal("1"),
    ) -> tuple[list[dict], list[str]]:
        """Valida y normaliza filas. Devuelve (registros_validos, errores).

        Cada registro válido: {proyecto_id, meter_id, measurement_date, generation_mwh}.
        `unit_factor` convierte el valor crudo a MWh (1 para MWh, 0.001 para kWh,
        1000 para GWh). Idempotente entre llamadas: dedup interno por
        (proyecto_id, fecha, meter_id), conservando la última aparición.
        """
        valid: dict[tuple, dict] = {}
        errors: list[str] = []

        for i, row in enumerate(raw_rows, start=2):  # fila 1 = encabezado
            # Proyecto: id explícito o resolución por nombre.
            proyecto_id: int | None = None
            raw_pid = row.get("proyecto_id")
            if raw_pid not in (None, ""):
                try:
                    proyecto_id = int(raw_pid)
                except (ValueError, TypeError):
                    proyecto_id = None
            if proyecto_id is None:
                norm_name = _norm(row.get("proyecto"))
                proyecto_id = proyecto_lookup.get(norm_name) if norm_name else None
            if proyecto_id is None:
                ref = row.get("proyecto") or row.get("proyecto_id")
                errors.append(f"Fila {i}: proyecto no reconocido ('{ref}')")
                continue

            fecha = self._parse_fecha(row.get("fecha"))
            if fecha is None:
                errors.append(f"Fila {i}: fecha inválida o vacía ('{row.get('fecha')}')")
                continue

            gen = self._parse_generacion(row.get("generacion"))
            if gen is None:
                errors.append(
                    f"Fila {i}: generación inválida o vacía ('{row.get('generacion')}')")
                continue
            if gen < 0:
                errors.append(f"Fila {i}: generación negativa ({gen})")
                continue

            meter_raw = row.get("meter_id")
            meter_id = str(meter_raw).strip() if meter_raw not in (None, "") else ""
            if not meter_id:
                errors.append(f"Fila {i}: meter_id (medidor/frontera) vacío")
                continue

            key = (proyecto_id, fecha, meter_id)
            valid[key] = {
                "proyecto_id": proyecto_id,
                "meter_id": meter_id,
                "measurement_date": fecha,
                "generation_mwh": gen * unit_factor,
            }

        return list(valid.values()), errors

    # ── upsert masivo ─────────────────────────────────────────────────────────
    def _upsert(self, records: list[dict], source_file: str | None) -> int:
        if not records:
            return 0
        payload = [{**r, "source_file": source_file} for r in records]
        stmt = pg_insert(XMGeneracionHistorico).values(payload)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_xm_gen_hist_proj_date_meter",
            set_={
                "generation_mwh": stmt.excluded.generation_mwh,
                "source_file": stmt.excluded.source_file,
            },
        )
        self.db.execute(stmt)
        self.db.commit()
        return len(payload)

    # ── orquestación ────────────────────────────────────────────────────────────
    def ingest(self, file: str | BinaryIO, source_file: str | None = None) -> dict:
        """Ingesta completa: lee, valida, hace upsert y devuelve un resumen."""
        raw_rows, col_map = self._read_rows(file)
        # `_read_rows` ya garantizó que la unidad es reconocible (si no, levantó
        # ValueError); aquí la usamos para normalizar a MWh.
        unit = _detect_unit(col_map["generacion"])
        unit_factor = _UNIT_TO_MWH[unit]
        lookup = self.build_proyecto_lookup()
        records, errors = self.parse_rows(raw_rows, lookup, unit_factor=unit_factor)
        uploaded = self._upsert(records, source_file)

        sample = [
            {
                "proyecto_id": r["proyecto_id"],
                "meter_id": r["meter_id"],
                "measurement_date": r["measurement_date"].isoformat(),
                "generation_mwh": float(r["generation_mwh"]),
            }
            for r in records[:5]
        ]
        return {
            "uploaded_count": uploaded,
            "skipped_count": len(errors),
            "errors": errors,
            "sample_data": sample,
            "columns_detected": col_map,
            "source_unit": unit,  # unidad detectada en el archivo (kwh/mwh/gwh)
        }

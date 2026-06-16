"""
Carga y parseo de archivos "Estado de Resultados" (ER) por proyecto.

Cada ER es un xlsx con fórmulas. El flujo es:
  1. recalcular_er(): LibreOffice headless recalcula las fórmulas y reescribe el
     archivo, para poder leer los valores con openpyxl(data_only=True).
  2. parsear_er(): extrae ingreso bruto, comercialización XM desglosada, costos
     operativos, IVA y fee de administración → dict estructurado.
  3. match_proyecto(): se reutiliza el de liquidaciones_loader.

El parser es tolerante a la posición exacta de las filas: detecta secciones y
columnas por etiqueta, usando las filas indicadas en la especificación solo como
guía. Todos los valores quedan editables en el frontend.
"""
import os
import re
import shutil
import subprocess
import tempfile
import unicodedata

# Reutilizamos los helpers de matching ya probados en liquidaciones.
from app.utils.liquidaciones_loader import normalizar, match_proyecto  # noqa: F401

IVA = 0.19
FEE_ADMIN = 0.0380  # Fee Administración = Total Ingresos × 3.80%

# Nombres de comercializadores conocidos (solo informativo; el parser detecta la
# columna "Venta ($)" sin importar el nombre).
COMERCIALIZADORES = [
    "terpel", "neu", "biac", "sol&cielo", "sol & cielo", "solycielo",
    "enermas", "unergy",
]

# Tokens de razón social que NO aparecen nunca en un nombre de proyecto y que
# contaminan el match (ej. "S.A.S." matcheaba con "M.D.M. Cientifica S.A.S").
_RAZON_SOCIAL = re.compile(r'^(s\.?a\.?s\.?|e\.?s\.?p\.?|s\.?a\.?|ltda\.?|cia\.?)$', re.I)


# ── Nombre de proyecto desde el nombre de archivo ───────────────────────────────

def extraer_proyecto_de_archivo(nombre: str, proyectos_db: list[dict]) -> dict | None:
    """
    Los ER se llaman "Estado resultados {INVERSIONISTA} {PROYECTO} N 2026.xlsx".
    El inversionista tiene longitud variable (1 a 8+ palabras), así que en vez de
    adivinar dónde termina, probamos match_proyecto() con ventanas deslizantes
    desde la derecha (el proyecto siempre va al final) y nos quedamos con el match
    de la ventana más larga que exista en DB. Reusa el match_proyecto compartido
    (con sus aliases de NAOS, Valencia Oriente, etc.).
    """
    base = re.sub(r'^Estado\s+resultados\s+', '', nombre or '', flags=re.I)
    base = re.sub(r'\.xlsx?$', '', base, flags=re.I)
    # Quitar el sufijo " N 2026" (consecutivo + año), conservando números que
    # son parte del nombre del proyecto (ej. "Valencia Oriente 1", "GD NAOS 1").
    base = re.sub(r'\s+\d+\s+20\d{2}$', '', base)

    palabras = [w for w in base.split() if not _RAZON_SOCIAL.match(w)]
    for n in range(min(6, len(palabras)), 0, -1):
        candidato = " ".join(palabras[-n:])
        m = match_proyecto(proyectos_db, candidato)
        if m:
            return m
    return None


# ── LibreOffice: recalcular fórmulas ────────────────────────────────────────────

def recalcular_er(path: str) -> str:
    """
    Convierte el xlsx a uno con las fórmulas ya calculadas usando LibreOffice
    headless. Devuelve la ruta del archivo recalculado (en un tmpdir).

    Si LibreOffice no está disponible (entorno local sin soffice), se devuelve
    el archivo original tal cual; openpyxl data_only leerá los valores cacheados
    que Excel/LibreOffice hayan guardado.
    """
    soffice = _find_soffice()
    if not soffice:
        return path

    outdir = tempfile.mkdtemp(prefix="er_recalc_")
    try:
        # --convert-to xlsx fuerza el recálculo de fórmulas al reescribir.
        subprocess.run(
            [
                soffice, "--headless", "--calc",
                "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                "--outdir", outdir, path,
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=120,
            env={**os.environ, "HOME": outdir},
        )
    except Exception:
        shutil.rmtree(outdir, ignore_errors=True)
        return path

    base = os.path.splitext(os.path.basename(path))[0] + ".xlsx"
    recalc = os.path.join(outdir, base)
    return recalc if os.path.exists(recalc) else path


def _find_soffice() -> str | None:
    for cand in ("soffice", "libreoffice"):
        found = shutil.which(cand)
        if found:
            return found
    # Rutas típicas en contenedor / Windows
    for p in (
        "/usr/bin/soffice", "/usr/bin/libreoffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
    ):
        if os.path.exists(p):
            return p
    return None


# ── Helpers de lectura ──────────────────────────────────────────────────────────

def _num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[$\s]", "", str(v).strip())
    s = s.replace(".", "").replace(",", ".") if s.count(",") == 1 and s.count(".") > 1 else s.replace(",", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s or s in ("-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _txt(v) -> str:
    return str(v).strip() if v is not None else ""


def _norm(s) -> str:
    s = unicodedata.normalize("NFD", str(s or "").lower().strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


# ── Parser principal ─────────────────────────────────────────────────────────────

def parsear_er(path: str) -> dict:
    """
    Extrae del ER (ya recalculado) un dict estructurado:

    {
      "comercializador": str | None,
      "tiene_bolsa": bool,
      "ingreso_bruto": float,
      "total_ingresos": float,
      "comercializacion": [ {concepto, valor}, ... ],   # XM desglosada (negativos)
      "costos": [ {concepto, valor, iva} ],             # operativos
      "facturas": [ {concepto, valor} ],                # representación, CGM, admin
      "kwh": float | None,
      "warnings": [str],
    }
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    sh = wb[wb.sheetnames[0]]
    grid = [[c.value for c in row] for row in sh.iter_rows()]
    wb.close()

    warnings: list[str] = []

    ing = _parse_ingresos(grid, warnings)
    ingreso_bruto = ing["ingreso_bruto"]
    total_ingresos = ing["total_ingresos"]

    comercializacion = _parse_comercializacion(grid)
    costos = _parse_costos(grid)
    kwh = _parse_kwh(grid)

    # Facturas de servicio: Representación y CGM = kWh × 5 c/u (si hay kWh);
    # Administración = Total Ingresos × 3.80%.
    facturas: list[dict] = []
    if kwh:
        facturas.append({"concepto": "Representación", "valor": -round(kwh * 5, 2)})
        facturas.append({"concepto": "CGM", "valor": -round(kwh * 5, 2)})
    facturas.append({"concepto": "Administración", "valor": -round(total_ingresos * FEE_ADMIN, 2)})

    return {
        "comercializador": ing["comercializador"],
        "tiene_bolsa": ing["tiene_bolsa"],
        "ingreso_bruto": round(ingreso_bruto, 2),
        "total_ingresos": round(total_ingresos, 2),
        "venta_bolsa": round(ing["venta_bolsa"], 2),
        "compra_bolsa": round(ing["compra_bolsa"], 2),
        "comercializacion": comercializacion,
        "costos": costos,
        "facturas": facturas,
        "kwh": kwh,
        "warnings": warnings,
    }


def _parse_ingresos(grid: list[list], warnings: list[str]) -> dict:
    """
    Filas 4-34: tabla de generación diaria + venta por comercializador.
    Ingreso bruto = suma de la(s) columna(s) "Venta ($)".
    Si hay bolsa: Venta + Venta bolsa − Compra bolsa.
    Si hay 2 puntos (Terpel 1+2): suma ambas columnas de venta.
    """
    # Localizar fila de encabezado que contenga "venta" en alguna celda.
    header_row = None
    for i, row in enumerate(grid[:40]):
        joined = " ".join(_norm(c) for c in row if c is not None)
        if "venta" in joined and ("$" in joined or "venta ($)" in joined or "venta($)" in joined):
            header_row = i
            break

    venta_cols: list[int] = []
    venta_bolsa_cols: list[int] = []
    compra_bolsa_cols: list[int] = []
    comercializador = None

    if header_row is not None:
        for j, c in enumerate(grid[header_row]):
            h = _norm(c)
            if not h:
                continue
            if "compra" in h and "bolsa" in h:
                compra_bolsa_cols.append(j)
            elif "venta" in h and "bolsa" in h:
                venta_bolsa_cols.append(j)
            elif "venta" in h and ("$" in h or "cop" in h or "pes" in h):
                venta_cols.append(j)
            elif h == "venta" or h.startswith("venta "):
                venta_cols.append(j)
        # Detectar comercializador: nombre en la fila de encabezado o la anterior.
        for r in (header_row - 1, header_row):
            if r < 0:
                continue
            for c in grid[r]:
                t = _norm(c)
                for com in COMERCIALIZADORES:
                    if com in t:
                        comercializador = _txt(c)
                        break
                if comercializador:
                    break
            if comercializador:
                break

    def _sum_cols(cols: list[int]) -> float:
        total = 0.0
        if header_row is None:
            return total
        for r in range(header_row + 1, min(header_row + 60, len(grid))):
            row = grid[r]
            for j in cols:
                if j < len(row):
                    v = _num(row[j])
                    if v is not None:
                        total += v
        return total

    venta = _sum_cols(venta_cols)
    venta_bolsa = _sum_cols(venta_bolsa_cols)
    compra_bolsa = _sum_cols(compra_bolsa_cols)
    tiene_bolsa = bool(venta_bolsa_cols or compra_bolsa_cols) and (venta_bolsa != 0 or compra_bolsa != 0)

    if header_row is None or not venta_cols:
        warnings.append("No se detectó la columna 'Venta ($)'; ingreso bruto = 0.")

    ingreso_bruto = venta
    total_ingresos = venta + venta_bolsa - compra_bolsa

    return {
        "comercializador": comercializador,
        "tiene_bolsa": tiene_bolsa,
        "ingreso_bruto": ingreso_bruto,
        "total_ingresos": total_ingresos,
        "venta_bolsa": venta_bolsa,
        "compra_bolsa": compra_bolsa,
    }


def _xm_concepto(nm: str) -> str | None:
    """
    Resuelve la etiqueta de un renglón de Comercialización XM a su concepto.
    `nm` viene normalizado y SIN puntos (para que 'i.v.a.' == 'iva').
    Distingue Generador/Comercializador a partir de la propia etiqueta.
    """
    gen = "generador" in nm
    com = "comercializador" in nm
    suf = " (Gen)" if gen else (" (Com)" if com else "")
    if "arranque y parada" in nm:
        return "Arranque y parada"
    if "energia en bolsa" in nm:
        return "Energía en Bolsa" + suf
    if "iva" in nm:
        return "IVA " + ("Generador" if gen else ("Comercializador" if com else "")).strip()
    if "despacho" in nm and "cnd" in nm:
        return "Serv. Despacho CND" + suf
    if ("administracion sic" in nm) or ("admin" in nm and "sic" in nm):
        return "Serv. Admin SIC" + suf
    return None


def _parse_comercializacion(grid: list[list]) -> list[dict]:
    """
    Comercialización XM desglosada. La etiqueta está en una columna (C) y el
    valor TOTAL en COP en la columna inmediatamente siguiente (D). A la derecha
    hay una sección de tarifas (Tarifa plataforma/operación…) que NO debe leerse:
    por eso el valor se toma como el PRIMER numérico tras la etiqueta, nunca el
    último. Los renglones en 0 (p.ej. la versión Comercializador cuando solo
    aplica el Generador) se descartan para no duplicar conceptos.
    """
    # Acotar al bloque entre "Ingresos y costos XM" y "Total Comercialización".
    start = end = None
    for i, row in enumerate(grid):
        nm = _norm(_row_label(row)).replace(".", "")
        if start is None and "ingresos y costos xm" in nm:
            start = i
        elif start is not None and "total comercializa" in nm:
            end = i
            break
    rows = grid[start + 1:end] if (start is not None and end is not None) else grid

    out: list[dict] = []
    vistos: set[str] = set()
    for row in rows:
        etiqueta = _row_label(row)
        if not etiqueta:
            continue
        nm = _norm(etiqueta).replace(".", "")
        concepto = _xm_concepto(nm)
        if not concepto:
            continue
        val = _value_after_label(row)
        if val is None or round(val, 2) == 0:
            continue
        if concepto in vistos:
            continue
        vistos.add(concepto)
        out.append({"concepto": concepto, "valor": -abs(val)})
    return out


# Costos operativos buscados por etiqueta. iva=True → aplica IVA 19% como línea aparte.
_COSTO_CONCEPTOS = [
    ("arrend", "Arrendamiento", False),
    ("arriendo", "Arrendamiento", False),
    ("fondo de mantenimiento", "Fondo de mantenimiento", True),
    ("fondo mantenimiento", "Fondo de mantenimiento", True),
    ("mantenimiento", "Mantenimiento", True),
    ("poliza", "Póliza", False),
    ("internet", "Servicio de Internet", True),
    ("servicios publicos", "Servicios públicos", False),
    ("servicio publico", "Servicios públicos", False),
]


def _parse_costos(grid: list[list]) -> list[dict]:
    """
    Costos operativos. IVA = 19% sobre Mantenimiento e Internet (siempre),
    se devuelve como flag `iva` por línea para que el frontend/loader lo expanda.
    """
    out: list[dict] = []
    vistos: set[str] = set()
    for row in grid:
        etiqueta = _row_label(row)
        if not etiqueta:
            continue
        ne = _norm(etiqueta)
        for key, label, aplica_iva in _COSTO_CONCEPTOS:
            if key in ne:
                if label in vistos:
                    continue
                # Igual que en XM: el valor en COP está en la columna contigua a la
                # etiqueta; tomar el primer numérico (no el último, que sería una tarifa).
                val = _value_after_label(row)
                if val is None or round(val, 2) == 0:
                    break
                vistos.add(label)
                out.append({"concepto": label, "valor": -abs(val), "iva": aplica_iva})
                break
    return out


def _parse_kwh(grid: list[list]) -> float | None:
    """Energía generada (kWh) del período — para Representación/CGM (kWh×5)."""
    for row in grid:
        etiqueta = _row_label(row)
        ne = _norm(etiqueta)
        if not ne:
            continue
        if ("kwh" in ne or "energia" in ne or "generacion" in ne) and "bolsa" not in ne:
            if "total" in ne or "generad" in ne or "kwh" in ne:
                v = _row_value(row)
                if v and v > 0:
                    return v
    return None


def _row_label(row: list) -> str:
    """Primera celda de texto no numérico de la fila (la etiqueta del concepto)."""
    for c in row:
        if c is None:
            continue
        if isinstance(c, str) and c.strip() and _num(c) is None:
            return c.strip()
    return ""


def _row_value(row: list):
    """Último valor numérico de la fila (el monto del concepto)."""
    val = None
    for c in row:
        n = _num(c) if not isinstance(c, str) else (_num(c) if re.search(r"\d", str(c)) else None)
        if n is not None:
            val = n
    return val


def _value_after_label(row: list):
    """
    Primer valor numérico que aparece DESPUÉS de la celda de etiqueta (la primera
    celda de texto). En los ER el monto en COP está justo a la derecha del concepto
    (columna D si la etiqueta está en C); las columnas más a la derecha contienen
    tarifas/factores que no deben leerse.
    """
    label_idx = None
    for i, c in enumerate(row):
        if isinstance(c, str) and c.strip() and _num(c) is None:
            label_idx = i
            break
    if label_idx is None:
        return None
    for c in row[label_idx + 1:]:
        n = _num(c) if not isinstance(c, str) else (_num(c) if re.search(r"\d", str(c)) else None)
        if n is not None:
            return n
    return None

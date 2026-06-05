"""
Funciones puras para cargar el panel de seguimiento contable (XLSX)
directamente a la base de datos via SQLAlchemy.
"""
import re
import unicodedata
from datetime import date as date_type

from sqlalchemy.orm import Session, selectinload

from app.models.clientes import Cliente
from app.models.liquidaciones import (
    Liquidacion, LiquidacionCosto, LiquidacionMandato,
    LiquidacionMandatoLinea, LiquidacionFactura,
)
from app.models.proyectos import Proyecto, ProyectoInversionista


# ── Constantes ─────────────────────────────────────────────────────────────────

TIPO_LINEA_MAP = [
    (r"ingreso bruto",                         "ingreso_bruto"),
    (r"^ingreso$",                             "ingreso_bruto"),
    (r"despacho",                              "despacho"),
    (r"ventas en bolsa",                       "ventas_en_bolsa"),
    (r"compras en bolsa|compras a ",           "compras_en_bolsa"),
    (r"redistribuci",                          "redistribucion_ingresos"),
    (r"comercializaci",                        "ajuste_comercializacion"),
    (r"ajuste.*xm|xm.*ajuste",                "ajuste_xm"),
    (r"ajuste.*unergy",                        "ajuste_unergy"),
    (r"ajuste.*administr",                     "otro_costo"),
    (r"arriendo",                              "arriendo"),
    (r"mantenimiento",                         "mantenimiento"),
    (r"iva.*internet|internet.*iva",           "iva"),
    (r"internet",                              "servicio_internet"),
    (r"poliza.*cumplimiento|p.liza.*cumpl",    "poliza_cumplimiento"),
    (r"poliza|incendio|lucro cesante",         "seguro"),
    (r"servicios p.blicos|consumo de energ",   "servicios_publicos_consumo"),
    (r"cambio.*equipo|equipo.*medida",         "cambio_equipos_medida"),
    (r"representaci",                          "representacion"),
    (r"^cgm",                                  "cgm"),
    (r"administraci|valor final",              "administracion"),
    (r"reteica",                               "reteica"),
    (r"ica opex",                              "ica_opex"),
    (r"iva|i\.v\.a",                           "iva"),
    (r"retenci",                               "retencion_fuente"),
    (r"porcentaje.*participaci",               "porcentaje_participacion"),
    (r"valor.*pagar|valor.*ganancia|utilidad", "valor_a_pagar"),
    (r"intereses",                             "intereses"),
]

_DOC_CONTABLE_MAP = {
    "informacion": "Información",
    "mandato":     "Mandato",
    "costos":      "Costos",
    "factura":     "Factura",
    "buenaventura": "Buenaventura",
}
TIPOS_VALIDOS_DOC = set(_DOC_CONTABLE_MAP.keys())

ALIASES: dict[str, str] = {
    "ayura sas":                    "ayura",
    "ayura s.a.s":                  "ayura",
    "solenium sas":                 "solenium",
    "solenium s.a.s":               "solenium",
    "patrimonios autonomos":        "patrimonios",
    "suno activos sostenibles":     "suno",
    "rodriguez velez beatriz":      "rodriguez velez",
    "inversiones estrada arbelaez": "estrada",
    "strada asociados":             "estrada",
}

ALIASES_INVERSIONISTA: dict[str, str | None] = {
    "17844 sol de la sierra":       "patrimonios autonomos fiduciaria bancolombia",
    "patrimonios autonomos fiduciaria bancolombia s a sociedad fiduciaria - 17844 sol de la sierra":
                                    "patrimonios autonomos fiduciaria bancolombia",
}

_OMITIR_PROY = re.compile(r'\btrading\b|duplicado|dulicado|\bdup\b|\bgasto\b', re.IGNORECASE)


# ── Helpers puros ──────────────────────────────────────────────────────────────

def normalizar(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or "").lower().strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def normalizar_alias(nombre: str) -> str:
    n = normalizar(nombre)
    for patron, alias in ALIASES.items():
        if patron in n:
            return alias
    return n


def normalizar_alias_inv(nombre: str) -> str | None:
    """Retorna alias canónico, o None si el inversionista debe ignorarse."""
    n = normalizar(nombre)
    for patron, alias in ALIASES_INVERSIONISTA.items():
        if patron in n:
            return alias  # None = sin match explícito
    return n


def normalizar_doc_contable(s: str) -> str:
    return _DOC_CONTABLE_MAP.get(normalizar(s), s.strip())


def concepto_a_tipo_linea(concepto: str) -> str:
    norm = normalizar(concepto)
    for patron, tipo in TIPO_LINEA_MAP:
        if re.search(patron, norm):
            return tipo
    return "otro_ingreso"


def concepto_a_tipo_costo(concepto: str) -> str:
    norm = normalizar(concepto)
    if "arriendo" in norm:                          return "arriendo"
    if "mantenimiento" in norm:                     return "mantenimiento"
    if "internet" in norm:                          return "internet"
    if "poliza" in norm or "p.liza" in norm:        return "polizas"
    if "servicios p" in norm:                       return "servicios_publicos"
    if "cambio" in norm and "equipo" in norm:       return "cambio_equipos_medida"
    if "comercializaci" in norm:                    return "comercializacion_xm"
    return "otro"


def concepto_a_tipo_factura(concepto: str) -> str:
    norm = normalizar(concepto)
    if "representaci" in norm:  return "representacion"
    if "cgm" in norm:           return "cgm"
    return "administracion_operacion"


def parse_valor(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[$,\s]", "", str(v).strip())
    if not s or s in ("-", ""):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def celda(row, idx: int):
    return row[idx] if idx < len(row) else None


def link(row, idx: int) -> str | None:
    cell = row[idx] if idx < len(row) else None
    if cell and hasattr(cell, "hyperlink") and cell.hyperlink:
        return cell.hyperlink.target
    return None


def valor(row, idx: int):
    c = celda(row, idx)
    return c.value if c is not None else None


# ── Leer hoja ──────────────────────────────────────────────────────────────────

def leer_hoja(xlsx_path: str, hoja: str) -> tuple[list[dict], dict[str, str]]:
    """
    Devuelve (filas, er_map).
    - filas: lista de dicts con datos de la tabla principal (cols A-K)
    - er_map: dict {nombre_proy_normalizado: er_url} de la tabla lateral (N-O)
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    if hoja not in wb.sheetnames:
        raise ValueError(f"Hoja '{hoja}' no encontrada. Disponibles: {wb.sheetnames}")
    sh = wb[hoja]

    rows = []
    er_map: dict[str, str] = {}

    for xl_row in sh.iter_rows(min_row=2, max_row=sh.max_row):
        if len(xl_row) > 14:
            n_val = valor(xl_row, 13)
            o_cell = xl_row[14]
            if n_val:
                er_url = (o_cell.hyperlink.target if o_cell.hyperlink else None) \
                         or str(o_cell.value or "").strip()
                if er_url and er_url.startswith("http"):
                    er_map[normalizar(str(n_val).strip())] = er_url

        proy = valor(xl_row, 0)
        if not proy:
            continue
        rows.append({
            "proyecto":        str(proy).strip(),
            "inversionista":   str(valor(xl_row, 1) or "").strip(),
            "doc_contable":    normalizar_doc_contable(str(valor(xl_row, 2) or "")),
            "concepto":        str(valor(xl_row, 5) or "").strip(),
            "total":           parse_valor(valor(xl_row, 6)),
            "ref_factura":     str(valor(xl_row, 7) or "").strip(),
            "ref_factura_url": link(xl_row, 7),
            "cons_ing_txt":    str(valor(xl_row, 8) or "").strip(),
            "cons_ing_url":    link(xl_row, 8),
            "cons_cos_txt":    str(valor(xl_row, 9) or "").strip() if len(xl_row) > 9 else "",
            "comprobante":     str(valor(xl_row, 10) or "").strip() if len(xl_row) > 10 else "",
            "carpeta_url":     link(xl_row, 15),
        })
    wb.close()
    return rows, er_map


def obtener_nombres_hojas(xlsx_path: str) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


# ── Match helpers ──────────────────────────────────────────────────────────────

def match_proyecto(proyectos_db: list[dict], nombre: str) -> dict | None:
    norm = normalizar(nombre)

    for p in proyectos_db:
        if normalizar(p["nombre_comercial"]) == norm:
            return p

    _SUFIJOS = re.compile(
        r'\s*(trading|dup\b|duplicado|dulicado|terpel|excedentes|excendentes'
        r'|gasto|gastos)\s*$'
    )
    norm_sin_sufijo = _SUFIJOS.sub('', norm).strip()
    _m_trail = re.search(r'\b(\d+)\s*$', norm_sin_sufijo)
    trailing_num: str | None = _m_trail.group(1) if _m_trail else None
    _m_code = re.search(r'\b(mgs[\s\-]*\d+)\b', norm)
    excel_mgs: str | None = re.sub(r'[\s\-]+', '', _m_code.group(1)) if _m_code else None

    def _db_trailing(p: dict) -> str | None:
        m = re.search(r'\bn?(\d+)\s*$', normalizar(p["nombre_comercial"]))
        return m.group(1) if m else None

    def _db_mgs(p: dict) -> str | None:
        m = re.search(r'\b(mgs[\s\-]*\d+)\b', normalizar(p["nombre_comercial"]))
        return re.sub(r'[\s\-]+', '', m.group(1)) if m else None

    def _num_ok(p: dict) -> bool:
        if trailing_num is None:
            return True
        db_num = _db_trailing(p)
        # Si el Excel tiene número final, exigir que el proyecto DB también lo tenga
        # y que coincida. Evita que proyectos sin número (ej. "Chima Oriente")
        # capturen nombres numerados (ej. "Valencia Oriente 1").
        if db_num is None:
            return False
        return db_num == trailing_num

    if excel_mgs:
        for p in proyectos_db:
            if _db_mgs(p) == excel_mgs:
                return p

    for p in proyectos_db:
        n = normalizar(p["nombre_comercial"])
        if n and (n in norm or norm in n):
            if not _num_ok(p):
                continue
            return p

    partes = [t for t in norm.split() if len(t) >= 4 and not t.isdigit()]
    for parte in reversed(partes):
        for p in proyectos_db:
            if parte in normalizar(p["nombre_comercial"]):
                if not _num_ok(p):
                    continue
                return p

    return None


def match_inversionista(
    inversionistas_db: list[dict],
    nombre: str,
    periodo_date=None,
) -> dict | None:
    if not nombre or nombre.upper() == "TOTAL":
        return None

    if periodo_date:
        activos = [
            inv for inv in inversionistas_db
            if (not inv.get("fecha_inicio") or
                inv["fecha_inicio"] <= periodo_date.isoformat())
            and (not inv.get("fecha_fin") or
                inv["fecha_fin"] >= periodo_date.isoformat())
        ]
        if activos:
            inversionistas_db = activos

    def _nombre_db(inv: dict) -> str:
        return (inv.get("cliente_nombre") or inv.get("razon_social_nombre") or
                inv.get("nombre") or inv.get("razon_social") or "")

    norm_inv = normalizar_alias_inv(nombre)
    if norm_inv is None:
        return None  # alias explícito "sin match"
    norm = norm_inv
    for inv in inversionistas_db:
        cn = normalizar(_nombre_db(inv))
        if cn and (cn == norm or cn in norm or norm in cn):
            return inv
    palabras = [w for w in norm.split() if len(w) > 4]
    for inv in inversionistas_db:
        cn = normalizar(_nombre_db(inv))
        if palabras and cn and all(w in cn for w in palabras[:2]):
            return inv

    norm_alias = normalizar_alias(nombre)
    if norm_alias != norm:
        for inv in inversionistas_db:
            cn = normalizar_alias(_nombre_db(inv))
            if cn and cn == norm_alias:
                return inv

    return None


# ── DB helpers ─────────────────────────────────────────────────────────────────

def _tiene_mandatos_db(db: Session, liq_id: int) -> bool:
    return db.query(LiquidacionMandato).filter(
        LiquidacionMandato.liquidacion_id == liq_id
    ).first() is not None


def _limpiar_db(db: Session, liq_id: int):
    mandatos = db.query(LiquidacionMandato).filter(
        LiquidacionMandato.liquidacion_id == liq_id
    ).all()
    for m in mandatos:
        db.query(LiquidacionMandatoLinea).filter(
            LiquidacionMandatoLinea.mandato_id == m.id
        ).delete(synchronize_session=False)
        db.delete(m)
    db.query(LiquidacionCosto).filter(
        LiquidacionCosto.liquidacion_id == liq_id
    ).delete(synchronize_session=False)
    db.query(LiquidacionFactura).filter(
        LiquidacionFactura.liquidacion_id == liq_id
    ).delete(synchronize_session=False)
    db.commit()


# ── Carga principal ────────────────────────────────────────────────────────────

def cargar_desde_db(
    db: Session,
    filas: list[dict],
    er_map: dict[str, str],
    periodo_date: str,
    limpiar: bool,
    dry_run: bool,
    usuario_id: int,
) -> dict:
    """
    Versión DB-directa de cargar(). Usa SQLAlchemy en lugar de llamadas HTTP.
    periodo_date: 'YYYY-MM-DD' (primer día del mes).
    """
    errores: list[str] = []

    # Validar tipos de documento desconocidos
    tipos_desconocidos: dict[str, list[int]] = {}
    for i, f in enumerate(filas, start=2):
        nd = normalizar(f["doc_contable"])
        if nd and nd not in TIPOS_VALIDOS_DOC:
            tipos_desconocidos.setdefault(f["doc_contable"], []).append(i)

    y_str, m_str, _ = periodo_date.split("-")
    period = date_type(int(y_str), int(m_str), 1)

    proyectos_db_objs = db.query(Proyecto).order_by(Proyecto.id).all()
    proyectos_db = [{"id": p.id, "nombre_comercial": p.nombre_comercial} for p in proyectos_db_objs]

    grupos: dict[str, dict[str, list]] = {}
    for f in filas:
        grupos.setdefault(f["proyecto"], {}).setdefault(f["inversionista"], []).append(f)

    stats = {
        "proyectos_cargados": 0,
        "proyectos_omitidos": 0,
        "proyectos_sin_match": [],
        "mandatos": 0,
        "lineas": 0,
        "costos": 0,
        "facturas": 0,
        "inversionistas_sin_match": [],
    }

    if dry_run:
        encontrados = []
        no_encontrados = []
        for nombre_proy in grupos:
            if _OMITIR_PROY.search(nombre_proy):
                continue
            proy_db = match_proyecto(proyectos_db, nombre_proy)
            if proy_db:
                encontrados.append(nombre_proy)
            else:
                no_encontrados.append(nombre_proy)
        return {
            "ok": True,
            "dry_run": True,
            "proyectos_encontrados": encontrados,
            "proyectos_no_encontrados": no_encontrados,
        }

    for nombre_proy, inv_grupos in grupos.items():
        if _OMITIR_PROY.search(nombre_proy):
            stats["proyectos_omitidos"] += 1
            continue

        proy_db = match_proyecto(proyectos_db, nombre_proy)
        if not proy_db:
            stats["proyectos_sin_match"].append(nombre_proy)
            continue

        pid = proy_db["id"]
        er_url = er_map.get(normalizar(nombre_proy))

        # Crear o recuperar liquidación
        liq = (
            db.query(Liquidacion)
            .filter(Liquidacion.proyecto_id == pid, Liquidacion.periodo == period)
            .filter(Liquidacion.deleted_at.is_(None))
            .first()
        )

        if liq is None:
            liq = Liquidacion(
                proyecto_id=pid,
                periodo=period,
                tipo_venta="ppa",
                generado_por_id=usuario_id,
            )
            db.add(liq)
            try:
                db.commit()
                db.refresh(liq)
            except Exception as exc:
                db.rollback()
                errores.append(f"{nombre_proy}: error creando liquidación — {exc}")
                continue
        else:
            if not limpiar and _tiene_mandatos_db(db, liq.id):
                stats["proyectos_omitidos"] += 1
                continue
            if limpiar:
                _limpiar_db(db, liq.id)

        liq_id = liq.id

        # Actualizar metadatos
        patch: dict = {}
        if er_url:
            patch["estado_resultados_url"] = er_url
        total_rows = inv_grupos.get("Total", [])
        for f in total_rows:
            if f["doc_contable"].lower() == "información":
                if f["comprobante"]:
                    patch["comprobante_contable_ref"] = f["comprobante"]
                if f["cons_ing_txt"].isdigit():
                    patch["consecutivo_inicial_ingresos"] = int(f["cons_ing_txt"])
                if f["cons_cos_txt"].isdigit():
                    patch["consecutivo_inicial_costos"] = int(f["cons_cos_txt"])
        if patch:
            for k, v in patch.items():
                setattr(liq, k, v)
            db.commit()

        # Obtener inversionistas del proyecto con JOIN explícito para garantizar
        # que razon_social_nombre esté disponible sin depender de lazy loading.
        inv_rows = (
            db.query(
                ProyectoInversionista.id,
                ProyectoInversionista.es_patrimonio_autonomo,
                ProyectoInversionista.fecha_inicio,
                ProyectoInversionista.fecha_fin,
                Cliente.razon_social_nombre,
            )
            .outerjoin(Cliente, ProyectoInversionista.cliente_id == Cliente.id)
            .filter(ProyectoInversionista.proyecto_id == pid)
            .all()
        )
        inversionistas_db = [
            {
                "id": row.id,
                "cliente_nombre": row.razon_social_nombre or "",
                "es_patrimonio_autonomo": row.es_patrimonio_autonomo,
                "fecha_inicio": row.fecha_inicio.isoformat() if row.fecha_inicio else None,
                "fecha_fin": row.fecha_fin.isoformat() if row.fecha_fin else None,
            }
            for row in inv_rows
        ]

        for inv_nombre, filas_inv in inv_grupos.items():
            es_total = inv_nombre.upper() == "TOTAL"
            inv_db = None if es_total else match_inversionista(inversionistas_db, inv_nombre, period)
            inv_id = inv_db["id"] if inv_db else None

            if not es_total and inv_db is None and inv_nombre.strip():
                stats["inversionistas_sin_match"].append(f"{nombre_proy} → {inv_nombre}")

            cons_ing = cons_cos = None
            for f in filas_inv:
                if f["doc_contable"].lower() == "información":
                    if f["cons_ing_txt"].isdigit():
                        cons_ing = int(f["cons_ing_txt"])
                    if f["cons_cos_txt"].isdigit():
                        cons_cos = int(f["cons_cos_txt"])

            filas_ing = [f for f in filas_inv if f["doc_contable"].lower() == "mandato"]
            filas_cos = [f for f in filas_inv if f["doc_contable"].lower() == "costos"]
            filas_fac = [f for f in filas_inv if f["doc_contable"].lower() in ("factura", "buenaventura")]

            # ── Mandato de ingresos ──
            lineas_ing = [
                f for f in filas_ing
                if f["concepto"]
                and normalizar(f["concepto"]) not in ("porcentaje de participacion", "valor a pagar")
                and f["total"] is not None
            ]
            if lineas_ing:
                mandato = LiquidacionMandato(
                    liquidacion_id=liq_id,
                    tipo="ingresos",
                    inversionista_id=inv_id,
                    beneficiario_nombre=inv_nombre if not es_total else None,
                    consecutivo=cons_ing,
                    pa_aplica=inv_db.get("es_patrimonio_autonomo", False) if inv_db else False,
                )
                db.add(mandato)
                try:
                    db.commit()
                    db.refresh(mandato)
                    mid = mandato.id
                    stats["mandatos"] += 1
                except Exception as exc:
                    db.rollback()
                    errores.append(f"{nombre_proy}/{inv_nombre}: error creando mandato ingresos — {exc}")
                    continue

                neto_pagar = None
                for orden, f in enumerate(filas_ing):
                    if not f["concepto"] or f["total"] is None:
                        continue
                    if normalizar(f["concepto"]) == "valor a pagar":
                        neto_pagar = f["total"]
                        continue
                    tipo_l = concepto_a_tipo_linea(f["concepto"])
                    ref = (f["ref_factura"].split(" | ")[0].strip()
                           if " | " in f["ref_factura"] else f["ref_factura"]) or None
                    if ref and len(ref) > 255:
                        ref = ref[:252] + "..."
                    soporte = f["ref_factura_url"] or f["cons_ing_url"] or None
                    try:
                        linea = LiquidacionMandatoLinea(
                            mandato_id=mid,
                            tipo_linea=tipo_l,
                            concepto=f["concepto"],
                            valor_cop=f["total"],
                            referencia_factura=ref,
                            soporte_url=soporte,
                            orden=orden,
                        )
                        db.add(linea)
                        db.commit()
                        stats["lineas"] += 1
                    except Exception as exc:
                        db.rollback()
                        errores.append(f"{nombre_proy}/{inv_nombre}: línea '{f['concepto']}' — {exc}")

                if neto_pagar is not None:
                    try:
                        mandato.valor_neto_cop = neto_pagar
                        db.commit()
                    except Exception as exc:
                        db.rollback()
                        errores.append(f"{nombre_proy}/{inv_nombre}: patch valor_neto — {exc}")

            # ── Mandato de costos ──
            lineas_cos = [
                f for f in filas_cos
                if f["concepto"] and f["total"] is not None
                and normalizar(f["concepto"]) not in ("porcentaje de participacion", "valor a pagar")
            ]
            if lineas_cos:
                mandato_c = LiquidacionMandato(
                    liquidacion_id=liq_id,
                    tipo="costos",
                    inversionista_id=inv_id,
                    beneficiario_nombre=inv_nombre if not es_total else None,
                    consecutivo=cons_cos,
                    pa_aplica=inv_db.get("es_patrimonio_autonomo", False) if inv_db else False,
                )
                db.add(mandato_c)
                try:
                    db.commit()
                    db.refresh(mandato_c)
                    mcid = mandato_c.id
                    stats["mandatos"] += 1
                except Exception as exc:
                    db.rollback()
                    errores.append(f"{nombre_proy}/{inv_nombre}: error creando mandato costos — {exc}")
                    continue

                neto_cos = None
                for orden, f in enumerate(filas_cos):
                    if not f["concepto"] or f["total"] is None:
                        continue
                    if normalizar(f["concepto"]) == "valor a pagar":
                        neto_cos = f["total"]
                        continue
                    tipo_l = concepto_a_tipo_linea(f["concepto"])
                    ref = (f["ref_factura"].split(" | ")[0].strip()
                           if " | " in f["ref_factura"] else f["ref_factura"]) or None
                    if ref and len(ref) > 255:
                        ref = ref[:252] + "..."
                    soporte = f["ref_factura_url"] or f["cons_ing_url"] or None
                    try:
                        linea_c = LiquidacionMandatoLinea(
                            mandato_id=mcid,
                            tipo_linea=tipo_l,
                            concepto=f["concepto"],
                            valor_cop=f["total"],
                            referencia_factura=ref,
                            soporte_url=soporte,
                            orden=orden,
                        )
                        db.add(linea_c)
                        db.commit()
                        stats["lineas"] += 1
                    except Exception as exc:
                        db.rollback()
                        errores.append(f"{nombre_proy}/{inv_nombre}: línea costo '{f['concepto']}' — {exc}")
                        continue

                    if es_total and f["total"] != 0:
                        try:
                            costo = LiquidacionCosto(
                                liquidacion_id=liq_id,
                                tipo_costo=concepto_a_tipo_costo(f["concepto"]),
                                descripcion=f["concepto"],
                                nro_soporte=(f["ref_factura"] or f["comprobante"]) or None,
                                soporte_url=f["cons_ing_url"],
                                valor_cop=f["total"],
                            )
                            db.add(costo)
                            db.commit()
                            stats["costos"] += 1
                        except Exception as exc:
                            db.rollback()
                            errores.append(f"{nombre_proy}: costo '{f['concepto']}' — {exc}")

                if neto_cos is not None:
                    try:
                        mandato_c.valor_neto_cop = neto_cos
                        db.commit()
                    except Exception as exc:
                        db.rollback()
                        errores.append(f"{nombre_proy}/{inv_nombre}: patch valor_neto costos — {exc}")

            # ── Facturas de servicio — Total y por inversionista ──
            for f in filas_fac:
                if not f["concepto"] or f["total"] is None:
                    continue
                try:
                    nro_fac = (f["ref_factura"].split(" | ")[0].strip()
                               if " | " in f["ref_factura"]
                               else f["ref_factura"]) or None
                    factura = LiquidacionFactura(
                        liquidacion_id=liq_id,
                        proyecto_inversionista_id=inv_db["id"] if inv_db else None,
                        tipo_servicio=concepto_a_tipo_factura(f["concepto"]),
                        numero_factura=nro_fac,
                        nro_soporte=f["cons_ing_txt"] or None,
                        soporte_url=f["cons_ing_url"],
                        valor_cop=f["total"],
                    )
                    db.add(factura)
                    db.commit()
                    stats["facturas"] += 1
                except Exception as exc:
                    db.rollback()
                    errores.append(f"{nombre_proy}: factura '{f['concepto']}' — {exc}")

        stats["proyectos_cargados"] += 1

    return {"ok": True, "stats": stats, "errores": errores}

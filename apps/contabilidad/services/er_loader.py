"""
Carga y parseo de archivos "Estado de Resultados" (ER) por proyecto.

Cada ER es un xlsx con fórmulas. El flujo es:
  1. recalcular_er(): LibreOffice headless recalcula las fórmulas y reescribe el
     archivo, para poder leer los valores con openpyxl(data_only=True).
  2. parsear_er(): extrae ingreso bruto, comercialización XM desglosada, costos
     operativos, IVA y fee de administración → dict estructurado.
  3. match_proyecto(): se reutiliza el de liquidaciones_loader.

El parser es tolerante a la posición exacta de las filas: detecta secciones y
columnas por etiqueta, usando las filas indicadas en la especificación solo como
guía. Todos los valores quedan editables en el frontend.
"""
import os
import re
import shutil
import subprocess
import tempfile
import unicodedata

# Reutilizamos los helpers de matching ya probados en liquidaciones.
from apps.comun.nombres_excel import match_proyecto, normalizar  # noqa: F401

IVA = 0.19
FEE_ADMIN = 0.0380  # Fee Administración = Total Ingresos × 3.80%

# Nombres de comercializadores conocidos (solo informativo; el parser detecta la
# columna "Venta ($)" sin importar el nombre).
COMERCIALIZADORES = [
    "terpel", "neu", "biac", "sol&cielo", "sol & cielo", "solycielo",
    "enermas", "unergy",
]

# Tokens de razón social que NO aparecen nunca en un nombre de proyecto y que
# contaminan el match (ej. "S.A.S." matcheaba con "M.D.M. Cientifica S.A.S").
_RAZON_SOCIAL = re.compile(r'^(s\.?a\.?s\.?|e\.?s\.?p\.?|s\.?a\.?|ltda\.?|cia\.?)$', re.I)


# ── Nombre de proyecto desde el nombre de archivo ───────────────────────────────

# Meses (para limpiar sufijos tipo "_MAYO", "_Abril 2026" en los ER ALL_DATA).
_MESES_RE = (
    r'enero|febrero|marzo|abril|mayo|junio|julio|agosto|'
    r'septiembre|setiembre|octubre|noviembre|diciembre'
)


def extraer_proyecto_de_archivo(nombre: str, proyectos_db: list[dict]) -> dict | None:
    """
    Soporta dos patrones de nombre de archivo:
      1. Normal: "Estado resultados {INVERSIONISTA} {PROYECTO} N 2026.xlsx".
      2. NITRO/NEU ALL_DATA: "{PLANTA}_NITRO_ALL_DATA_{MES}.xlsx" (o NEU/sin tipo),
         ej. "CACICA_NITRO_ALL_DATA_MAYO.xlsx" → "CACICA".
    En el normal, el inversionista tiene longitud variable, así que probamos
    match_proyecto() con ventanas deslizantes desde la derecha (el proyecto va al
    final). Reusa match_proyecto (con sus aliases de NAOS, Cacica, etc.).
    """
    base = re.sub(r'^Estado\s+resultados\s+', '', nombre or '', flags=re.I)
    base = re.sub(r'\.xlsx?$', '', base, flags=re.I)

    # Patrón ALL_DATA: quitar "_(NITRO|NEU)_ALL_DATA_..." / "_ALL_DATA..." y el mes,
    # y pasar guiones bajos a espacios → queda solo el nombre de la planta.
    base = re.sub(r'[_\s]*(nitro|neu)?[_\s]*all[_\s]*data.*$', '', base, flags=re.I)
    base = re.sub(rf'[_\s]*({_MESES_RE})([_\s]*20\d{{2}})?$', '', base, flags=re.I)
    base = base.replace('_', ' ').strip()

    # Quitar el sufijo " N 2026" (consecutivo + año), conservando números que
    # son parte del nombre del proyecto (ej. "Valencia Oriente 1", "GD NAOS 1").
    base = re.sub(r'\s+\d+\s+20\d{2}$', '', base)

    palabras = [w for w in base.split() if not _RAZON_SOCIAL.match(w)]
    for n in range(min(6, len(palabras)), 0, -1):
        candidato = " ".join(palabras[-n:])
        m = match_proyecto(proyectos_db, candidato)
        if m:
            return m
    return None


# ── LibreOffice: recalcular fórmulas ────────────────────────────────────────────

def recalcular_er(path: str) -> str:
    """
    Convierte el xlsx a uno con las fórmulas ya calculadas usando LibreOffice
    headless. Devuelve la ruta del archivo recalculado (en un tmpdir).

    Si LibreOffice no está disponible (entorno local sin soffice), se devuelve
    el archivo original tal cual; openpyxl data_only leerá los valores cacheados
    que Excel/LibreOffice hayan guardado.
    """
    soffice = _find_soffice()
    if not soffice:
        return path

    outdir = tempfile.mkdtemp(prefix="er_recalc_")
    try:
        # --convert-to xlsx fuerza el recálculo de fórmulas al reescribir.
        subprocess.run(
            [
                soffice, "--headless", "--calc",
                "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                "--outdir", outdir, path,
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=120,
            env={**os.environ, "HOME": outdir},
        )
    except Exception:
        shutil.rmtree(outdir, ignore_errors=True)
        return path

    base = os.path.splitext(os.path.basename(path))[0] + ".xlsx"
    recalc = os.path.join(outdir, base)
    return recalc if os.path.exists(recalc) else path


def _find_soffice() -> str | None:
    for cand in ("soffice", "libreoffice"):
        found = shutil.which(cand)
        if found:
            return found
    # Rutas típicas en contenedor / Windows
    for p in (
        "/usr/bin/soffice", "/usr/bin/libreoffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
    ):
        if os.path.exists(p):
            return p
    return None


# ── Helpers de lectura ──────────────────────────────────────────────────────────

def _num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[$\s]", "", str(v).strip())
    s = s.replace(".", "").replace(",", ".") if s.count(",") == 1 and s.count(".") > 1 else s.replace(",", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s or s in ("-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _txt(v) -> str:
    return str(v).strip() if v is not None else ""


def _norm(s) -> str:
    s = unicodedata.normalize("NFD", str(s or "").lower().strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


# ── Parser principal ─────────────────────────────────────────────────────────────

def parsear_er(path: str, tipo: str = "normal", mapeos: dict | None = None,
               aliases: dict | None = None, proyecto_nombre: str | None = None) -> dict:
    """
    Extrae del ER (ya recalculado) un dict estructurado. La hoja
    "Estado_de_resultados" es la misma en los 3 tipos; SOLO cambia cómo se lee
    la sección de ingresos (`tipo` ∈ {'normal','neu','nitro'}). Comercialización
    XM, costos operativos y fee de administración se leen igual en los 3.

    El parser PROPONE: por cada concepto devuelve el valor y la celda exacta de
    donde lo tomó (`hoja`/`celda`, ej. "Sheet1"/"H35"). Si `mapeos` trae una
    celda guardada para un concepto (RECORDAR), lee ESA celda directamente en vez
    de proponer. `mapeos`: {concepto_normalizado: {"hoja": str, "celda": str}}.

    {
      "tipo": str,
      "comercializador": str | None,
      "tiene_bolsa": bool,
      "ingreso_bruto": float,
      "total_ingresos": float,
      "ingresos_detalle": [ {concepto, valor, hoja, celda}, ... ],
      "comercializacion": [ {concepto, valor, hoja, celda}, ... ],   # XM (negativos)
      "costos": [ {concepto, valor, iva, hoja, celda} ],             # operativos
      "facturas": [ {concepto, valor, hoja, celda} ],                # repr, CGM, admin
      "kwh": float | None,
      "snapshot": {hoja: {coord: valor}},   # celdas numéricas del ER recalculado
      "warnings": [str],
    }
    """
    from openpyxl import load_workbook

    tipo = (tipo or "normal").strip().lower()
    if tipo not in ("normal", "neu", "nitro"):
        tipo = "normal"
    mapeos = mapeos or {}
    aliases = aliases or {}

    wb = load_workbook(path, data_only=True)
    primera = wb.sheetnames[0]
    sh = wb[primera]
    grid = [[c.value for c in row] for row in sh.iter_rows()]
    snapshot = _construir_snapshot(wb)
    # Grid de la hoja Summary (NITRO la usa para el ingreso bruto / bolsa).
    _summary_name = next((n for n in wb.sheetnames if n.lower() == "summary"), None)
    summary_grid = (
        [[c.value for c in row] for row in wb[_summary_name].iter_rows()]
        if _summary_name else None
    )
    wb.close()

    warnings: list[str] = []

    if tipo == "neu":
        ing = _parse_ingresos_neu(grid, warnings)
    elif tipo == "nitro":
        ing = _parse_ingresos_nitro(grid, warnings)
    else:
        ing = _parse_ingresos(grid, warnings)
    ingreso_bruto = ing["ingreso_bruto"]
    total_ingresos = ing["total_ingresos"]

    comercializacion = _parse_comercializacion(grid)
    # NEU/NITRO suelen traer la Comercialización como un único total (celda
    # Estado_de_resultados!D20 o Sheet1!D52), no como desglose XM. Si no se
    # detectó desglose, emitir un renglón único mapeable a esa celda.
    if tipo in ("neu", "nitro") and not comercializacion:
        comercializacion = [{"concepto": "Comercialización", "valor": 0.0}]
    costos = _parse_costos(grid)
    kwh = _parse_kwh(grid)

    # Facturas de servicio. Representación y CGM se proponen leyendo la fila
    # "Cobro OPEX: Representación/CGM" del ER (su tarifa por kWh varía por proyecto,
    # ej. ×5.26 o ×6.00); si no se encuentra la etiqueta, se cae a kWh × 5.
    # Administración = Total Ingresos × 3.80% (mapeable a una celda si difiere).
    rep = _buscar_etiqueta_valor(grid, [("representacion",)])
    cgm = _buscar_etiqueta_valor(grid, [("cgm",)])
    rep_val = rep if rep is not None else (kwh * 5 if kwh else 0.0)
    cgm_val = cgm if cgm is not None else (kwh * 5 if kwh else 0.0)
    facturas: list[dict] = [
        {"concepto": "Representación", "valor": -round(rep_val, 2)},
        {"concepto": "CGM", "valor": -round(cgm_val, 2)},
        {"concepto": "Administración", "valor": -round(total_ingresos * FEE_ADMIN, 2)},
    ]

    ingresos_detalle = ing.get("ingresos_detalle", [])

    # Garantizar que existan los renglones canónicos del tipo, aunque el parser no
    # los haya encontrado en la hoja principal (NEU/NITRO traen el desglose en
    # tabs Mandato!/Summary! que no se autodetectan): así la usuaria puede mapear
    # cada concepto a su celda y el sistema lo recuerda. Comercialización se maneja
    # aparte vía la sección XM (su suma = la "Comercialización" del Excel).
    canonicos = {
        "neu": ["Despacho de energía", "Venta en bolsa", "Compra en bolsa",
                "Redistribución de ingresos"],
        "nitro": ["Ingreso Bruto", "Venta en bolsa", "Compra en bolsa"],
    }.get(tipo, [])
    presentes = {_norm(d["concepto"]) for d in ingresos_detalle}
    for concepto in canonicos:
        if _norm(concepto) not in presentes:
            ingresos_detalle.append({"concepto": concepto, "valor": 0.0})

    # NITRO: el Ingreso Bruto / Ventas / Compras en bolsa viven en la hoja Summary
    # (fila "Total" del proyecto), NO en la sección de Sheet1 (que trae el balance).
    # Proponerlos desde ahí para que la carga fresca ya salga bien (la usuaria igual
    # puede remapear). Cacica=fila Total Cacica, Piloneras=fila Total Piloneras.
    if tipo == "nitro" and summary_grid:
        sm = _summary_nitro(summary_grid, _summary_name, proyecto_nombre)
        for d in ingresos_detalle:
            nd = _norm(d["concepto"])
            clave = ("ib" if "bruto" in nd else
                     "venta" if ("venta" in nd and "bolsa" in nd) else
                     "compra" if ("compra" in nd and "bolsa" in nd) else None)
            if clave and clave in sm:
                val, hoja_s, celda_s = sm[clave]
                d["valor"] = -abs(round(val, 2)) if clave == "compra" else round(val, 2)
                d["hoja"], d["celda"] = hoja_s, celda_s

    # Las fuentes de ingreso ya traen su celda exacta (columna Venta($) en la fila
    # TOTAL de Sheet1); fijar su hoja a la principal SOLO si no tienen una ya
    # (las de NITRO/Summary traen hoja="Summary" y no deben sobreescribirse).
    usados: set[str] = set()
    for ln in ingresos_detalle:
        if ln.get("celda"):
            if not ln.get("hoja"):
                ln["hoja"] = primera
            usados.add(f"{ln['hoja']}!{ln['celda']}")

    # PROPONER: para los renglones sin celda, buscarla en el snapshot por el valor
    # propuesto, sin reusar la misma celda dos veces (desambigua Repr/CGM, iguales).
    for grupo in (ingresos_detalle, comercializacion, costos, facturas):
        for ln in grupo:
            if ln.get("celda"):
                continue
            hoja, celda = _localizar_celda(snapshot, ln["valor"], primera, usados)
            ln["hoja"], ln["celda"] = hoja, celda
            if celda:
                usados.add(f"{hoja}!{celda}")

    # RECORDAR (ingresos): aplicar los alias de fuente guardados (celda → etiqueta).
    # Relabela las fuentes detectadas y RESUCITA las fuentes manuales que la usuaria
    # agregó otro mes (cuyo alias apunta a una celda con valor en el ER).
    if aliases:
        for ln in ingresos_detalle:
            if ln.get("hoja") and ln.get("celda"):
                a = aliases.get(f"{ln['hoja']}!{ln['celda']}".lower())
                if a:
                    ln["concepto"] = a["etiqueta"]
                    ln["orden_alias"] = a.get("orden", 0)
        presentes_celda = {f"{ln.get('hoja')}!{ln.get('celda')}".lower()
                           for ln in ingresos_detalle if ln.get("celda")}
        for col_origen, a in aliases.items():
            if col_origen in presentes_celda:
                continue
            try:
                hoja_a, celda_a = col_origen.split("!", 1)
            except ValueError:
                continue
            val = leer_celda(snapshot, hoja_a, celda_a)
            if val is None:
                continue
            es_compra = "compra" in _norm(a["etiqueta"]) and "bolsa" in _norm(a["etiqueta"])
            ingresos_detalle.append({
                "concepto": a["etiqueta"],
                "valor": -abs(round(val, 2)) if es_compra else round(val, 2),
                "hoja": hoja_a, "celda": celda_a.upper(), "orden_alias": a.get("orden", 0),
            })
        # Ordenar por el orden de alias cuando esté definido (estable para el resto).
        ingresos_detalle.sort(key=lambda d: d.get("orden_alias", 1_000_000))

    # RECORDAR: si hay una celda guardada para el concepto, leerla y sustituir el
    # valor propuesto (la usuaria ya la confirmó un mes anterior).
    for grupo_nombre, grupo in (
        ("ingresos", ingresos_detalle), ("comercializacion", comercializacion),
        ("costos", costos), ("facturas", facturas),
    ):
        for ln in grupo:
            m = mapeos.get(_norm(ln["concepto"]))
            if not m:
                continue
            val = leer_celda(snapshot, m["hoja"], m["celda"])
            if val is None:
                warnings.append(
                    f"Mapeo {ln['concepto']} → {m['hoja']}!{m['celda']} sin valor en el ER."
                )
                continue
            ln["valor"] = _aplicar_signo(grupo_nombre, ln["concepto"], val)
            ln["hoja"], ln["celda"] = m["hoja"], m["celda"]

    # Si el mapeo cambió ingreso bruto / detalle, recomputar los totales de ingresos.
    # Un proyecto puede tener VARIAS fuentes de ingreso bruto (Terpel 1, Terpel 2…):
    # sumar todas las líneas "bruto"/"despacho", no solo la primera.
    if ingresos_detalle:
        ib_lineas = [d["valor"] for d in ingresos_detalle
                     if "bruto" in _norm(d["concepto"]) or "despacho" in _norm(d["concepto"])]
        if ib_lineas:
            ingreso_bruto = sum(ib_lineas)
        total_ingresos = sum(d["valor"] for d in ingresos_detalle)

    return {
        "tipo": tipo,
        "comercializador": ing["comercializador"],
        "tiene_bolsa": ing["tiene_bolsa"],
        "ingreso_bruto": round(ingreso_bruto, 2),
        "total_ingresos": round(total_ingresos, 2),
        "venta_bolsa": round(ing["venta_bolsa"], 2),
        "compra_bolsa": round(ing["compra_bolsa"], 2),
        "ingresos_detalle": ingresos_detalle,
        "comercializacion": comercializacion,
        "costos": costos,
        "facturas": facturas,
        "kwh": kwh,
        "snapshot": snapshot,
        "warnings": warnings,
    }


# ── Snapshot y lectura de celdas por mapeo ──────────────────────────────────────

def _construir_snapshot(wb) -> dict:
    """
    {hoja: {coord: valor}} con las celdas NUMÉRICAS de todas las hojas del ER ya
    recalculado. Permite releer una celda al cambiar el mapeo sin re-subir el
    archivo, y localizar la celda de origen de un valor propuesto.
    """
    snap: dict[str, dict[str, float]] = {}
    for name in wb.sheetnames:
        celdas: dict[str, float] = {}
        for row in wb[name].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, bool):
                    continue
                if isinstance(v, (int, float)):
                    celdas[c.coordinate] = float(v)
        snap[name] = celdas
    return snap


def leer_celda(snapshot: dict, hoja: str, celda: str) -> float | None:
    """Valor numérico de hoja!celda en el snapshot (tolerante a mayúsculas/espacios)."""
    if not snapshot or not hoja or not celda:
        return None
    celda = str(celda).strip().upper().replace("$", "")
    hoja = str(hoja).strip()
    hojas = snapshot
    # Coincidencia exacta o case-insensitive del nombre de hoja.
    if hoja in hojas:
        celdas = hojas[hoja]
    else:
        match = next((h for h in hojas if h.lower() == hoja.lower()), None)
        if match is None:
            return None
        celdas = hojas[match]
    if celda in celdas:
        return celdas[celda]
    match = next((c for c in celdas if c.upper() == celda), None)
    return celdas[match] if match else None


def _localizar_celda(snapshot: dict, valor, hoja_pref: str, usados: set) -> tuple:
    """
    Celda (hoja, coord) cuyo valor coincide con `valor` (en magnitud), priorizando
    la hoja principal y sin reusar celdas ya asignadas. Devuelve (None, None) si no
    hay coincidencia (p.ej. Administración, que es un cálculo y no una celda).
    """
    if valor is None:
        return None, None
    objetivo = abs(round(float(valor), 2))
    if objetivo == 0:
        return None, None
    tol = max(1.0, objetivo * 1e-6)
    orden = [hoja_pref] + [h for h in snapshot if h != hoja_pref]
    for hoja in orden:
        for coord, v in snapshot.get(hoja, {}).items():
            if f"{hoja}!{coord}" in usados:
                continue
            if abs(abs(v) - objetivo) <= tol:
                return hoja, coord
    return None, None


def _aplicar_signo(grupo: str, concepto: str, valor: float) -> float:
    """
    Aplica la convención de signo del panel al valor leído de una celda:
    comercialización/costos/facturas son negativos; 'compra' en bolsa es negativa;
    el resto se respeta tal cual.
    """
    n = _norm(concepto)
    if grupo in ("comercializacion", "costos", "facturas"):
        return -abs(round(valor, 2))
    if "compra" in n and "bolsa" in n:
        return -abs(round(valor, 2))
    return round(valor, 2)


def _col_letter(j: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(j + 1)  # j es índice 0-based de la columna


def _fila_total(grid: list[list], header_row: int) -> int:
    """
    Fila de TOTAL de la tabla diaria (la que tiene la etiqueta 'TOTAL' en su
    columna de texto). Ahí están las celdas que la usuaria mapea (ej. G35). Si no
    se encuentra, cae a header_row+31 (31 días) como heurística.
    """
    for r in range(header_row + 1, min(header_row + 60, len(grid))):
        if "total" == _norm(_row_label(grid[r])):
            return r
    return min(header_row + 31, len(grid) - 1)


def _parse_ingresos(grid: list[list], warnings: list[str]) -> dict:
    """
    Tabla de generación diaria + venta por comercializador. Cada columna "Venta ($)"
    es una FUENTE de ingreso independiente (Terpel 1, Terpel 2, …); se devuelve una
    por columna, con su etiqueta propuesta, su celda de la fila TOTAL (ej. G35) y su
    valor. Si hay bolsa, Venta/Compra en bolsa también son fuentes. La usuaria puede
    renombrar cada fuente (alias persistente) y el sistema lo recuerda.
    """
    # Localizar fila de encabezado que contenga "venta" en alguna celda.
    header_row = None
    for i, row in enumerate(grid[:40]):
        joined = " ".join(_norm(c) for c in row if c is not None)
        if "venta" in joined and ("$" in joined or "venta ($)" in joined or "venta($)" in joined):
            header_row = i
            break

    venta_cols: list[int] = []
    venta_bolsa_cols: list[int] = []
    compra_bolsa_cols: list[int] = []
    comercializador = None
    headers: dict[int, str] = {}

    if header_row is not None:
        for j, c in enumerate(grid[header_row]):
            h = _norm(c)
            if not h:
                continue
            headers[j] = _txt(c)
            if "compra" in h and "bolsa" in h:
                compra_bolsa_cols.append(j)
            elif "venta" in h and "bolsa" in h:
                venta_bolsa_cols.append(j)
            elif "venta" in h and ("$" in h or "cop" in h or "pes" in h):
                venta_cols.append(j)
            elif h == "venta" or h.startswith("venta "):
                venta_cols.append(j)
        for r in (header_row - 1, header_row):
            if r < 0:
                continue
            for c in grid[r]:
                t = _norm(c)
                for com in COMERCIALIZADORES:
                    if com in t:
                        comercializador = _txt(c)
                        break
                if comercializador:
                    break
            if comercializador:
                break

    total_row = _fila_total(grid, header_row) if header_row is not None else None

    def _valor_col(j: int) -> float:
        """Valor TOTAL de la columna j: la celda de la fila TOTAL, o la suma diaria."""
        if total_row is not None and j < len(grid[total_row]):
            v = _num(grid[total_row][j])
            if v is not None:
                return v
        total = 0.0
        for r in range(header_row + 1, min(header_row + 60, len(grid))):
            if r == total_row:
                continue
            row = grid[r]
            if j < len(row):
                v = _num(row[j])
                if v is not None:
                    total += v
        return total

    def _celda(j: int) -> str | None:
        return f"{_col_letter(j)}{total_row + 1}" if total_row is not None else None

    def _nombre_fuente(j: int) -> str:
        """Comercializador del encabezado de la columna (ej. 'Terpel'), sin ruido."""
        h = _norm(headers.get(j, ""))
        for tok in ("venta", "($)", "$", "(cop)", "cop", "pesos", "(kwh)", "kwh", "_x", "_y"):
            h = h.replace(tok, " ")
        nombre = " ".join(w for w in h.split() if w)
        return nombre.title() if nombre else (comercializador or "")

    if header_row is None or not venta_cols:
        warnings.append("No se detectó la columna 'Venta ($)'; ingreso bruto = 0.")

    detalle: list[dict] = []
    venta_total = 0.0
    multi = len(venta_cols) > 1
    for idx, j in enumerate(venta_cols, start=1):
        val = round(_valor_col(j), 2)
        venta_total += val
        nombre = _nombre_fuente(j)
        if multi:
            etiqueta = f"Ingreso Bruto {nombre} {idx}".strip()
        else:
            etiqueta = f"Ingreso Bruto {nombre}".strip() if nombre else "Ingreso Bruto"
        detalle.append({"concepto": etiqueta, "valor": val, "celda": _celda(j)})

    venta_bolsa = sum(round(_valor_col(j), 2) for j in venta_bolsa_cols)
    compra_bolsa = sum(round(_valor_col(j), 2) for j in compra_bolsa_cols)
    tiene_bolsa = bool(venta_bolsa_cols or compra_bolsa_cols) and (venta_bolsa != 0 or compra_bolsa != 0)
    if tiene_bolsa:
        if venta_bolsa:
            detalle.append({"concepto": "Venta en bolsa", "valor": round(venta_bolsa, 2),
                            "celda": _celda(venta_bolsa_cols[0]) if venta_bolsa_cols else None})
        if compra_bolsa:
            detalle.append({"concepto": "Compra en bolsa", "valor": -abs(round(compra_bolsa, 2)),
                            "celda": _celda(compra_bolsa_cols[0]) if compra_bolsa_cols else None})

    ingreso_bruto = venta_total
    total_ingresos = venta_total + venta_bolsa - compra_bolsa

    return {
        "comercializador": comercializador,
        "tiene_bolsa": tiene_bolsa,
        "ingreso_bruto": ingreso_bruto,
        "total_ingresos": total_ingresos,
        "venta_bolsa": venta_bolsa,
        "compra_bolsa": compra_bolsa,
        "ingresos_detalle": detalle,
    }


# ── NITRO: ingreso bruto / bolsa desde la hoja Summary ──────────────────────────

def _plant_token(nombre: str | None) -> str | None:
    """Última palabra significativa del nombre del proyecto (ej. 'cacica')."""
    if not nombre:
        return None
    toks = [t for t in _norm(nombre).split()
            if len(t) >= 4 and t not in ("solar", "minigranja", "granja")]
    return toks[-1] if toks else None


def _summary_nitro(grid: list[list], hoja: str, proyecto_nombre: str | None) -> dict:
    """
    De la hoja Summary, fila 'Total' del proyecto: Despacho/Ingreso Bruto, Ventas en
    bolsa y Compras en bolsa. Columnas localizadas por encabezado (despacho/ventas/
    compras $). Devuelve {clave: (valor, hoja, 'I6')} para 'ib'/'venta'/'compra'.
    """
    from openpyxl.utils import get_column_letter
    if not grid:
        return {}
    col_ib = col_vta = col_cmp = None
    for row in grid[:6]:
        for j, c in enumerate(row):
            h = _norm(c)
            if not h:
                continue
            if "despacho" in h and ("$" in h or "cop" in h) and col_ib is None:
                col_ib = j
            elif "venta" in h and ("$" in h or "cop" in h) and col_vta is None:
                col_vta = j
            elif "compra" in h and ("$" in h or "cop" in h) and col_cmp is None:
                col_cmp = j
        if col_ib is not None:
            break
    if col_ib is None:
        return {}
    tok = _plant_token(proyecto_nombre)
    fila = None
    for i, row in enumerate(grid):
        txt = " ".join(_norm(x) for x in row if isinstance(x, str))
        if "total" in txt and (tok is None or tok in txt):
            fila = i
            break
    if fila is None:
        return {}
    out: dict = {}
    for clave, j in (("ib", col_ib), ("venta", col_vta), ("compra", col_cmp)):
        if j is None or j >= len(grid[fila]):
            continue
        v = _num(grid[fila][j])
        if v is None:
            continue
        out[clave] = (v, hoja, f"{get_column_letter(j + 1)}{fila + 1}")
    return out


# ── Ingresos NEU / NITRO (sección "Ingresos y costos", NO la XM) ────────────────

def _ingresos_section_bounds(grid: list[list]) -> tuple[int | None, int | None]:
    """
    Acota la sección "Ingresos y costos" (la de conceptos de ingreso, NO la
    "Ingresos y costos XM"). Empieza en su encabezado y termina en la fila de
    "Total ingresos" (inclusive) o, si no la hay, al toparse con "Costos
    operativos" / "Comercialización XM".
    """
    start = None
    for i, row in enumerate(grid):
        nm = _norm(_row_label(row)).replace(".", "")
        if nm.startswith("ingresos y costos") and "xm" not in nm:
            start = i
            break
    if start is None:
        return None, None
    end = len(grid)
    for j in range(start + 1, len(grid)):
        nm = _norm(_row_label(grid[j])).replace(".", "")
        if not nm:
            continue
        if "total ingreso" in nm:
            end = j + 1  # incluir la fila de total
            break
        if "costos operativos" in nm or "ingresos y costos xm" in nm or "comercializacion xm" in nm:
            end = j
            break
    return start, end


def _buscar_concepto(grid: list[list], start: int, end: int, variantes: list[tuple]) -> float | None:
    """
    Primer valor (columna contigua) de la fila cuya etiqueta normalizada (sin
    puntos) contenga TODOS los tokens de alguna de las `variantes`. Recorre solo
    el bloque [start, end) para no leer conceptos de otras secciones.
    """
    for r in range(start, end):
        nm = _norm(_row_label(grid[r])).replace(".", "")
        if not nm:
            continue
        for tokens in variantes:
            if all(tok in nm for tok in tokens):
                return _value_after_label(grid[r])
    return None


def _parse_ingresos_neu(grid: list[list], warnings: list[str]) -> dict:
    """
    NEU (Baraya, El Son, Ibirico, Mapale, Puya). Ingresos =
      Despacho de energía + Ventas en bolsa − Compras en bolsa
      + Distribución Superávit (Redistribución de ingresos) + Ajuste.
    """
    start, end = _ingresos_section_bounds(grid)
    if start is None:
        warnings.append("NEU: no se encontró la sección 'Ingresos y costos'; ingresos = 0.")
        return _ingresos_vacio()

    despacho = _buscar_concepto(grid, start, end, [("despacho",)])
    ventas = _buscar_concepto(grid, start, end, [("ventas", "bolsa"), ("venta", "bolsa")])
    compras = _buscar_concepto(grid, start, end, [("compras", "bolsa"), ("compra", "bolsa")])
    distrib = _buscar_concepto(grid, start, end, [("distribucion", "superavit"), ("redistribucion",)])
    ajuste = _buscar_concepto(grid, start, end, [("ajuste",)])

    detalle: list[dict] = []
    if despacho is None:
        warnings.append("NEU: no se encontró 'Despacho de energía'.")
    else:
        detalle.append({"concepto": "Despacho de energía", "valor": round(despacho, 2)})
    if ventas:
        detalle.append({"concepto": "Venta en bolsa", "valor": round(ventas, 2)})
    if compras:
        detalle.append({"concepto": "Compra en bolsa", "valor": -abs(round(compras, 2))})
    if distrib:
        detalle.append({"concepto": "Distribución Superávit", "valor": round(distrib, 2)})
    if ajuste:
        detalle.append({"concepto": "Ajuste", "valor": round(ajuste, 2)})

    total = sum(d["valor"] for d in detalle)
    return {
        "comercializador": None,
        "tiene_bolsa": bool(ventas or compras),
        "ingreso_bruto": round(despacho or 0.0, 2),
        "total_ingresos": round(total, 2),
        "venta_bolsa": round(ventas or 0.0, 2),
        "compra_bolsa": round(compras or 0.0, 2),
        "ingresos_detalle": detalle,
    }


def _parse_ingresos_nitro(grid: list[list], warnings: list[str]) -> dict:
    """
    NITRO (La Cacica, Las Piloneras). Ingresos =
      Ingreso Bruto + Ventas en bolsa − Compras en bolsa + Comercialización.
    """
    start, end = _ingresos_section_bounds(grid)
    if start is None:
        warnings.append("NITRO: no se encontró la sección 'Ingresos y costos'; ingresos = 0.")
        return _ingresos_vacio()

    bruto = _buscar_concepto(grid, start, end, [("ingreso", "bruto")])
    ventas = _buscar_concepto(grid, start, end, [("ventas", "bolsa"), ("venta", "bolsa")])
    compras = _buscar_concepto(grid, start, end, [("compras", "bolsa"), ("compra", "bolsa")])
    comerc = _buscar_concepto(grid, start, end, [("comercializacion",)])

    detalle: list[dict] = []
    if bruto is None:
        warnings.append("NITRO: no se encontró 'Ingreso Bruto'.")
    else:
        detalle.append({"concepto": "Ingreso Bruto", "valor": round(bruto, 2)})
    if ventas:
        detalle.append({"concepto": "Venta en bolsa", "valor": round(ventas, 2)})
    if compras:
        detalle.append({"concepto": "Compra en bolsa", "valor": -abs(round(compras, 2))})
    if comerc:
        detalle.append({"concepto": "Comercialización", "valor": round(comerc, 2)})

    total = sum(d["valor"] for d in detalle)
    return {
        "comercializador": None,
        "tiene_bolsa": bool(ventas or compras),
        "ingreso_bruto": round(bruto or 0.0, 2),
        "total_ingresos": round(total, 2),
        "venta_bolsa": round(ventas or 0.0, 2),
        "compra_bolsa": round(compras or 0.0, 2),
        "ingresos_detalle": detalle,
    }


def _ingresos_vacio() -> dict:
    return {
        "comercializador": None, "tiene_bolsa": False,
        "ingreso_bruto": 0.0, "total_ingresos": 0.0,
        "venta_bolsa": 0.0, "compra_bolsa": 0.0, "ingresos_detalle": [],
    }


def _xm_concepto(nm: str) -> str | None:
    """
    Resuelve la etiqueta de un renglón de Comercialización XM a su concepto.
    `nm` viene normalizado y SIN puntos (para que 'i.v.a.' == 'iva').
    Distingue Generador/Comercializador a partir de la propia etiqueta.
    """
    gen = "generador" in nm
    com = "comercializador" in nm
    suf = " (Gen)" if gen else (" (Com)" if com else "")
    if "arranque y parada" in nm:
        return "Arranque y parada"
    if "cargo" in nm and "confiabilidad" in nm:
        return "Cargo por confiabilidad"
    if "fazni" in nm:
        return "Fazni"
    if "energia en bolsa" in nm:
        return "Energía en Bolsa" + suf
    if "iva" in nm:
        return "IVA " + ("Generador" if gen else ("Comercializador" if com else "")).strip()
    if "despacho" in nm and "cnd" in nm:
        return "Serv. Despacho CND" + suf
    if ("administracion sic" in nm) or ("admin" in nm and "sic" in nm):
        return "Serv. Admin SIC" + suf
    return None


def _parse_comercializacion(grid: list[list]) -> list[dict]:
    """
    Comercialización XM desglosada. La etiqueta está en una columna (C) y el
    valor TOTAL en COP en la columna inmediatamente siguiente (D). A la derecha
    hay una sección de tarifas (Tarifa plataforma/operación…) que NO debe leerse:
    por eso el valor se toma como el PRIMER numérico tras la etiqueta, nunca el
    último. Los renglones en 0 (p.ej. la versión Comercializador cuando solo
    aplica el Generador) se descartan para no duplicar conceptos.
    """
    # Acotar al bloque entre "Ingresos y costos XM" y "Total Comercialización".
    start = end = None
    for i, row in enumerate(grid):
        nm = _norm(_row_label(row)).replace(".", "")
        if start is None and "ingresos y costos xm" in nm:
            start = i
        elif start is not None and "total comercializa" in nm:
            end = i
            break
    rows = grid[start + 1:end] if (start is not None and end is not None) else grid

    out: list[dict] = []
    vistos: set[str] = set()
    for row in rows:
        etiqueta = _row_label(row)
        if not etiqueta:
            continue
        nm = _norm(etiqueta).replace(".", "")
        concepto = _xm_concepto(nm)
        if not concepto:
            continue
        val = _value_after_label(row)
        if val is None or round(val, 2) == 0:
            continue
        if concepto in vistos:
            continue
        vistos.add(concepto)
        out.append({"concepto": concepto, "valor": -abs(val)})
    return out


# Costos operativos buscados por etiqueta. iva=True → aplica IVA 19% como línea aparte.
# El orden importa: la primera clave que aparezca en la etiqueta gana. "cambio
# equipos de medida" va antes que cualquier clave genérica para no confundirse.
_COSTO_CONCEPTOS = [
    ("arrend", "Arrendamiento", False),
    ("arriendo", "Arrendamiento", False),
    ("fondo de mantenimiento", "Fondo de mantenimiento", True),
    ("fondo mantenimiento", "Fondo de mantenimiento", True),
    ("mantenimiento", "Mantenimiento", True),
    ("poliza", "Póliza", False),
    ("internet", "Servicio de Internet", True),
    ("servicios publicos", "Servicios públicos", False),
    ("servicio publico", "Servicios públicos", False),
    # Cambio de equipos de medida: es COSTO OPERATIVO. Antes no se capturaba en
    # ninguna sección, así que quedaba sin agrupar (Bug 4). Variantes de etiqueta.
    ("cambio equipos de medida", "Cambio equipos de medida", False),
    ("cambio de equipos de medida", "Cambio equipos de medida", False),
    ("cambio equipos medida", "Cambio equipos de medida", False),
]


def _parse_costos(grid: list[list]) -> list[dict]:
    """
    Costos operativos. IVA = 19% sobre Mantenimiento e Internet (siempre),
    se devuelve como flag `iva` por línea para que el frontend/loader lo expanda.
    """
    out: list[dict] = []
    vistos: set[str] = set()
    for row in grid:
        etiqueta = _row_label(row)
        if not etiqueta:
            continue
        ne = _norm(etiqueta)
        for key, label, aplica_iva in _COSTO_CONCEPTOS:
            if key in ne:
                if label in vistos:
                    continue
                # Igual que en XM: el valor en COP está en la columna contigua a la
                # etiqueta; tomar el primer numérico (no el último, que sería una tarifa).
                val = _value_after_label(row)
                if val is None or round(val, 2) == 0:
                    break
                vistos.add(label)
                out.append({"concepto": label, "valor": -abs(val), "iva": aplica_iva})
                break
    return out


def _buscar_etiqueta_valor(grid: list[list], variantes: list[tuple]) -> float | None:
    """
    Primer valor (columna contigua) de la fila cuya etiqueta normalizada (sin
    puntos) contenga TODOS los tokens de alguna variante. Igual que
    _buscar_concepto pero sin acotar a una sección. Para Representación/CGM, que
    en el ER son filas "Cobro OPEX: Representación/CGM" con la tarifa por kWh.
    """
    for row in grid:
        nm = _norm(_row_label(row)).replace(".", "")
        if not nm:
            continue
        for tokens in variantes:
            if all(tok in nm for tok in tokens):
                v = _value_after_label(row)
                if v is not None and round(v, 2) != 0:
                    return v
    return None


def _parse_kwh(grid: list[list]) -> float | None:
    """Energía generada (kWh) del período — para Representación/CGM (kWh×5)."""
    for row in grid:
        etiqueta = _row_label(row)
        ne = _norm(etiqueta)
        if not ne:
            continue
        if ("kwh" in ne or "energia" in ne or "generacion" in ne) and "bolsa" not in ne:
            if "total" in ne or "generad" in ne or "kwh" in ne:
                v = _row_value(row)
                if v and v > 0:
                    return v
    return None


def _row_label(row: list) -> str:
    """Primera celda de texto no numérico de la fila (la etiqueta del concepto)."""
    for c in row:
        if c is None:
            continue
        if isinstance(c, str) and c.strip() and _num(c) is None:
            return c.strip()
    return ""


def _row_value(row: list):
    """Último valor numérico de la fila (el monto del concepto)."""
    val = None
    for c in row:
        n = _num(c) if not isinstance(c, str) else (_num(c) if re.search(r"\d", str(c)) else None)
        if n is not None:
            val = n
    return val


def _value_after_label(row: list):
    """
    Primer valor numérico que aparece DESPUÉS de la celda de etiqueta (la primera
    celda de texto). En los ER el monto en COP está justo a la derecha del concepto
    (columna D si la etiqueta está en C); las columnas más a la derecha contienen
    tarifas/factores que no deben leerse.
    """
    label_idx = None
    for i, c in enumerate(row):
        if isinstance(c, str) and c.strip() and _num(c) is None:
            label_idx = i
            break
    if label_idx is None:
        return None
    for c in row[label_idx + 1:]:
        n = _num(c) if not isinstance(c, str) else (_num(c) if re.search(r"\d", str(c)) else None)
        if n is not None:
            return n
    return None

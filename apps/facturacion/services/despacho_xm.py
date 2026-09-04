"""Lectura del archivo de despachos de XM (`dspcttos_txf_MM.xlsx`).

Parseo puro: recibe los bytes del Excel y devuelve la energía mensual y diaria
por contrato. No toca la base — de eso se encarga `guardar_despacho`.

**El archivo se RECHAZA si no trae exactamente 24 columnas de hora.** Hay
archivos mal armados (p. ej. `dspcttos_txF` de abril) que mezclan energía y
PRECIO por hora, dejando 48 columnas con nombres tan corruptos que no se pueden
leer de forma fiable. Sumarlas daría un total inflado sin que nadie se enterase,
así que se prefiere el error explícito y pedir el archivo correcto.
"""

import io
import re
from datetime import date

PATRONES_HORA = (
    r"^DESP_HORA(\d{1,2})(?:\.\d+)?$",
    r"^H(\d{1,2})(?:\.\d+)?$",          # respaldo para formatos H01..H24
)
HORAS_ESPERADAS = 24


class ExcelInvalido(ValueError):
    pass


def _encabezado(fila, nombres: set[str]) -> int | None:
    for indice, celda in enumerate(fila):
        if isinstance(celda, str) and celda.strip().upper() in nombres:
            return indice
    return None


def _columnas_de_hora(encabezado) -> tuple[dict[int, int], int]:
    """`{hora: índice}` más cuántas columnas de hora se vieron en total.

    Los dos números se devuelven aparte a propósito: si difieren, el archivo
    trae columnas duplicadas y hay que rechazarlo.
    """
    for patron in PATRONES_HORA:
        encontradas: dict[int, int] = {}
        total = 0
        for indice, celda in enumerate(encabezado):
            if not isinstance(celda, str):
                continue
            hallado = re.match(patron, celda.strip().upper().replace(" ", ""))
            if not hallado:
                continue
            total += 1
            hora = int(hallado.group(1))
            if 1 <= hora <= HORAS_ESPERADAS and hora not in encontradas:
                encontradas[hora] = indice
        if encontradas:
            return encontradas, total
    return {}, 0


def _fecha(valor) -> date | None:
    if valor is None:
        return None
    if hasattr(valor, "date"):
        return valor.date()
    if hasattr(valor, "year"):
        return valor
    try:
        return date.fromisoformat(str(valor)[:10])
    except ValueError:
        return None


def leer(contenido: bytes) -> tuple[dict, dict]:
    """Devuelve (por contrato, por contrato y día).

    Por contrato: `{codigo: {kwh, vendedor, comprador, tipo, fechas}}`.
    Por día: `{(codigo, fecha): kwh}`.
    """
    import openpyxl

    try:
        libro = openpyxl.load_workbook(
            io.BytesIO(contenido), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ExcelInvalido(f"No se pudo leer el Excel: {exc}") from exc

    hoja = libro[libro.sheetnames[0]]
    filas = hoja.iter_rows(values_only=True)
    try:
        encabezado = list(next(filas))
    except StopIteration:
        raise ExcelInvalido("El archivo está vacío")

    i_contrato = _encabezado(encabezado, {"CONTRATO"})
    if i_contrato is None:
        raise ExcelInvalido("No encontré la columna CONTRATO en el archivo")

    i_vendedor = _encabezado(encabezado, {"VENDEDOR"})
    i_comprador = _encabezado(encabezado, {"COMPRADOR"})
    i_tipo = _encabezado(encabezado, {"TIPO"})
    i_fecha = _encabezado(
        encabezado, {"FECHADOCUMENTO", "FECHA DOCUMENTO", "FECHA"}
    )

    horas, columnas_vistas = _columnas_de_hora(encabezado)
    if not horas:
        raise ExcelInvalido(
            "No encontré columnas de horas (DESP_HORA 01..24) en el archivo."
        )
    if len(horas) != HORAS_ESPERADAS or columnas_vistas != HORAS_ESPERADAS:
        raise ExcelInvalido(
            f"El archivo parece mal armado: encontré {columnas_vistas} columnas "
            f"de hora y {len(horas)} horas distintas (deben ser 24 columnas = 24 "
            f"horas). Suele pasar cuando el Excel trae la energía y el PRECIO por "
            f"hora mezclados (ej. 'dspcttos_txF'). Sube el archivo de despacho "
            f"correcto (24 columnas DESP_HORA 01..24, ej. 'dspcttos_txr' / "
            f"'dspcttos_txf')."
        )
    indices_hora = [horas[h] for h in sorted(horas)]

    por_contrato: dict[str, dict] = {}
    por_dia: dict[tuple[str, date], float] = {}

    for fila in filas:
        if not fila or i_contrato >= len(fila) or fila[i_contrato] is None:
            continue
        crudo = fila[i_contrato]
        # Los códigos vienen a veces como float de Excel (89902.0).
        codigo = str(int(crudo)) if isinstance(crudo, float) else str(crudo).strip()

        kwh = sum(
            fila[i] for i in indices_hora
            if i < len(fila) and isinstance(fila[i], (int, float))
        )
        acumulado = por_contrato.setdefault(codigo, {
            "kwh": 0.0, "vendedor": None, "comprador": None,
            "tipo": None, "fechas": set(),
        })
        acumulado["kwh"] += kwh

        if i_fecha is not None and i_fecha < len(fila):
            fecha = _fecha(fila[i_fecha])
            if fecha is not None:
                acumulado["fechas"].add(fecha)
                clave = (codigo, fecha)
                por_dia[clave] = round(por_dia.get(clave, 0.0) + kwh, 4)

        for indice, campo in (
            (i_vendedor, "vendedor"), (i_comprador, "comprador"), (i_tipo, "tipo"),
        ):
            if indice is not None and indice < len(fila) and fila[indice]:
                acumulado[campo] = str(fila[indice]).strip()

    return por_contrato, por_dia

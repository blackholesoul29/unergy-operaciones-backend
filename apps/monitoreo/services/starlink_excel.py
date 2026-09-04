"""Construcción del libro Excel de la factura Starlink.

Movido tal cual desde `app/api/v1/starlink.py`: son ~120 líneas de formato de
celdas que no tienen nada que ver con HTTP. La vista solo lo llama y envuelve el
resultado en una respuesta.
"""

def construir_excel(items: list[dict], agrupado: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Colores
    HEADER_COLOR = "1F4E79"   # azul oscuro
    ROW_ALT      = "BDD7EE"   # azul claro
    TOTAL_COLOR  = "0D2137"   # azul muy oscuro
    WHITE        = "FFFFFF"
    HEADER_FONT  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    TOTAL_FONT   = Font(bold=True, color=WHITE, name="Calibri", size=10)
    NORMAL_FONT  = Font(name="Calibri", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left", vertical="center")

    thin_side = Side(style="thin", color="B8CCE4")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def money_fmt(ws, row, cols):
        for c in cols:
            ws.cell(row, c).number_format = '#,##0.00'

    wb = Workbook()

    # ── Hoja 1: Detalle ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Detalle"

    headers1 = ["Tipo", "Descripción", "Precio unitario", "Cantidad",
                "Total impuestos", "Sin IVA", "IVA", "Monto total"]
    money_cols1 = [3, 5, 6, 7, 8]   # columnas con formato dinero

    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(1, ci, h)
        cell.fill   = fill(HEADER_COLOR)
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border

    totals1 = {c: 0.0 for c in money_cols1}
    for ri, item in enumerate(items, 2):
        vals = [
            item["tipo"], item["descripcion"], item["precio_unitario"],
            item["cantidad"], item["total_impuestos"],
            item["sin_iva"], item["iva"], item["monto_total"],
        ]
        row_fill = fill(ROW_ALT) if ri % 2 == 0 else fill(WHITE)
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(ri, ci, v)
            cell.fill   = row_fill
            cell.font   = NORMAL_FONT
            cell.border = thin_border
            cell.alignment = CENTER if ci != 2 else LEFT
        money_fmt(ws1, ri, money_cols1)
        for c in money_cols1:
            totals1[c] += ws1.cell(ri, c).value or 0

    # Fila TOTAL hoja 1
    tr = len(items) + 2
    ws1.cell(tr, 1, "TOTAL").fill = fill(TOTAL_COLOR)
    ws1.cell(tr, 1).font = TOTAL_FONT
    ws1.cell(tr, 1).alignment = CENTER
    ws1.cell(tr, 2, "").fill = fill(TOTAL_COLOR)
    ws1.cell(tr, 3, "").fill = fill(TOTAL_COLOR)
    ws1.cell(tr, 4, "").fill = fill(TOTAL_COLOR)
    for c in money_cols1:
        cell = ws1.cell(tr, c, round(totals1[c], 2))
        cell.fill      = fill(TOTAL_COLOR)
        cell.font      = TOTAL_FONT
        cell.alignment = CENTER
        cell.number_format = '#,##0.00'
        cell.border    = thin_border
    for ci in range(1, len(headers1) + 1):
        ws1.cell(tr, ci).border = thin_border

    # Anchos automáticos hoja 1
    _auto_width(ws1, headers1)

    # ── Hoja 2: Agrupado ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Agrupado")
    headers2 = ["Descripción", "Cantidad total", "Precio unit. promedio",
                "Sin IVA", "IVA", "Monto total"]
    money_cols2 = [3, 4, 5, 6]

    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(1, ci, h)
        cell.fill   = fill(HEADER_COLOR)
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border

    totals2 = {c: 0.0 for c in money_cols2}
    for ri, item in enumerate(agrupado, 2):
        vals = [
            item["descripcion"], item["cantidad_total"],
            item["precio_unitario_promedio"],
            item["sin_iva"], item["iva"], item["monto_total"],
        ]
        row_fill = fill(ROW_ALT) if ri % 2 == 0 else fill(WHITE)
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(ri, ci, v)
            cell.fill   = row_fill
            cell.font   = NORMAL_FONT
            cell.border = thin_border
            cell.alignment = CENTER if ci != 1 else LEFT
        money_fmt(ws2, ri, money_cols2)
        for c in money_cols2:
            totals2[c] += ws2.cell(ri, c).value or 0

    # Fila TOTAL hoja 2
    tr2 = len(agrupado) + 2
    for ci in range(1, len(headers2) + 1):
        cell = ws2.cell(tr2, ci)
        cell.fill   = fill(TOTAL_COLOR)
        cell.font   = TOTAL_FONT
        cell.border = thin_border
        cell.alignment = CENTER
    ws2.cell(tr2, 1, "TOTAL")
    for c in money_cols2:
        cell = ws2.cell(tr2, c, round(totals2[c], 2))
        cell.number_format = '#,##0.00'

    _auto_width(ws2, headers2)

    return wb


def _auto_width(ws, headers):
    from openpyxl.utils import get_column_letter
    for ci, h in enumerate(headers, 1):
        max_len = len(h)
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

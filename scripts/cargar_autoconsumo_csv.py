#!/usr/bin/env python3
"""
Carga el panel de seguimiento contable de AUTOCONSUMO al backend de Unergy.

  - Sin inversionistas individuales (siempre Total / inversionista_id = null)
  - tipo_venta = "autoconsumo"
  - Solo filas Mandato. Filas Factura se OMITEN completamente.
  - Proyectos con Ingreso Bruto = 0 / None / #N/A se omiten
  - Lee hoja principal + hoja "*Data" complementaria por mes
  - Hipervínculos de col H leídos via cell.hyperlink.target (soporte_url)

Uso:
  python scripts/cargar_autoconsumo_csv.py \
      --xlsx data/2026_Autoconsumo_Panel_Seguimiento_Contable.xlsx \
      --hojas Enero Febrero Marzo Abril \
      --api-url https://backend-production-63d8.up.railway.app \
      --usuario TU_USUARIO@unergy.io \
      --password TU_PASSWORD \
      --limpiar
"""

import argparse, re, sys, unicodedata
from openpyxl import load_workbook
import requests

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

HOJA_MES = {
    "Enero":1,"Febrero":2,"Marzo":3,"Abril":4,"Mayo":5,"Junio":6,
    "Julio":7,"Agosto":8,"Septiembre":9,"Octubre":10,"Noviembre":11,"Diciembre":12,
}

TIPO_LINEA_MAP = [
    (r"ingreso bruto|^ingreso$",        "ingreso_bruto"),
    (r"inter[eé]s|interes",             "intereses"),
    (r"retenci[oó]n",                   "retencion_fuente"),
    (r"^ica$|industria.*comercio",      "ica_opex"),
    (r"iva",                            "iva"),
    (r"valor.*pagar|utilidad|ganancia", "valor_a_pagar"),
]

# Aliases: nombre en Excel (normalizado) → nombre_comercial exacto en DB
ALIAS_MAP = {
    "m.d.m. cientifica s.a.s":                                "MDM",
    "mdm cientifica s.a.s":                                   "MDM",
    "servicios industriales y metalmecanicos sas":            "Seridme",
    "sociedad educativa ngs sas":                             "Nuevo Gimnasio School",
    "the pub s. a. s.":                                       "Pola del Pub",
    "the pub s.a.s.":                                         "Pola del Pub",
    "dairy partners americas manufacturing colombia ltda":    "Nestlé",
    "corredor nino frank edwin":                              "IBES",
    "correa tobon alberto de jesus":                          "Los Coches",
    "centro comercial y empresarial obelisco - p.h.-":        "Obelisco",
    "urbanizacion arboleda de castilla propiedad horizontal": "Arboleda Castilla",
    "urbanizacion arboleda de castilla p.h.":                 "Arboleda Castilla",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _norm(s):
    s = unicodedata.normalize("NFD", str(s or "").lower().strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _tipo_linea(concepto):
    n = _norm(concepto)
    for patron, tipo in TIPO_LINEA_MAP:
        if re.search(patron, n):
            return tipo
    return "otro_ingreso"


def _valor(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r"[$,\s]", "", str(v).strip())
    if not s or s.startswith("#"): return None
    try: return float(s)
    except ValueError: return None


def _es_vacio(v):
    if v is None: return True
    if isinstance(v, str) and v.strip().startswith("#"): return True
    try: return float(str(v).replace(",", "").replace("$", "").strip()) == 0
    except ValueError: return True


def _parse_codigo(codigo):
    if codigo is None: return None
    try: return int(float(str(codigo).strip()))
    except (ValueError, TypeError): return None


def _celda_link(row, idx):
    """Lee el hyperlink.target de la celda en col idx si existe, si no None."""
    c = row[idx] if idx < len(row) else None
    if c is None: return None
    try:
        return c.hyperlink.target if c.hyperlink else None
    except Exception:
        return None


# ── Lectura del Excel ─────────────────────────────────────────────────────────

def leer_hoja_principal(sh):
    """Lee la hoja principal; accede a Cell objects para capturar hipervínculos."""
    filas = []
    for row in sh.iter_rows(min_row=2, max_row=sh.max_row):
        proy = str(row[0].value or "").strip()
        doc  = str(row[3].value or "").strip().title()
        conc = str(row[5].value or "").strip()
        if not proy or not conc or doc != "Mandato":
            continue
        total       = _valor(row[6].value)
        factura_val = str(row[7].value or "").strip()
        factura     = factura_val if factura_val and not factura_val.startswith("#") else None
        soporte_url = _celda_link(row, 7)   # URL real de Google Drive embebido
        consecutivo = _parse_codigo(row[9].value) if len(row) > 9 else None
        filas.append({
            "proyecto":    proy,
            "concepto":    conc,
            "total":       total,
            "factura":     factura,
            "soporte_url": soporte_url,
            "consecutivo": consecutivo,
        })
    return filas


def leer_hoja_data(sh):
    """Hoja complementaria '<Mes> Data' — nombres completos de proyectos."""
    filas = []
    for row in sh.iter_rows(min_row=2, max_row=sh.max_row, values_only=True):
        proy = str(row[0] or "").strip()
        conc = str(row[1] or "").strip()
        if not proy or not conc: continue
        if _norm(conc) in ("administracion", "total"): continue
        total   = _valor(row[2])
        factura = str(row[3] or "").strip()
        filas.append({
            "proyecto":    proy,
            "concepto":    conc,
            "total":       total,
            "factura":     factura if factura and not factura.startswith("#") else None,
            "soporte_url": None,
            "consecutivo": None,
        })
    return filas


def agrupar(filas):
    grupos = {}
    for f in filas:
        grupos.setdefault(f["proyecto"], []).append(f)
    omitidos = []
    for nombre in list(grupos.keys()):
        ing = next((f["total"] for f in grupos[nombre]
                    if _norm(f["concepto"]) == "ingreso bruto"), None)
        if _es_vacio(ing):
            omitidos.append(nombre)
            del grupos[nombre]
    if omitidos:
        print(f"  Omitidos sin ingreso ({len(omitidos)}):")
        for n in omitidos: print(f"    - {n}")
    return grupos


# ── Match de proyecto ─────────────────────────────────────────────────────────

def match_proyecto(proyectos_db, nombre):
    norm = _norm(nombre)
    # 1. Alias explícito
    if norm in ALIAS_MAP:
        alias_norm = _norm(ALIAS_MAP[norm])
        for p in proyectos_db:
            if _norm(p["nombre_comercial"]) == alias_norm: return p
    # 2. Exacto
    for p in proyectos_db:
        if _norm(p["nombre_comercial"]) == norm: return p
    # 3. Substring
    for p in proyectos_db:
        n = _norm(p["nombre_comercial"])
        if n in norm or norm in n: return p
    # 4. Palabras clave (>= 4 chars)
    palabras = [w for w in norm.split() if len(w) >= 4]
    for palabra in palabras:
        for p in proyectos_db:
            if palabra in _norm(p["nombre_comercial"]): return p
    return None


# ── API client ────────────────────────────────────────────────────────────────

class API:
    def __init__(self, base):
        self.base = base.rstrip("/")
        self.token = None

    def login(self, usuario, password):
        r = requests.post(f"{self.base}/api/v1/auth/token",
                          data={"username": usuario, "password": password})
        r.raise_for_status()
        self.token = r.json()["access_token"]

    def _h(self): return {"Authorization": f"Bearer {self.token}"}

    def get(self, path, **kw):
        r = requests.get(f"{self.base}{path}", headers=self._h(), **kw)
        r.raise_for_status()
        return r.json()

    def post(self, path, body):
        r = requests.post(f"{self.base}{path}", json=body, headers=self._h())
        r.raise_for_status()
        return r.json()

    def patch(self, path, body):
        r = requests.patch(f"{self.base}{path}", json=body, headers=self._h())
        r.raise_for_status()
        return r.json()

    def delete(self, path):
        r = requests.delete(f"{self.base}{path}", headers=self._h())
        if r.status_code not in (200, 204, 404): r.raise_for_status()

    def post_linea(self, path, body):
        """POST tolerante al 500 post-commit de Railway."""
        try:
            r = requests.post(f"{self.base}{path}", json=body, headers=self._h())
            if r.status_code == 500:
                print(f"     500 post-commit (INSERT OK, continua)")
                return {}
            r.raise_for_status()
            return r.json()
        except requests.HTTPError:
            return {}


# ── Limpiar liquidación ───────────────────────────────────────────────────────

def limpiar_liq(api, liq_id):
    print(f"    Limpiando liquidacion {liq_id}...")
    api.delete(f"/api/v1/liquidaciones/{liq_id}/limpiar")


# ── Carga de un mes ───────────────────────────────────────────────────────────

def cargar_mes(api, proyectos_db, grupos, periodo_date, do_limpiar, usuario_id, stats):
    for nombre_proy, filas in grupos.items():
        proy_db = match_proyecto(proyectos_db, nombre_proy)
        if not proy_db:
            stats["sin_match"].append(nombre_proy)
            print(f"  Sin match: '{nombre_proy}'")
            continue
        pid = proy_db["id"]
        print(f"  -> {nombre_proy[:50]}  (id={pid})")

        liq_id = None
        try:
            liq_resp = api.post("/api/v1/liquidaciones", {
                "proyecto_id": pid, "periodo": periodo_date,
                "tipo_venta": "autoconsumo", "generado_por_id": usuario_id,
            })
            liq_id = liq_resp["id"]
            stats["liq_nuevas"] += 1
        except requests.HTTPError as e:
            if e.response.status_code in (409, 422, 500):
                resp = api.get("/api/v1/liquidaciones",
                               params={"proyecto_id": pid, "size": 50})
                y, m = periodo_date[:7].split("-")
                liq_id = next(
                    (l["id"] for l in resp["items"]
                     if l["periodo"].startswith(f"{y}-{m.zfill(2)}")), None)
                if not liq_id:
                    print(f"    No se pudo obtener liquidacion")
                    continue
                stats["liq_existentes"] += 1
            else:
                raise

        if do_limpiar:
            limpiar_liq(api, liq_id)

        consecutivo = next(
            (f["consecutivo"] for f in filas
             if f.get("consecutivo") and _norm(f["concepto"]) == "ingreso bruto"), None)
        nro_factura = next(
            (f["factura"] for f in filas
             if f.get("factura") and _norm(f["concepto"]) == "ingreso bruto"), None)

        m_resp = api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos", {
            "tipo": "ingresos", "inversionista_id": None,
            "consecutivo": consecutivo, "pa_aplica": False,
        })
        mid = m_resp["id"]
        stats["mandatos"] += 1

        neto = None
        for orden, f in enumerate(filas):
            if _norm(f["concepto"]) == "valor a pagar":
                neto = f["total"]
                continue
            if f["total"] is None: continue
            api.post_linea(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mid}/lineas", {
                "tipo_linea":         _tipo_linea(f["concepto"]),
                "concepto":           f["concepto"],
                "valor_cop":          f["total"],
                "referencia_factura": f.get("factura") or nro_factura,
                "soporte_url":        f.get("soporte_url"),
                "orden":              orden,
            })
            stats["lineas"] += 1

        if neto is not None:
            try:
                api.patch(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mid}",
                          {"valor_neto_cop": neto})
                print(f"     valor_neto_cop = $ {neto:,.0f}")
            except Exception:
                pass

        stats["ok"] += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Carga autoconsumo desde Excel al sistema de liquidaciones"
    )
    p.add_argument("--xlsx",      required=True,  help="Ruta al Excel")
    p.add_argument("--hojas",     nargs="+",       required=True,
                   help="Hojas a procesar (ej: Enero Febrero Marzo Abril)")
    p.add_argument("--anio-base", type=int,        default=2026)
    p.add_argument("--api-url",   default="https://backend-production-63d8.up.railway.app")
    p.add_argument("--usuario",   required=True)
    p.add_argument("--password",  required=True)
    p.add_argument("--limpiar",   action="store_true",
                   help="Borra mandatos/costos/facturas antes de reimportar")
    p.add_argument("--dry-run",   action="store_true",
                   help="Imprime qué haría sin escribir nada")
    args = p.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)

    print(f"\nAutenticando en {args.api_url}...")
    api = API(args.api_url)
    api.login(args.usuario, args.password)
    me = api.get("/api/v1/auth/me")
    usuario_id = me["id"]
    print("Autenticado OK")

    proyectos_db = api.get("/api/v1/proyectos", params={"size": 500})["items"]
    print(f"{len(proyectos_db)} proyectos en DB\n")

    stats = {"ok": 0, "liq_nuevas": 0, "liq_existentes": 0,
             "mandatos": 0, "lineas": 0, "sin_match": []}

    for hoja in args.hojas:
        if hoja not in HOJA_MES:
            print(f"Hoja '{hoja}' no reconocida, omitiendo.")
            continue
        mes  = HOJA_MES[hoja]
        anio = 2025 if hoja == "Diciembre" else args.anio_base
        periodo_date = f"{anio}-{mes:02d}-01"

        print(f"\n{'='*50}")
        print(f"HOJA: {hoja}  ->  {periodo_date}")
        print(f"{'='*50}")

        if hoja not in wb.sheetnames:
            print(f"  Hoja '{hoja}' no encontrada en el Excel")
            continue

        filas_main = leer_hoja_principal(wb[hoja])
        grupos = agrupar(filas_main)
        print(f"  Principal: {len(grupos)} proyectos con ingreso")

        hoja_data = f"{hoja} Data"
        if hoja_data in wb.sheetnames:
            filas_data = leer_hoja_data(wb[hoja_data])
            grupos_data = agrupar(filas_data)
            norms_main = {_norm(k) for k in grupos}
            nuevos = sum(1 for n in grupos_data if _norm(n) not in norms_main)
            for nombre, filas in grupos_data.items():
                if _norm(nombre) not in norms_main:
                    grupos[nombre] = filas
            print(f"  Data:      +{nuevos} proyectos adicionales")

        print(f"  Total:     {len(grupos)} proyectos\n")

        if args.dry_run:
            print("  [DRY RUN]")
            for n in grupos: print(f"    -> {n}")
            continue

        cargar_mes(api, proyectos_db, grupos, periodo_date,
                   args.limpiar, usuario_id, stats)

    print(f"\n{'='*50}")
    print(f"RESUMEN FINAL")
    print(f"{'='*50}")
    print(f"Proyectos cargados:       {stats['ok']}")
    print(f"Liquidaciones nuevas:     {stats['liq_nuevas']}")
    print(f"Liquidaciones existentes: {stats['liq_existentes']}")
    print(f"Mandatos creados:         {stats['mandatos']}")
    print(f"Lineas creadas:           {stats['lineas']}")
    if stats["sin_match"]:
        print(f"\nSin match en DB ({len(stats['sin_match'])}):")
        for n in stats["sin_match"]: print(f"  - {n}")
    else:
        print("\nTodos los proyectos encontrados en DB.")


if __name__ == "__main__":
    main()

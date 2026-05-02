"""
Carga el Diccionario de contratos de GESCON.xlsx en gescon_diccionario_contratos.
Uso: python scripts/cargar_diccionario.py [--dry-run]
"""
import sys
import requests
import openpyxl

XLSX = r"C:\Users\juan_\OneDrive\Documentos\Dev\cumplimiento\GESCON\GESCON.xlsx"
BASE_URL = "https://backend-production-63d8.up.railway.app/api/v1"
EMAIL    = "juanjose@unergy.io"
PASSWORD = "Unergy2025!"
DRY_RUN  = "--dry-run" in sys.argv


def login():
    r = requests.post(
        f"{BASE_URL}/auth/token",
        data={"username": EMAIL, "password": PASSWORD},
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def main():
    print("[1/3] Autenticando...")
    token = login()
    headers = {"Authorization": f"Bearer {token}"}

    print("[2/3] Leyendo Diccionario de contratos...")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Diccionario de contratos"]
    # Row 1 = header ("Contrato", ""), data from row 2
    # Col 1 = codigo_contrato, Col 2 = nombre
    filas = [(ws.cell(r, 1).value, ws.cell(r, 2).value)
             for r in range(2, ws.max_row + 1)
             if ws.cell(r, 1).value]
    print(f"      {len(filas)} entradas")

    ok = err = 0
    print("[3/3] Insertando...")
    for codigo, nombre in filas:
        payload = {
            "codigo_contrato": str(codigo).strip(),
            "nombre": str(nombre).strip() if nombre else None,
        }

        if DRY_RUN:
            print(f"  [DRY] {payload['codigo_contrato']} -> {payload['nombre']}")
            ok += 1
            continue

        try:
            resp = requests.post(f"{BASE_URL}/asic/gescon/diccionario", json=payload, headers=headers, timeout=15)
            resp.raise_for_status()
            ok += 1
        except Exception as e:
            body = ""
            try:
                body = resp.text[:200]
            except Exception:
                pass
            print(f"  [ERR] {codigo}: {e} | {body}")
            err += 1

    print(f"\n[DONE] OK={ok}  ERR={err}")


if __name__ == "__main__":
    main()

"""
Carga las Fronteras comerciales de GESCON.xlsx en la tabla fronteras.
Hace upsert por codigo_frontera. Intenta asociar proyecto_id por nombre.
Uso: python scripts/cargar_fronteras_gescon.py [--dry-run]
"""
import sys, os, unicodedata, re
from datetime import datetime
import requests
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app.utils.nombre_matching import mejor_candidato

XLSX = r"C:\Users\juan_\OneDrive\Documentos\Dev\cumplimiento\GESCON\GESCON.xlsx"
BASE_URL = "https://backend-production-63d8.up.railway.app/api/v1"
from _creds import api_credentials

EMAIL, PASSWORD = api_credentials()
DRY_RUN  = "--dry-run" in sys.argv

TIPO_MAP = {
    "generacion": "generacion",
    "consumo auxiliar": "consumo_auxiliar",
    "consumo propio": "consumo_propio",
    "consumo": "consumo",
    "generacion consumo": "generacion_consumo",
}


def norm(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip().lower()


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def safe_int(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def safe_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()[:10]
    return s if len(s) == 10 else None


def safe_bool(v):
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("si", "sí", "yes", "true", "1", "s")


def login():
    r = requests.post(
        f"{BASE_URL}/auth/token",
        data={"username": EMAIL, "password": PASSWORD},
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def cargar_proyectos(token):
    headers = {"Authorization": f"Bearer {token}"}
    page, items = 1, []
    while True:
        r = requests.get(f"{BASE_URL}/proyectos", params={"page": page, "size": 200}, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()
        batch = data["items"] if isinstance(data, dict) else data
        if not batch:
            break
        items.extend(batch)
        if isinstance(data, dict) and len(items) >= data.get("total", 0):
            break
        page += 1
    return items


# ── Matching frontera -> proyecto ──────────────────────────────────────────────
# El algoritmo real vive en app/utils/nombre_matching.py, compartido con
# app/utils/proyecto_matching.py (find_proyecto_by_name, usada en producción)
# para no mantener dos implementaciones del mismo problema. La versión anterior
# de match_proyecto() comparaba el nombre completo con SequenceMatcher (sin quitar
# prefijos como "GD"/"Minigranja Solar"/"MGS") y tomaba el score más alto sin
# verificar si había un segundo candidato casi igual de bueno -- eso fue la causa
# de varios mismatches contra producción (Cañahuate, El Molino, Gandalf, Perijá,
# El Son, Naos 2/3), corregidos en 2026-07-02.
def match_proyecto(nombre_frontera, codigo_propio, proyectos):
    """Encuentra el proyecto correcto para una frontera. Returns (proyecto_id, match_reason) or (None, None)."""
    candidatos = [(p["id"], [p.get("nombre_comercial"), p.get("nombre_bitacora")]) for p in proyectos]

    # 1. Nombre exacto (normalizado)
    n_frontera = norm(nombre_frontera)
    for pid, nombres in candidatos:
        for n in nombres:
            if n and norm(n) == n_frontera:
                return pid, "nombre exacto"

    # 2. codigo_propio exacto contra el nombre del proyecto
    if codigo_propio:
        n_codigo = norm(codigo_propio)
        for pid, nombres in candidatos:
            for n in nombres:
                if n and norm(n) == n_codigo:
                    return pid, "codigo_propio exacto"

    # 3. Scoring por tokens + similitud, con desambiguación (ver nombre_matching.py)
    mejor_id, mejor_score = mejor_candidato(nombre_frontera, candidatos)
    if mejor_id is None:
        return None, None
    return mejor_id, f"fuzzy {mejor_score:.2f}"


def main():
    print("[1/4] Autenticando...")
    token = login()
    headers = {"Authorization": f"Bearer {token}"}

    print("[2/4] Cargando proyectos...")
    proyectos = cargar_proyectos(token)
    print(f"      {len(proyectos)} proyectos")

    print("[3/4] Leyendo Fronteras comerciales...")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Fronteras comerciales"]
    filas = [list(ws.cell(r, c).value for c in range(1, 75))
             for r in range(2, ws.max_row + 1)
             if ws.cell(r, 1).value is not None]
    print(f"      {len(filas)} fronteras")

    ok = err = 0
    sin_proyecto = []

    print("[4/4] Insertando...")
    for i, cols in enumerate(filas):
        # Map columns (0-indexed)
        (codigo_sic, nombre, codigo_propio, registrada_por, fecha_prim_reg,
         factor_perdidas, niu, nit, voltaje_kv, nivel_tension,
         tipo_punto_med, transf_max, representante, f_inicio_rep, rep_anterior,
         op_red, op_red_zona, ag_exportador, ag_importador, tipo_frontera,
         cap_transporte, cap_transporte_comp, cap_efectiva, nombre_cgm, sic_ddv,
         predio_id, nombre_predio, rep_ddv,
         nro_ppal, marca_ppal, modelo_ppal, clase_medidor, num_elem_ppal,
         clase_ct, clase_pt, fecha_cambio_ppal, entidad_cal_ppal, fecha_cal_ppal, fecha_act_ppal,
         nro_resp, marca_resp, modelo_resp, num_elem_resp,
         fecha_cambio_resp, entidad_cal_resp, fecha_cal_resp, fecha_act_resp,
         es_agrupadora, factor_psf, agrupada_bajo,
         es_principal_emb, embebida_bajo, factor_acordado, factor_ajuste, fp_principal,
         codigo_ciiu, clasif_ind_gen, clasif_ind_esp,
         departamento, municipio, centro_poblado, direccion,
         latitud, longitud, altitud,
         sic_sub_export, sic_frt_gen, nombre_recurso, clasif_recurso,
         potencia_max, tipo_tecnologia, sic_frt_usuario, sic_sub_consumo, sic_sub_usuario
         ) = (cols + [None]*74)[:74]

        tipo_mapped = TIPO_MAP.get(norm(tipo_frontera), "generacion")

        proyecto_id, match_reason = match_proyecto(str(nombre or ""), str(codigo_propio or ""), proyectos)
        if proyecto_id is None:
            sin_proyecto.append({"codigo": str(codigo_sic), "nombre": str(nombre), "codigo_propio": str(codigo_propio or "")})

        payload = {
            "codigo_frontera": str(codigo_sic).strip() if codigo_sic else None,
            "nombre_frontera": str(nombre).strip() if nombre else "SIN NOMBRE",
            "codigo_propio": str(codigo_propio).strip() if codigo_propio else None,
            "tipo_frontera": tipo_mapped,
            "estado": "activa",
            "proyecto_id": proyecto_id,
            "registrada_por": str(registrada_por).strip() if registrada_por else None,
            "fecha_primer_registro_asic": safe_date(fecha_prim_reg),
            "factor_perdidas": safe_float(factor_perdidas),
            "niu": str(niu).strip() if niu and str(niu).strip() else None,
            "nit": str(nit).strip() if nit and str(nit).strip() not in ("NA", "N/A", "") else None,
            "nivel_tension_kv": safe_float(voltaje_kv),
            "nivel_tension": safe_int(nivel_tension),
            "tipo_punto_medicion": safe_int(tipo_punto_med),
            "transferencia_maxima_kwh": safe_float(transf_max),
            "representante_frontera": str(representante).strip() if representante else None,
            "fecha_inicio_representacion": safe_date(f_inicio_rep),
            "representante_anterior": str(rep_anterior).strip() if rep_anterior else None,
            "operador_red": str(op_red).strip() if op_red else None,
            "operador_red_zona": str(op_red_zona).strip() if op_red_zona else None,
            "agente_exportador": str(ag_exportador).strip() if ag_exportador else None,
            "agente_importador": str(ag_importador).strip() if ag_importador else None,
            "capacidad_transporte_mw": safe_float(cap_transporte),
            "capacidad_transporte_compartida_mw": safe_float(cap_transporte_comp),
            "capacidad_efectiva_mw": safe_float(cap_efectiva),
            "nombre_cgm": str(nombre_cgm).strip() if nombre_cgm else None,
            "codigo_sic_ddv": str(sic_ddv).strip() if sic_ddv else None,
            "predio_id": str(predio_id).strip() if predio_id else None,
            "nombre_predio": str(nombre_predio).strip() if nombre_predio else None,
            "representante_ddv": str(rep_ddv).strip() if rep_ddv else None,
            # Medidor principal
            "nro_serie_med_ppal": str(nro_ppal).strip() if nro_ppal else None,
            "marca_med_ppal": str(marca_ppal).strip() if marca_ppal else None,
            "modelo_med_ppal": str(modelo_ppal).strip() if modelo_ppal else None,
            "clase_medidor": str(clase_medidor).strip() if clase_medidor else None,
            "num_elementos_med_ppal": safe_int(num_elem_ppal),
            "clase_ct": str(clase_ct).strip() if clase_ct else None,
            "clase_pt": str(clase_pt).strip() if clase_pt else None,
            "fecha_cambio_med_ppal": safe_date(fecha_cambio_ppal),
            "entidad_calibradora_med_ppal": str(entidad_cal_ppal).strip() if entidad_cal_ppal else None,
            "fecha_calibracion_med_ppal": safe_date(fecha_cal_ppal),
            "fecha_actualizacion_ppal": safe_date(fecha_act_ppal),
            # Medidor respaldo
            "nro_serie_med_resp": str(nro_resp).strip() if nro_resp else None,
            "marca_med_resp": str(marca_resp).strip() if marca_resp else None,
            "modelo_med_resp": str(modelo_resp).strip() if modelo_resp else None,
            "num_elementos_med_resp": safe_int(num_elem_resp),
            "fecha_cambio_med_resp": safe_date(fecha_cambio_resp),
            "entidad_calibradora_med_resp": str(entidad_cal_resp).strip() if entidad_cal_resp else None,
            "fecha_calibracion_med_resp": safe_date(fecha_cal_resp),
            "fecha_actualizacion_resp": safe_date(fecha_act_resp),
            # Agrupación
            "es_agrupadora": safe_bool(es_agrupadora),
            "factor_psf": safe_float(factor_psf),
            "agrupada_bajo": str(agrupada_bajo).strip() if agrupada_bajo else None,
            "es_principal_embebido": safe_bool(es_principal_emb),
            "embebida_bajo": str(embebida_bajo).strip() if embebida_bajo else None,
            "factor_acordado": safe_float(factor_acordado),
            "factor_ajuste": safe_float(factor_ajuste),
            "factor_perdidas_frontera_principal": safe_float(fp_principal),
            # Clasificación industrial
            "codigo_ciiu": str(codigo_ciiu).strip() if codigo_ciiu else None,
            "clasificacion_industrial_general": str(clasif_ind_gen).strip() if clasif_ind_gen else None,
            "clasificacion_industrial_especifica": str(clasif_ind_esp).strip() if clasif_ind_esp else None,
            # Ubicación
            "departamento": str(departamento).strip() if departamento else None,
            "municipio": str(municipio).strip() if municipio else None,
            "centro_poblado": str(centro_poblado).strip() if centro_poblado else None,
            "direccion": str(direccion).strip() if direccion else None,
            "latitud": safe_float(latitud),
            "longitud": safe_float(longitud),
            "altitud_msnm": safe_int(altitud),
            # Códigos SIC adicionales
            "codigo_sic_submercado_exportador": str(sic_sub_export).strip() if sic_sub_export else None,
            "codigo_sic_frontera_generacion": str(sic_frt_gen).strip() if sic_frt_gen else None,
            "nombre_recurso_generacion": str(nombre_recurso).strip() if nombre_recurso else None,
            "clasificacion_recurso": str(clasif_recurso).strip() if clasif_recurso else None,
            "potencia_maxima_declarada": safe_float(potencia_max),
            "tipo_tecnologia": str(tipo_tecnologia).strip() if tipo_tecnologia else None,
            "codigo_sic_frontera_usuario": str(sic_frt_usuario).strip() if sic_frt_usuario else None,
            "codigo_sic_submercado_consumo": str(sic_sub_consumo).strip() if sic_sub_consumo else None,
            "codigo_sic_submercado_usuario": str(sic_sub_usuario).strip() if sic_sub_usuario else None,
        }

        if DRY_RUN:
            proyecto_str = f"proyecto_id={proyecto_id} ({match_reason})" if proyecto_id else "SIN PROYECTO"
            print(f"  [DRY] {codigo_sic} | {nombre} | {tipo_mapped} | {proyecto_str}")
            ok += 1
            continue

        try:
            resp = requests.post(f"{BASE_URL}/fronteras", json=payload, headers=headers, timeout=15)
            resp.raise_for_status()
            ok += 1
            if (i + 1) % 20 == 0:
                print(f"  ... {i+1}/{len(filas)} procesadas")
        except Exception as e:
            body = ""
            try:
                body = resp.text[:300]
            except Exception:
                pass
            print(f"  [ERR] {codigo_sic} {nombre}: {e} | {body}")
            err += 1

    print(f"\n[DONE] OK={ok}  ERR={err}")
    print(f"\nFronteras SIN proyecto ({len(sin_proyecto)}):")
    for f in sin_proyecto:
        print(f"  - {f['codigo']} | {f['nombre']} | cp={f['codigo_propio']}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Carga el panel de seguimiento contable (XLSX) al sistema de liquidaciones.
Lee hipervínculos directamente desde Excel para obtener links a Drive.

Uso:
  python scripts/cargar_liquidaciones.py \\
      --xlsx "2026_Panel de seguimiento contable.xlsx" \\
      --hoja Enero \\
      --periodo 2026-01 \\
      --usuario admin@unergy.co \\
      --password TU_PASSWORD \\
      --limpiar          # ← borra mandatos/costos/facturas antes de cargar

Flags:
  --limpiar    Elimina todos los mandatos, costos y facturas de la
               liquidación antes de cargar. Úsalo si el script corrió
               antes y hay datos duplicados. La liquidación en sí NO
               se borra (conserva id, estado, período).
  --dry-run    Solo imprime lo que haría, sin escribir nada en la API.
  --api-url    URL del backend (default: https://backend-production-63d8.up.railway.app)

Dependencias:
  pip install openpyxl requests
"""

import argparse
import re
import sys
import unicodedata

# Windows: forzar UTF-8 en la salida estándar y deshabilitar buffering
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stdout.reconfigure(line_buffering=True)  # flush en cada línea (visible en background)

from openpyxl import load_workbook
import requests


# ─────────────────────────────────────────────────────────────────────────────
# Mapa concepto → tipo_linea
# ─────────────────────────────────────────────────────────────────────────────
TIPO_LINEA_MAP = [
    (r"ingreso bruto",                          "ingreso_bruto"),
    (r"^ingreso$",                              "ingreso_bruto"),
    (r"despacho",                               "despacho"),
    (r"ventas en bolsa",                        "ventas_en_bolsa"),
    (r"compras en bolsa|compras a ",            "compras_en_bolsa"),
    (r"redistribuci",                           "redistribucion_ingresos"),
    (r"ajuste.*xm|xm.*ajuste",                 "ajuste_xm"),
    (r"ajuste.*unergy",                         "ajuste_unergy"),
    (r"ajuste.*administr",                      "otro_costo"),
    (r"comercializaci",                         "ajuste_comercializacion"),
    (r"arriendo",                               "arriendo"),
    (r"mantenimiento",                          "mantenimiento"),
    (r"iva.*internet|internet.*iva",            "iva"),          # antes que regla genérica
    (r"internet",                               "servicio_internet"),
    (r"poliza.*cumplimiento|p.liza.*cumpl",     "poliza_cumplimiento"),
    (r"poliza|incendio|lucro cesante",          "seguro"),
    (r"servicios p.blicos|consumo de energ",    "servicios_publicos_consumo"),
    (r"cambio.*equipo|equipo.*medida",          "cambio_equipos_medida"),
    (r"representaci",                           "representacion"),
    (r"^cgm",                                   "cgm"),
    (r"administraci|valor final",               "administracion"),
    (r"reteica",                                "reteica"),
    (r"ica opex",                               "ica_opex"),
    (r"iva|i\.v\.a",                            "iva"),
    (r"retenci",                                "retencion_fuente"),
    (r"porcentaje.*participaci",                "porcentaje_participacion"),
    (r"valor.*pagar|valor.*ganancia|utilidad",  "valor_a_pagar"),
    (r"intereses",                              "intereses"),
]

# Tipos de Documento contable válidos → forma canónica
DOC_MAP = {
    "informacion":  "Información",
    "información":  "Información",
    "mandato":      "Mandato",
    "costos":       "Costos",
    "factura":      "Factura",
    "buenaventura": "Factura",  # alias → Factura
}


def concepto_a_tipo_factura(concepto: str) -> str:
    n = _norm(concepto)
    if "representaci" in n:
        return "representacion"
    if "cgm" in n:
        return "cgm"
    return "administracion_operacion"


def concepto_a_tipo_costo(concepto: str) -> str:
    """Mapea concepto → TipoCostoEnum (valores usados en liquidacion_costos)."""
    n = _norm(concepto)
    if "arriendo" in n:                    return "arriendo"
    if "mantenimiento" in n:               return "mantenimiento"
    if "internet" in n:                    return "internet"
    if "poliza" in n:                      return "polizas"
    if "servicios p" in n:                 return "servicios_publicos"
    if "cambio" in n and "equipo" in n:    return "cambio_equipos_medida"
    if "comercializaci" in n:              return "comercializacion_xm"
    return "otro"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or "").lower().strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _doc(s: str) -> str | None:
    return DOC_MAP.get(_norm(s), None)


def _tipo_linea(concepto: str) -> str:
    n = _norm(concepto)
    for patron, tipo in TIPO_LINEA_MAP:
        if re.search(patron, n):
            return tipo
    return "otro_ingreso"


def _valor(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[$,\s]", "", str(v).strip())
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _celda_val(row, idx):
    return row[idx].value if idx < len(row) else None


def _celda_link(row, idx) -> str | None:
    c = row[idx] if idx < len(row) else None
    if c and hasattr(c, "hyperlink") and c.hyperlink:
        return c.hyperlink.target
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Leer hoja XLSX
# ─────────────────────────────────────────────────────────────────────────────
def leer_hoja(xlsx_path: str, hoja: str) -> tuple[list[dict], dict[str, str]]:
    """
    Devuelve (filas, er_map).

    Columnas esperadas (0-based):
      0  = Proyecto
      1  = Inversionista
      2  = Doc. Contable
      5  = Concepto
      6  = Total
      7  = Ref. Factura / Soporte
      8  = Consecutivo Ingresos / URL soporte
      9  = Consecutivo Costos
      10 = Comprobante Contable
      13 = Proyecto (tabla lateral ER)
      14 = URL Estado de Resultados
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if hoja not in wb.sheetnames:
        raise ValueError(f"Hoja '{hoja}' no encontrada. Disponibles: {wb.sheetnames}")
    sh = wb[hoja]

    filas: list[dict] = []
    er_map: dict[str, str] = {}

    for xl_row in sh.iter_rows(min_row=2, max_row=sh.max_row):
        # Tabla lateral: col N (13) = proyecto, col O (14) = URL estado de resultados
        if len(xl_row) > 14:
            n_val = _celda_val(xl_row, 13)
            o_cell = xl_row[14]
            if n_val:
                er_url = (o_cell.hyperlink.target if o_cell.hyperlink else None) \
                         or str(o_cell.value or "").strip()
                if er_url and er_url.startswith("http"):
                    er_map[_norm(str(n_val).strip())] = er_url

        # Tabla principal: col A (0) = proyecto (requerido)
        proy = _celda_val(xl_row, 0)
        if not proy:
            continue

        # Excluir proyectos "trading" (actividades de bolsa propias de Unergy)
        if "trading" in _norm(str(proy)):
            continue

        doc_raw = str(_celda_val(xl_row, 2) or "").strip()
        doc_tipo = _doc(doc_raw)
        if not doc_tipo:
            continue  # tipo desconocido → omitir

        filas.append({
            "proyecto":        str(proy).strip(),
            "inversionista":   str(_celda_val(xl_row, 1) or "").strip(),
            "doc_contable":    doc_tipo,
            "contato1":        str(_celda_val(xl_row, 3) or "").strip(),
            "contato2":        str(_celda_val(xl_row, 4) or "").strip(),
            "concepto":        str(_celda_val(xl_row, 5) or "").strip(),
            "total":           _valor(_celda_val(xl_row, 6)),
            "ref_factura":     str(_celda_val(xl_row, 7) or "").strip(),
            "ref_factura_url": _celda_link(xl_row, 7),
            "cons_ing_txt":    str(_celda_val(xl_row, 8) or "").strip(),
            "cons_ing_url":    _celda_link(xl_row, 8),
            "cons_cos_txt":    str(_celda_val(xl_row, 9) or "").strip() if len(xl_row) > 9 else "",
            "comprobante":     str(_celda_val(xl_row, 10) or "").strip() if len(xl_row) > 10 else "",
        })

    wb.close()
    return filas, er_map


# ─────────────────────────────────────────────────────────────────────────────
# Cliente HTTP
# ─────────────────────────────────────────────────────────────────────────────
_RETRY_ON = {500, 502, 503, 504}  # crashes transitorios de Railway
_MAX_RETRIES = 5
_RETRY_WAIT = 20  # espera base en segundos (Railway tarda ~15s en recuperarse)
_CALL_PAUSE = 0.2  # pausa mínima entre llamadas para no saturar el worker


def _retry(fn, *args, **kwargs):
    """Reintenta una llamada HTTP si Railway devuelve un error transitorio."""
    import time
    time.sleep(_CALL_PAUSE)  # throttle básico entre llamadas
    for attempt in range(_MAX_RETRIES):
        r = fn(*args, **kwargs)
        if r.status_code not in _RETRY_ON:
            return r
        wait = _RETRY_WAIT * (attempt + 1)
        print(f"  ⚡ {r.status_code} — reintentando en {wait}s... ({attempt+1}/{_MAX_RETRIES})")
        time.sleep(wait)
    return r  # devuelve el último aunque siga fallando


class API:
    def __init__(self, base: str):
        self.base = base.rstrip("/")
        self.token: str | None = None

    def login(self, usuario: str, password: str):
        r = requests.post(
            f"{self.base}/api/v1/auth/token",
            data={"username": usuario, "password": password},
        )
        r.raise_for_status()
        self.token = r.json()["access_token"]

    def _h(self) -> dict:
        return {"Authorization": f"Bearer {self.token}"}

    def get(self, path: str, **kw):
        r = _retry(requests.get, f"{self.base}{path}", headers=self._h(), **kw)
        r.raise_for_status()
        return r.json()

    def post(self, path: str, body: dict):
        r = _retry(requests.post, f"{self.base}{path}", json=body, headers=self._h())
        if not r.ok:
            print(f"  ✗ POST {path} → {r.status_code}: {r.text[:400]}")
        r.raise_for_status()
        return r.json()

    def patch(self, path: str, body: dict):
        r = _retry(requests.patch, f"{self.base}{path}", json=body, headers=self._h())
        r.raise_for_status()
        return r.json()

    def delete(self, path: str):
        r = _retry(requests.delete, f"{self.base}{path}", headers=self._h())
        if r.status_code not in (200, 204, 404):
            r.raise_for_status()


# ─────────────────────────────────────────────────────────────────────────────
# Match helpers
# ─────────────────────────────────────────────────────────────────────────────
def match_proyecto(proyectos_db: list, nombre: str) -> dict | None:
    norm = _norm(nombre)
    # Coincidencia exacta
    for p in proyectos_db:
        if _norm(p["nombre_comercial"]) == norm:
            return p
    # Contención
    for p in proyectos_db:
        n = _norm(p["nombre_comercial"])
        if n in norm or norm in n:
            return p
    # Por palabra significativa
    for parte in reversed(norm.split()):
        if len(parte) < 4:
            continue
        for p in proyectos_db:
            if parte in _norm(p["nombre_comercial"]):
                return p
    return None


def match_inversionista(inversionistas_db: list, nombre: str) -> dict | None:
    if not nombre or nombre.upper() == "TOTAL":
        return None
    norm = _norm(nombre)
    for inv in inversionistas_db:
        inv_nombre = inv.get("inversionista_nombre") or inv.get("cliente_nombre") or ""
        cn = _norm(inv_nombre)
        if not cn:
            continue
        if cn == norm:
            return inv
        if cn in norm or norm in cn:
            return inv
        palabras = [w for w in norm.split() if len(w) > 3]
        if palabras and all(w in cn for w in palabras[:2]):
            return inv
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Limpiar liquidación
# ─────────────────────────────────────────────────────────────────────────────
def limpiar_liquidacion(api: API, liq_id: int):
    """Borra mandatos, costos y facturas. La liquidación en sí NO se toca."""
    print(f"  🧹 Limpiando liquidación {liq_id}...")
    detalle = api.get(f"/api/v1/liquidaciones/{liq_id}")

    mandatos = detalle.get("mandatos", [])
    costos   = detalle.get("costos", [])
    facturas = detalle.get("facturas", [])

    for m in mandatos:
        api.delete(f"/api/v1/liquidaciones/{liq_id}/mandatos/{m['id']}")
    for c in costos:
        api.delete(f"/api/v1/liquidaciones/{liq_id}/costos/{c['id']}")
    for f in facturas:
        api.delete(f"/api/v1/liquidaciones/{liq_id}/facturas/{f['id']}")

    print(f"     Eliminados: {len(mandatos)} mandatos, {len(costos)} costos, {len(facturas)} facturas")


# ─────────────────────────────────────────────────────────────────────────────
# Carga principal
# ─────────────────────────────────────────────────────────────────────────────
def cargar(
    api: API,
    filas: list[dict],
    er_map: dict[str, str],
    periodo_date: str,
    dry_run: bool,
    limpiar: bool,
):
    me = api.get("/api/v1/auth/me")
    usuario_id = me["id"]

    proyectos_db = api.get("/api/v1/proyectos", params={"size": 500})["items"]
    print(f"  {len(proyectos_db)} proyectos en DB\n")

    # Agrupar: proyecto → inversionista → filas
    grupos: dict[str, dict[str, list]] = {}
    for f in filas:
        grupos.setdefault(f["proyecto"], {}).setdefault(f["inversionista"], []).append(f)

    stats = {
        "ok": 0, "sin_match": [], "liq_nueva": 0, "liq_existente": 0,
        "mandatos": 0, "lineas": 0, "costos": 0, "facturas": 0,
        "duplicados": 0,
    }

    OMITIR = {"porcentaje de participacion", "valor a pagar"}

    # Rastrear proyecto_ids ya procesados para evitar cargar el mismo proyecto dos veces
    # (ocurre cuando el Excel tiene "GD NAOS 1", "GD NAOS 2", etc. todos → mismo proyecto DB)
    proy_ids_procesados: set[int] = set()

    for nombre_proy, inv_grupos in grupos.items():
        proy_db = match_proyecto(proyectos_db, nombre_proy)
        if not proy_db:
            stats["sin_match"].append(nombre_proy)
            print(f"  ⚠  Sin match en DB: '{nombre_proy}'")
            continue

        pid = proy_db["id"]

        # Evitar duplicados: si ya procesamos este proyecto en esta carga, omitir
        if pid in proy_ids_procesados:
            print(f"  ⤷ '{nombre_proy}' → duplicado de proyecto id={pid}, omitido")
            stats["duplicados"] += 1
            continue
        proy_ids_procesados.add(pid)
        er_url = er_map.get(_norm(nombre_proy))
        print(f"→ {nombre_proy}  (id={pid})")

        if dry_run:
            print(f"   [dry-run] inversionistas: {list(inv_grupos.keys())}")
            stats["ok"] += 1
            continue

        # ── Obtener o crear la liquidación ───────────────────────────────────
        liq_id = None
        try:
            liq_resp = api.post("/api/v1/liquidaciones", {
                "proyecto_id":     pid,
                "periodo":         periodo_date,
                "tipo_venta":      "ppa",
                "generado_por_id": usuario_id,
            })
            liq_id = liq_resp["id"]
            print(f"  ✓ Liquidación creada id={liq_id}")
            stats["liq_nueva"] += 1
        except requests.HTTPError as e:
            if e.response.status_code in (409, 422, 500):
                resp = api.get("/api/v1/liquidaciones",
                               params={"proyecto_id": pid, "size": 50})
                y, m = periodo_date[:7].split("-")
                liq_id = next(
                    (l["id"] for l in resp["items"]
                     if l["periodo"].startswith(f"{y}-{m.zfill(2)}")),
                    None,
                )
                if not liq_id:
                    print(f"  ✗ No se pudo crear ni encontrar liquidación para {nombre_proy}")
                    continue
                print(f"  ~ Liquidación existente id={liq_id}")
                stats["liq_existente"] += 1
            else:
                raise

        # ── --limpiar ────────────────────────────────────────────────────────
        if limpiar:
            limpiar_liquidacion(api, liq_id)

        # ── Actualizar metadatos de la liquidación ───────────────────────────
        patch_body: dict = {}
        if er_url:
            patch_body["estado_resultados_url"] = er_url
        for f in inv_grupos.get("Total", []):
            if f["doc_contable"] == "Información":
                if f["comprobante"]:
                    patch_body["comprobante_contable_ref"] = f["comprobante"]
                if f["cons_ing_txt"].isdigit():
                    patch_body["consecutivo_inicial_ingresos"] = int(f["cons_ing_txt"])
                if f["cons_cos_txt"].isdigit():
                    patch_body["consecutivo_inicial_costos"] = int(f["cons_cos_txt"])
        if patch_body:
            api.patch(f"/api/v1/liquidaciones/{liq_id}", patch_body)

        # ── Inversionistas del proyecto ──────────────────────────────────────
        proy_detail = api.get(f"/api/v1/proyectos/{pid}")
        inversionistas_db = proy_detail.get("inversionistas", [])

        # ── Procesar cada grupo de inversionista ─────────────────────────────
        for inv_nombre, filas_inv in inv_grupos.items():
            es_total = inv_nombre.upper() == "TOTAL"
            inv_db   = None if es_total else match_inversionista(inversionistas_db, inv_nombre)
            inv_id   = inv_db["id"] if inv_db else None

            if not es_total and inv_db is None:
                print(f"  ⚠  Inversionista sin match: '{inv_nombre}' → se carga sin asignar")

            # Consecutivos del row Información de este inversionista
            cons_ing = cons_cos = None
            for f in filas_inv:
                if f["doc_contable"] == "Información":
                    if f["cons_ing_txt"].isdigit():
                        cons_ing = int(f["cons_ing_txt"])
                    if f["cons_cos_txt"].isdigit():
                        cons_cos = int(f["cons_cos_txt"])

            filas_mandato = [f for f in filas_inv if f["doc_contable"] == "Mandato"]
            filas_costos  = [f for f in filas_inv if f["doc_contable"] == "Costos"]
            filas_factura = [f for f in filas_inv if f["doc_contable"] == "Factura"]

            # numero_mandato = Contato 1 (primer valor no vacío entre las filas del grupo)
            contato1 = next((f["contato1"] for f in filas_inv if f.get("contato1")), "")
            contato2 = next((f["contato2"] for f in filas_inv if f.get("contato2")), "")

            # ── Mandato de ingresos ──────────────────────────────────────────
            lineas_ing = [
                f for f in filas_mandato
                if f["concepto"] and f["total"] is not None
                and _norm(f["concepto"]) not in OMITIR
            ]
            if lineas_ing:
                m_resp = api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos", {
                    "tipo":                "ingresos",
                    "inversionista_id":    inv_id,
                    "numero_mandato":      contato1 or None,
                    "beneficiario_nombre": contato2 or (inv_nombre if not es_total else None),
                    "consecutivo":         cons_ing,
                    "pa_aplica":           inv_db.get("es_patrimonio_autonomo", False) if inv_db else False,
                })
                mid = m_resp["id"]
                stats["mandatos"] += 1

                neto_pagar = None
                for orden, f in enumerate(filas_mandato):
                    if not f["concepto"] or f["total"] is None:
                        continue
                    n = _norm(f["concepto"])
                    if n == "valor a pagar":
                        neto_pagar = f["total"]
                        continue
                    if n in OMITIR:
                        continue
                    # Referencia: combina ref_factura y cons_ing si son distintos
                    partes = [f["ref_factura"]]
                    ci = f["cons_ing_txt"]
                    if ci and not ci.isdigit() and ci != f["ref_factura"]:
                        partes.append(ci)
                    ref = " | ".join(p for p in partes if p) or None

                    api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mid}/lineas", {
                        "tipo_linea":         _tipo_linea(f["concepto"]),
                        "concepto":           f["concepto"],
                        "valor_cop":          f["total"],
                        "referencia_factura": ref,
                        "orden":              orden,
                    })
                    stats["lineas"] += 1

                if neto_pagar is not None:
                    api.patch(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mid}",
                              {"valor_neto_cop": neto_pagar})

            # ── Mandato de costos ────────────────────────────────────────────
            lineas_cos = [
                f for f in filas_costos
                if f["concepto"] and f["total"] is not None
                and _norm(f["concepto"]) not in OMITIR
            ]
            if lineas_cos:
                mc_resp = api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos", {
                    "tipo":                "costos",
                    "inversionista_id":    inv_id,
                    "numero_mandato":      contato1 or None,
                    "beneficiario_nombre": contato2 or (inv_nombre if not es_total else None),
                    "consecutivo":         cons_cos,
                    "pa_aplica":           inv_db.get("es_patrimonio_autonomo", False) if inv_db else False,
                })
                mcid = mc_resp["id"]
                stats["mandatos"] += 1

                neto_cos = None
                for orden, f in enumerate(filas_costos):
                    if not f["concepto"] or f["total"] is None:
                        continue
                    n = _norm(f["concepto"])
                    if n == "valor a pagar":
                        neto_cos = f["total"]
                        continue
                    if n in OMITIR:
                        continue

                    partes = [f["ref_factura"]]
                    ci = f["cons_ing_txt"]
                    if ci and not ci.isdigit() and ci != f["ref_factura"]:
                        partes.append(ci)
                    ref = " | ".join(p for p in partes if p) or None

                    api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mcid}/lineas", {
                        "tipo_linea":         _tipo_linea(f["concepto"]),
                        "concepto":           f["concepto"],
                        "valor_cop":          f["total"],
                        "referencia_factura": ref,
                        "soporte_url":        f.get("ref_factura_url") or f.get("cons_ing_url") or None,
                        "orden":              orden,
                    })
                    stats["lineas"] += 1

                    # LiquidacionCosto (nivel proyecto) solo para filas del Total
                    if es_total and f["total"] != 0:
                        api.post(f"/api/v1/liquidaciones/{liq_id}/costos", {
                            "tipo_costo":  concepto_a_tipo_costo(f["concepto"]),
                            "descripcion": f["concepto"],
                            "nro_soporte": f["ref_factura"] or f["comprobante"] or None,
                            "soporte_url": f["cons_ing_url"],
                            "valor_cop":   f["total"],
                        })
                        stats["costos"] += 1

                if neto_cos is not None:
                    api.patch(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mcid}",
                              {"valor_neto_cop": neto_cos})

            # ── Facturas de servicio (solo Total = nivel proyecto) ────────────
            if es_total:
                for f in filas_factura:
                    if not f["concepto"] or f["total"] is None:
                        continue
                    api.post(f"/api/v1/liquidaciones/{liq_id}/facturas", {
                        "tipo_servicio":  concepto_a_tipo_factura(f["concepto"]),
                        "numero_factura": f["ref_factura"] or None,
                        "nro_soporte":    f["cons_ing_txt"] or None,
                        "soporte_url":    f["cons_ing_url"],
                        "valor_cop":      f["total"],
                    })
                    stats["facturas"] += 1

        stats["ok"] += 1

    # ── Resumen final ─────────────────────────────────────────────────────────
    print("\n" + "=" * 55)
    print(f"Proyectos cargados:         {stats['ok']}")
    print(f"Liquidaciones nuevas:       {stats['liq_nueva']}")
    print(f"Liquidaciones existentes:   {stats['liq_existente']}")
    print(f"Mandatos creados:           {stats['mandatos']}")
    print(f"Líneas creadas:             {stats['lineas']}")
    print(f"Costos (nivel proyecto):    {stats['costos']}")
    print(f"Facturas de servicio:       {stats['facturas']}")
    print(f"Duplicados omitidos:        {stats['duplicados']}")
    if stats["sin_match"]:
        print(f"\nProyectos sin match en DB ({len(stats['sin_match'])}):")
        for n in stats["sin_match"]:
            print(f"  - {n}")


# ─────────────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(
        description="Carga panel de seguimiento contable al backend de Unergy"
    )
    p.add_argument("--xlsx",     required=True,  help="Ruta al archivo Excel")
    p.add_argument("--hoja",     default="Enero", help="Nombre de la hoja (default: Enero)")
    p.add_argument("--periodo",  required=True,  help="Período YYYY-MM, ej: 2026-01")
    p.add_argument("--api-url",  default="https://backend-production-63d8.up.railway.app")
    p.add_argument("--usuario",  required=True)
    p.add_argument("--password", required=True)
    p.add_argument("--limpiar",  action="store_true",
                   help="Elimina mandatos/costos/facturas antes de cargar")
    p.add_argument("--dry-run",  action="store_true",
                   help="Solo muestra lo que haría, sin escribir en la API")
    args = p.parse_args()

    y, m = args.periodo.split("-")
    periodo_date = f"{y}-{m.zfill(2)}-01"

    print(f"Leyendo '{args.hoja}' de {args.xlsx} ...")
    filas, er_map = leer_hoja(args.xlsx, args.hoja)
    print(f"  {len(filas)} filas válidas | {len(er_map)} proyectos con ER URL")

    if args.dry_run:
        print("\n[DRY RUN — no se escribe nada]\n")
    if args.limpiar and not args.dry_run:
        print("\n⚠  --limpiar activo: se borrarán mandatos/costos/facturas antes de cargar\n")

    print(f"\nAutenticando en {args.api_url} ...")
    api = API(args.api_url)
    api.login(args.usuario, args.password)
    print("Autenticado ✓\n")

    cargar(api, filas, er_map, periodo_date,
           dry_run=args.dry_run,
           limpiar=args.limpiar and not args.dry_run)


if __name__ == "__main__":
    main()

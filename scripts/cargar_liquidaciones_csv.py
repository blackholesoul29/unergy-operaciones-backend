#!/usr/bin/env python3
"""
Carga un panel de seguimiento contable (XLSX) al sistema de liquidaciones.
Lee hipervínculos directamente desde el archivo Excel para obtener los
links a Drive de cada soporte.

Uso:
  python scripts/cargar_liquidaciones_csv.py \
      --xlsx "ruta/al/2026_Panel de seguimiento contable.xlsx" \
      --hoja Enero \
      --periodo 2026-01 \
      --api-url https://backend-production-63d8.up.railway.app \
      --usuario admin@unergy.co \
      --password TU_PASSWORD

Hojas disponibles: Enero, Febrero, Marzo, Abril prueba costos, etc.
"""
import argparse, re, sys
from openpyxl import load_workbook
import requests

# Forzar UTF-8 en stdout/stderr para que los símbolos Unicode (⚠, ✓, ✗)
# no causen UnicodeEncodeError en entornos Windows con cp1252.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


# ── Mapping concepto → tipo_linea ─────────────────────────────────────────────
TIPO_LINEA_MAP = [
    (r"ingreso bruto",                         "ingreso_bruto"),
    (r"^ingreso$",                             "ingreso_bruto"),
    (r"despacho",                              "despacho"),
    (r"ventas en bolsa",                       "ventas_en_bolsa"),
    (r"compras en bolsa|compras a ",           "compras_en_bolsa"),
    (r"redistribuci",                          "redistribucion_ingresos"),
    (r"comercializaci",                        "ajuste_comercializacion"),
    (r"ajuste.*xm|xm.*ajuste",                "ajuste_xm"),
    (r"ajuste.*unergy",                        "ajuste_unergy"),
    (r"ajuste.*administr",                     "otro_costo"),
    (r"arriendo",                              "arriendo"),
    (r"mantenimiento",                         "mantenimiento"),
    (r"iva.*internet|internet.*iva",           "iva"),
    (r"internet",                              "servicio_internet"),
    (r"poliza.*cumplimiento|p.liza.*cumpl",    "poliza_cumplimiento"),
    (r"poliza|incendio|lucro cesante",         "seguro"),
    (r"servicios p.blicos|consumo de energ",   "servicios_publicos_consumo"),
    (r"cambio.*equipo|equipo.*medida",         "cambio_equipos_medida"),
    (r"representaci",                          "representacion"),
    (r"^cgm",                                  "cgm"),
    (r"administraci|valor final",              "administracion"),
    (r"reteica",                               "reteica"),
    (r"ica opex",                              "ica_opex"),
    (r"iva|i\.v\.a",                           "iva"),
    (r"retenci",                               "retencion_fuente"),
    (r"porcentaje.*participaci",               "porcentaje_participacion"),
    (r"valor.*pagar|valor.*ganancia|utilidad", "valor_a_pagar"),
    (r"intereses",                             "intereses"),
]

# Bug 4: doc types válidos y mapa de normalización a title-case
_DOC_CONTABLE_MAP = {
    "informacion": "Información",
    "mandato":     "Mandato",
    "costos":      "Costos",
    "factura":     "Factura",
    "buenaventura": "Buenaventura",
}
TIPOS_VALIDOS_DOC = set(_DOC_CONTABLE_MAP.keys())


def normalizar_doc_contable(s: str) -> str:
    """Normaliza el valor de Documento contable: trim + title-case canónico."""
    return _DOC_CONTABLE_MAP.get(normalizar(s), s.strip())


def normalizar(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize("NFD", str(s or "").lower().strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


ALIASES: dict[str, str] = {
    "ayura sas":                    "ayura",
    "ayura s.a.s":                  "ayura",
    "solenium sas":                 "solenium",
    "solenium s.a.s":               "solenium",
    "patrimonios autonomos":        "patrimonios",
    "suno activos sostenibles":     "suno",
    "rodriguez velez beatriz":      "rodriguez velez",
    "inversiones estrada arbelaez": "estrada",
    "strada asociados":             "estrada",
}


def normalizar_alias(nombre: str) -> str:
    n = normalizar(nombre)
    for patron, alias in ALIASES.items():
        if patron in n:
            return alias
    return n


def concepto_a_tipo_linea(concepto: str) -> str:
    norm = normalizar(concepto)
    for patron, tipo in TIPO_LINEA_MAP:
        if re.search(patron, norm):
            return tipo
    return "otro_ingreso"


def concepto_a_tipo_costo(concepto: str) -> str:
    norm = normalizar(concepto)
    if "arriendo" in norm:                          return "arriendo"
    if "mantenimiento" in norm:                     return "mantenimiento"
    if "internet" in norm:                          return "internet"
    if "poliza" in norm or "p.liza" in norm:        return "polizas"
    if "servicios p" in norm:                       return "servicios_publicos"
    if "cambio" in norm and "equipo" in norm:       return "cambio_equipos_medida"
    if "comercializaci" in norm:                    return "comercializacion_xm"
    return "otro"


def concepto_a_tipo_factura(concepto: str) -> str:
    norm = normalizar(concepto)
    if "representaci" in norm:  return "representacion"
    if "cgm" in norm:           return "cgm"
    return "administracion_operacion"


def parse_valor(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[$,\s]", "", str(v).strip())
    if not s or s in ("-", ""):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def celda(row, idx: int):
    """Valor de celda o None si fuera de rango."""
    return row[idx] if idx < len(row) else None


def link(row, idx: int) -> str | None:
    """Hipervínculo de la celda o None."""
    cell = row[idx] if idx < len(row) else None
    if cell and hasattr(cell, "hyperlink") and cell.hyperlink:
        return cell.hyperlink.target
    return None


def valor(row, idx: int):
    c = celda(row, idx)
    return c.value if c is not None else None


# ── Leer hoja XLSX preservando hipervínculos ──────────────────────────────────
def leer_hoja(xlsx_path: str, hoja: str) -> tuple[list[dict], dict[str, str]]:
    """
    Devuelve (filas, er_map) donde:
    - filas: lista de dicts con datos de la tabla principal (cols A-K)
    - er_map: dict {nombre_proyecto_normalizado: er_url} leído de la tabla
              lateral derecha (col N = proyecto, col O = ER URL)

    Índices 0-based:
      0=Proyecto  1=Inversionista  2=DocContable  5=Concepto  6=Total
      7=RefFactura  8=ConsIng(soporte_url)  9=ConsCostos  10=Comprobante
      13=ProyectoER  14=EstadoResultados(URL)  15=CarpetaProyecto
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if hoja not in wb.sheetnames:
        raise ValueError(f"Hoja '{hoja}' no encontrada. Disponibles: {wb.sheetnames}")
    sh = wb[hoja]

    rows = []
    er_map: dict[str, str] = {}

    for xl_row in sh.iter_rows(min_row=2, max_row=sh.max_row):
        # ── Tabla lateral: col N → proyecto, col O → ER URL ──
        if len(xl_row) > 14:
            n_val = valor(xl_row, 13)
            o_cell = xl_row[14]
            if n_val:
                er_url = (o_cell.hyperlink.target if o_cell.hyperlink else None) \
                         or str(o_cell.value or "").strip()
                if er_url and er_url.startswith("http"):
                    er_map[normalizar(str(n_val).strip())] = er_url

        # ── Tabla principal: col A debe tener proyecto ──
        proy = valor(xl_row, 0)
        if not proy:
            continue
        rows.append({
            "proyecto":        str(proy).strip(),
            "inversionista":   str(valor(xl_row, 1) or "").strip(),
            "doc_contable":    normalizar_doc_contable(str(valor(xl_row, 2) or "")),
            "concepto":        str(valor(xl_row, 5) or "").strip(),
            "total":           parse_valor(valor(xl_row, 6)),
            "ref_factura":     str(valor(xl_row, 7) or "").strip(),
            "ref_factura_url": link(xl_row, 7),
            "cons_ing_txt":    str(valor(xl_row, 8) or "").strip(),
            "cons_ing_url":    link(xl_row, 8),    # ← link directo al PDF en Drive
            "cons_cos_txt":    str(valor(xl_row, 9) or "").strip() if len(xl_row) > 9 else "",
            "comprobante":     str(valor(xl_row, 10) or "").strip() if len(xl_row) > 10 else "",
            "carpeta_url":     link(xl_row, 15),
        })
    wb.close()
    return rows, er_map


# ── API client ────────────────────────────────────────────────────────────────
class API:
    def __init__(self, base: str):
        self.base = base.rstrip("/")
        self.token = None

    def login(self, usuario: str, password: str):
        r = requests.post(f"{self.base}/api/v1/auth/token",
                          data={"username": usuario, "password": password})
        r.raise_for_status()
        self.token = r.json()["access_token"]

    def _h(self):
        return {"Authorization": f"Bearer {self.token}"}

    def get(self, path, **kw):
        import time
        kw.setdefault("timeout", 30)
        for attempt in range(4):
            try:
                r = requests.get(f"{self.base}{path}", headers=self._h(), **kw)
                r.raise_for_status()
                return r.json()
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
                if attempt == 3:
                    raise
                wait = 2 ** attempt
                print(f"    [retry {attempt+1}/3 en {wait}s — {e.__class__.__name__}]")
                time.sleep(wait)
            except requests.exceptions.HTTPError:
                raise

    def post(self, path, body):
        import time, sys
        for attempt in range(4):
            try:
                r = requests.post(f"{self.base}{path}", json=body, headers=self._h(), timeout=30)
                # Retry transient 5xx (Railway cold-start / overload)
                if r.status_code in (500, 502, 503, 504) and attempt < 3:
                    wait = 2 ** attempt
                    print(f"    [retry {attempt+1}/3 en {wait}s — 500: {r.text[:120]}]")
                    sys.stdout.flush()
                    time.sleep(wait)
                    continue
                r.raise_for_status()
                return r.json()
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
                if attempt == 3:
                    raise
                wait = 2 ** attempt
                print(f"    [retry {attempt+1}/3 en {wait}s — {e.__class__.__name__}]")
                sys.stdout.flush()
                time.sleep(wait)
            except requests.exceptions.HTTPError as e:
                if e.response is not None:
                    print(f"    [HTTP {e.response.status_code} — path={path} body={body}]")
                    print(f"    [response: {e.response.text[:400]}]")
                    sys.stdout.flush()
                raise

    def patch(self, path, body):
        import time, sys
        for attempt in range(4):
            try:
                r = requests.patch(f"{self.base}{path}", json=body, headers=self._h(), timeout=30)
                # Retry transient 5xx (Railway cold-start / overload)
                if r.status_code in (500, 502, 503, 504) and attempt < 3:
                    wait = 2 ** attempt
                    print(f"    [retry {attempt+1}/3 en {wait}s — {r.status_code}: {r.text[:120]}]")
                    sys.stdout.flush()
                    time.sleep(wait)
                    continue
                r.raise_for_status()
                return r.json()
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
                if attempt == 3:
                    raise
                wait = 2 ** attempt
                print(f"    [retry {attempt+1}/3 en {wait}s — {e.__class__.__name__}]")
                sys.stdout.flush()
                time.sleep(wait)
            except requests.exceptions.HTTPError as e:
                if e.response is not None:
                    print(f"    [HTTP {e.response.status_code} — path={path} body={body}]")
                    print(f"    [response: {e.response.text[:400]}]")
                    sys.stdout.flush()
                raise


# ── Match helpers ─────────────────────────────────────────────────────────────
def match_proyecto(proyectos_db: list, nombre: str) -> dict | None:
    norm = normalizar(nombre)

    # ── Paso 1: match exacto normalizado ────────────────────────────────────────
    for p in proyectos_db:
        if normalizar(p["nombre_comercial"]) == norm:
            return p

    # ── Número final y código MGS extraídos del nombre Excel ────────────────────
    # Ej: "Valencia Oriente 2"   → trailing_num="2"
    #     "Polaris 2 trading"    → trailing_num="2"  (número antes del sufijo)
    #     "MGS 0026 Valencia..." → mgs_code="mgs0026"
    #
    # Sufijos que no forman parte del nombre del proyecto en DB y deben
    # ignorarse para extraer el número significativo.
    _SUFIJOS = re.compile(
        r'\s*(trading|dup\b|duplicado|dulicado|terpel|excedentes|excendentes'
        r'|gasto|gastos)\s*$'
    )
    norm_sin_sufijo = _SUFIJOS.sub('', norm).strip()

    _m_trail = re.search(r'\b(\d+)\s*$', norm_sin_sufijo)
    trailing_num: str | None = _m_trail.group(1) if _m_trail else None

    _m_code = re.search(r'\b(mgs[\s\-]*\d+)\b', norm)
    excel_mgs: str | None = re.sub(r'[\s\-]+', '', _m_code.group(1)) if _m_code else None

    def _db_trailing(p: dict) -> str | None:
        # Reconoce números finales en estilos "Norte 2" Y "N1" / "N2"
        # \bn? captura la 'n' opcional antes del número (ej: "N1" → "1")
        m = re.search(r'\bn?(\d+)\s*$', normalizar(p["nombre_comercial"]))
        return m.group(1) if m else None

    def _db_mgs(p: dict) -> str | None:
        m = re.search(r'\b(mgs[\s\-]*\d+)\b', normalizar(p["nombre_comercial"]))
        return re.sub(r'[\s\-]+', '', m.group(1)) if m else None

    def _num_ok(p: dict) -> bool:
        """Si Excel tiene número final, el DB debe tener el mismo número.
        - db_num None (sin número en DB) se acepta: el proyecto podría
          tener un nombre abreviado sin número explícito.
        - Con _db_trailing mejorado, "N1" → "1", así "Chiriguana N1" ya
          devuelve "1" y no None, por lo que no será aceptado para "Chiriguana 2"."""
        if trailing_num is None:
            return True
        db_num = _db_trailing(p)
        return db_num is None or db_num == trailing_num

    # ── Paso 1.5: match por código MGS si el nombre Excel lo incluye ─────────────
    if excel_mgs:
        for p in proyectos_db:
            if _db_mgs(p) == excel_mgs:
                return p

    # ── Paso 2: substring match con guard de número final ───────────────────────
    for p in proyectos_db:
        n = normalizar(p["nombre_comercial"])
        if n and (n in norm or norm in n):
            if not _num_ok(p):
                continue
            return p

    # ── Paso 3: keyword match con guard de número final ─────────────────────────
    # Omitir tokens numéricos puros para no hacer match sólo por "2" == "2"
    partes = [t for t in norm.split() if len(t) >= 4 and not t.isdigit()]
    for parte in reversed(partes):
        for p in proyectos_db:
            if parte in normalizar(p["nombre_comercial"]):
                if not _num_ok(p):
                    continue
                return p

    return None


def match_inversionista(inversionistas_db: list, nombre: str) -> dict | None:
    if not nombre or nombre.upper() == "TOTAL":
        return None

    def _nombre_db(inv: dict) -> str:
        return (inv.get("cliente_nombre") or inv.get("razon_social_nombre") or
                inv.get("nombre") or inv.get("razon_social") or "")

    # Fase 1: match directo normalizado (sin aliases) — evita colisiones entre
    # entidades distintas que comparten el mismo alias (ej. ESTRADA vs STRADA)
    norm = normalizar(nombre)
    for inv in inversionistas_db:
        cn = normalizar(_nombre_db(inv))
        if cn and (cn == norm or cn in norm or norm in cn):
            return inv
    palabras = [w for w in norm.split() if len(w) > 4]
    for inv in inversionistas_db:
        cn = normalizar(_nombre_db(inv))
        if palabras and cn and all(w in cn for w in palabras[:2]):
            return inv

    # Fase 2: match por alias — solo si no hubo match directo
    norm_alias = normalizar_alias(nombre)
    if norm_alias != norm:
        for inv in inversionistas_db:
            cn = normalizar_alias(_nombre_db(inv))
            if cn and cn == norm_alias:
                return inv

    return None


# ── Carga principal ───────────────────────────────────────────────────────────
def _tiene_mandatos(api: API, liq_id: int, proyecto_id: int) -> bool:
    try:
        resp = api.get(
            "/api/v1/liquidaciones/vistas/por-proyecto",
            params={"proyecto_id": proyecto_id},
        )
        for proy in resp:
            for liq in proy.get("liquidaciones", []):
                if liq["liquidacion_id"] == liq_id:
                    for inv in liq.get("inversionistas", []):
                        if inv.get("mandatos_ingresos") or inv.get("mandatos_costos"):
                            return True
                    if liq.get("mandatos_total_ingresos") or liq.get("mandatos_total_costos"):
                        return True
        return False
    except Exception:
        return False


def cargar(api: API, filas: list[dict], er_map: dict[str, str], periodo_date: str, dry_run: bool, limpiar: bool = False):
    # Bug 4: validar tipos de Documento contable desconocidos
    tipos_desconocidos: dict[str, list[int]] = {}
    for i, f in enumerate(filas, start=2):
        nd = normalizar(f["doc_contable"])
        if nd and nd not in TIPOS_VALIDOS_DOC:
            tipos_desconocidos.setdefault(f["doc_contable"], []).append(i)
    if tipos_desconocidos:
        print("\n⚠  TIPOS NO RECONOCIDOS en 'Documento contable':")
        for tipo, filas_idx in tipos_desconocidos.items():
            print(f"   '{tipo}' → filas: {filas_idx[:10]}{'…' if len(filas_idx) > 10 else ''}")
        print("   Estas filas serán ignoradas. Revisar el archivo o agregar al mapa.\n")

    me = api.get("/api/v1/auth/me")
    usuario_id = me["id"]

    proyectos_db = api.get("/api/v1/proyectos", params={"size": 500})["items"]
    print(f"  {len(proyectos_db)} proyectos en DB")

    # Agrupar filas: proyecto → inversionista → lista de filas
    grupos: dict[str, dict[str, list]] = {}

    for f in filas:
        grupos.setdefault(f["proyecto"], {}).setdefault(f["inversionista"], []).append(f)

    # Patrón para proyectos que deben omitirse (trading, duplicados, gastos)
    _OMITIR_PROY = re.compile(
        r'\btrading\b|duplicado|dulicado|\bdup\b', re.IGNORECASE
    )

    stats = {"ok": 0, "omitidos": 0, "sin_match": [], "liq": 0,
             "mandatos": 0, "lineas": 0, "costos": 0, "facturas": 0,
             "sin_match_inv": []}

    for nombre_proy, inv_grupos in grupos.items():
        # Omitir filas de trading y duplicados — no tienen liquidación propia en DB
        if _OMITIR_PROY.search(nombre_proy):
            stats["omitidos"] += 1
            continue

        proy_db = match_proyecto(proyectos_db, nombre_proy)
        if not proy_db:
            stats["sin_match"].append(nombre_proy)
            print(f"  ⚠  Sin match: '{nombre_proy}'")
            continue

        pid = proy_db["id"]
        er_url = er_map.get(normalizar(nombre_proy))
        print(f"\n-> {nombre_proy} (id={pid})")

        if dry_run:
            invs = list(inv_grupos.keys())
            print(f"   Inversionistas: {invs}")
            stats["ok"] += 1
            continue

        try:
            inversionistas_db = api.get(f"/api/v1/proyectos/{pid}/inversionistas")
        except Exception as exc:

            print(f"  ✗ Error obteniendo inversionistas para {pid}: {exc}")
            sys.stdout.flush()
            continue

        # Crear / recuperar liquidación
        try:
            liq_resp = api.post("/api/v1/liquidaciones", {
                "proyecto_id": pid,
                "periodo": periodo_date,
                "tipo_venta": "ppa",
                "generado_por_id": usuario_id,
            })
            liq_id = liq_resp["id"]
            print(f"  ✓ Liquidación creada id={liq_id}")
            stats["liq"] += 1
        except requests.HTTPError as e:
            if e.response.status_code in (409, 422, 500):
                # 409 = duplicate; 500 might also be a duplicate before migration fix
                resp = api.get("/api/v1/liquidaciones",
                               params={"proyecto_id": pid, "size": 50})
                y, m = periodo_date[:7].split("-")
                liq_id = next(
                    (l["id"] for l in resp["items"]
                     if l["periodo"].startswith(f"{y}-{m.zfill(2)}")), None)
                if not liq_id:
                    print(f"  ✗ No se pudo crear ni encontrar liquidación (status={e.response.status_code})")
                    continue
                print(f"  ~ Liquidación existente id={liq_id}")

                if not limpiar and _tiene_mandatos(api, liq_id, pid):
                    print(
                        f"  ⚠  Liquidación {liq_id} ya tiene mandatos cargados.\n"
                        f"     Usa --limpiar para reimportar sin duplicados. Saltando proyecto."
                    )
                    stats["ok"] += 1
                    continue

                if limpiar:
                    if not dry_run:
                        import time as _time
                        for _attempt in range(3):
                            try:
                                r = requests.delete(
                                    f"{api.base}/api/v1/liquidaciones/{liq_id}/limpiar",
                                    headers=api._h(),
                                    timeout=60,
                                )
                                r.raise_for_status()
                                print(f"  ~ Liquidación {liq_id} limpiada (mandatos/costos/facturas borrados)")
                                sys.stdout.flush()
                                break
                            except (requests.exceptions.ConnectionError,
                                    requests.exceptions.Timeout) as exc:
                                if _attempt == 2:
                                    print(f"  ✗ Error limpiando liquidación {liq_id} (timeout): {exc}")
                                    sys.stdout.flush()
                                    break
                                _wait = 2 ** _attempt
                                print(f"  [retry limpiar {_attempt+1}/2 en {_wait}s]")
                                sys.stdout.flush()
                                _time.sleep(_wait)
                            except Exception as exc:
                                print(f"  ✗ Error limpiando liquidación {liq_id}: {exc}")
                                sys.stdout.flush()
                                break
                        else:
                            continue
                    else:
                        print(f"  [DRY RUN] Se limpiaría liquidación {liq_id}")
            else:
                raise

        # Actualizar metadatos de la liquidación
        patch_body: dict = {}
        if er_url:
            patch_body["estado_resultados_url"] = er_url
        total_rows = inv_grupos.get("Total", [])
        for f in total_rows:
            if f["doc_contable"].lower() == "información":
                if f["comprobante"]:
                    patch_body["comprobante_contable_ref"] = f["comprobante"]
                if f["cons_ing_txt"].isdigit():
                    patch_body["consecutivo_inicial_ingresos"] = int(f["cons_ing_txt"])
                if f["cons_cos_txt"].isdigit():
                    patch_body["consecutivo_inicial_costos"] = int(f["cons_cos_txt"])
        if patch_body:
            api.patch(f"/api/v1/liquidaciones/{liq_id}", patch_body)

        # Procesar cada grupo de inversionista
        for inv_nombre, filas_inv in inv_grupos.items():
            es_total = inv_nombre.upper() == "TOTAL"
            inv_db = None if es_total else match_inversionista(inversionistas_db, inv_nombre)
            inv_id = inv_db["id"] if inv_db else None

            if not es_total and inv_db is None and inv_nombre.strip():
                stats["sin_match_inv"].append(f"{nombre_proy} → {inv_nombre}")
                print(
                    f"  ⚠  Inversionista sin match en DB: '{inv_nombre}' "
                    f"(proyecto '{nombre_proy}'). Filas cargadas sin inversionista_id."
                )

            # Consecutivos del row Información de este inversionista
            cons_ing = cons_cos = None
            for f in filas_inv:
                if f["doc_contable"].lower() == "información":
                    if f["cons_ing_txt"].isdigit():
                        cons_ing = int(f["cons_ing_txt"])
                    if f["cons_cos_txt"].isdigit():
                        cons_cos = int(f["cons_cos_txt"])

            # Bug 1/3: "Buenaventura" contiene Representación/CGM → es Factura, no Mandato
            filas_ing = [f for f in filas_inv
                         if f["doc_contable"].lower() == "mandato"]
            filas_cos = [f for f in filas_inv
                         if f["doc_contable"].lower() == "costos"]
            filas_fac = [f for f in filas_inv
                         if f["doc_contable"].lower() in ("factura", "buenaventura")]

            # ── Mandato de ingresos ──
            lineas_ing = [f for f in filas_ing
                          if f["concepto"] and
                          normalizar(f["concepto"]) not in ("porcentaje de participacion", "valor a pagar")
                          and f["total"] is not None]
            if lineas_ing:
                m_resp = api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos", {
                    "tipo": "ingresos",
                    "inversionista_id": inv_id,
                    "beneficiario_nombre": inv_nombre if not es_total else None,
                    "consecutivo": cons_ing,
                    "pa_aplica": inv_db.get("es_patrimonio_autonomo", False) if inv_db else False,
                })
                mid = m_resp["id"]
                stats["mandatos"] += 1

                neto_pagar = None
                for orden, f in enumerate(filas_ing):
                    if not f["concepto"] or f["total"] is None:
                        continue
                    if normalizar(f["concepto"]) == "valor a pagar":
                        neto_pagar = f["total"]
                        continue
                    tipo_l = concepto_a_tipo_linea(f["concepto"])
                    # referencia: combina ref_factura + nombre soporte si son distintos
                    ref_parts = [f["ref_factura"]]
                    if f["cons_ing_txt"] and not f["cons_ing_txt"].isdigit() and f["cons_ing_txt"] != f["ref_factura"]:
                        ref_parts.append(f["cons_ing_txt"])
                    ref = " | ".join(p for p in ref_parts if p) or None
                    if ref and len(ref) > 255:
                        ref = ref[:252] + "..."
                    try:
                        api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mid}/lineas", {
                            "tipo_linea": tipo_l,
                            "concepto": f["concepto"],
                            "valor_cop": f["total"],
                            "referencia_factura": ref,
                            "orden": orden,
                        })
                    except Exception as exc:
            
                        print(f"    ✗ Linea-ing '{f['concepto']}' → {tipo_l}: {exc}")
                        sys.stdout.flush()
                        continue
                    stats["lineas"] += 1

                if neto_pagar is not None:
                    try:
                        api.patch(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mid}",
                                  {"valor_neto_cop": neto_pagar})
                    except Exception as exc:
            
                        print(f"    ✗ patch valor_neto mandato {mid}: {exc}")
                        sys.stdout.flush()

            # ── Mandato de costos ──
            lineas_cos = [f for f in filas_cos
                          if f["concepto"] and f["total"] is not None
                          and normalizar(f["concepto"]) not in ("porcentaje de participacion", "valor a pagar")]
            if lineas_cos:
                mc_resp = api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos", {
                    "tipo": "costos",
                    "inversionista_id": inv_id,
                    "beneficiario_nombre": inv_nombre if not es_total else None,
                    "consecutivo": cons_cos,
                    "pa_aplica": inv_db.get("es_patrimonio_autonomo", False) if inv_db else False,
                })
                mcid = mc_resp["id"]
                stats["mandatos"] += 1

                neto_cos = None
                for orden, f in enumerate(filas_cos):
                    if not f["concepto"] or f["total"] is None:
                        continue
                    if normalizar(f["concepto"]) == "valor a pagar":
                        neto_cos = f["total"]
                        continue
                    tipo_l = concepto_a_tipo_linea(f["concepto"])
                    ref_parts = [f["ref_factura"]]
                    if f["cons_ing_txt"] and not f["cons_ing_txt"].isdigit() and f["cons_ing_txt"] != f["ref_factura"]:
                        ref_parts.append(f["cons_ing_txt"])
                    ref = " | ".join(p for p in ref_parts if p) or None
                    if ref and len(ref) > 255:
                        ref = ref[:252] + "..."
                    try:
                        api.post(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mcid}/lineas", {
                            "tipo_linea": tipo_l,
                            "concepto": f["concepto"],
                            "valor_cop": f["total"],
                            "referencia_factura": ref,
                            "orden": orden,
                        })
                    except Exception as exc:
            
                        print(f"    ✗ Linea-cos '{f['concepto']}' → {tipo_l}: {exc}")
                        sys.stdout.flush()
                        continue
                    stats["lineas"] += 1

                    # LiquidacionCosto solo para el Total (nivel proyecto)
                    if es_total and f["total"] != 0:
                        api.post(f"/api/v1/liquidaciones/{liq_id}/costos", {
                            "tipo_costo": concepto_a_tipo_costo(f["concepto"]),
                            "descripcion": f["concepto"],
                            "nro_soporte": (f["ref_factura"] or f["comprobante"]) or None,
                            "soporte_url": f["cons_ing_url"],  # ← link al PDF en Drive
                            "valor_cop": f["total"],
                        })
                        stats["costos"] += 1

                if neto_cos is not None:
                    api.patch(f"/api/v1/liquidaciones/{liq_id}/mandatos/{mcid}",
                              {"valor_neto_cop": neto_cos})

            # ── Facturas de servicio (solo Total = nivel proyecto) ──
            if es_total:
                for f in filas_fac:
                    if not f["concepto"] or f["total"] is None:
                        continue
                    try:
                        api.post(f"/api/v1/liquidaciones/{liq_id}/facturas", {
                            "tipo_servicio": concepto_a_tipo_factura(f["concepto"]),
                            "numero_factura": f["ref_factura"] or None,
                            "nro_soporte": f["cons_ing_txt"] or None,
                            "soporte_url": f["cons_ing_url"],   # ← link al PDF en Drive
                            "valor_cop": f["total"],
                        })
                    except Exception as exc:
            
                        print(f"    ✗ Factura '{f['concepto']}' falló: {exc}")
                        sys.stdout.flush()
                        continue
                    stats["facturas"] += 1

        stats["ok"] += 1

    print("\n" + "=" * 50)
    print(f"Proyectos cargados:     {stats['ok']}")
    print(f"Proyectos omitidos:     {stats['omitidos']}  (trading / duplicados)")
    print(f"Liquidaciones creadas:  {stats['liq']}")
    print(f"Mandatos creados:       {stats['mandatos']}")
    print(f"Líneas creadas:         {stats['lineas']}")
    print(f"Costos (nivel proy):    {stats['costos']}")
    print(f"Facturas creadas:       {stats['facturas']}")
    if stats["sin_match"]:
        print(f"\nProyectos sin match en DB ({len(stats['sin_match'])}):")
        for n in stats["sin_match"]:
            print(f"  - {n}")
    if stats["sin_match_inv"]:
        print(f"\nInversionistas sin match en DB ({len(stats['sin_match_inv'])}) — usaron beneficiario_nombre:")
        for n in stats["sin_match_inv"]:
            print(f"  - {n}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx",     required=True)
    p.add_argument("--hoja",     default="Enero")
    p.add_argument("--periodo",  required=True, help="YYYY-MM  ej: 2026-01")
    p.add_argument("--api-url",  default="https://backend-production-63d8.up.railway.app")
    p.add_argument("--usuario",  required=True)
    p.add_argument("--password", required=True)
    p.add_argument("--dry-run",  action="store_true")
    p.add_argument(
        "--limpiar",
        action="store_true",
        help="Borra mandatos/costos/facturas existentes antes de reimportar. "
             "Usar para corregir cargas duplicadas.",
    )
    args = p.parse_args()

    y, m = args.periodo.split("-")
    periodo_date = f"{y}-{m.zfill(2)}-01"

    print(f"Leyendo hoja '{args.hoja}' de {args.xlsx} ...")
    filas, er_map = leer_hoja(args.xlsx, args.hoja)
    print(f"  {len(filas)} filas leídas, {len(er_map)} proyectos con ER URL")

    if args.dry_run:
        print("\n[DRY RUN — no se escribira nada en la API]\n")

    print(f"Autenticando en {args.api_url}...")
    api = API(args.api_url)
    api.login(args.usuario, args.password)
    print("Autenticado\n")

    cargar(api, filas, er_map, periodo_date, dry_run=args.dry_run, limpiar=args.limpiar)


if __name__ == "__main__":
    main()

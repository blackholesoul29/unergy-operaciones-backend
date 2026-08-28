"""Genera app/services/registros_proyecto/catalogo_parametros.py desde la Hoja de Vida.

La hoja "Definicion de campos" del formato oficial de Hoja de Vida del sistema de
medicion (CREG 038/2014) es un diccionario de datos completo: por cada campo trae
ID, titulo, tipo de dato, si es requerido y la descripcion regulatoria. Este script
la lee y emite el catalogo de parametros UNICOS aplicando las reglas de
deduplicacion decididas en docs/registros-proyecto-decisiones.md.

Uso (el .py generado es la fuente de verdad en el repo; esto solo lo regenera):

    python scripts/generar_catalogo_parametros.py \
        "<ruta>/ASIC/01_Hoja_de_Vida/<archivo>.xlsx" \
        app/services/registros_proyecto/catalogo_parametros.py

Reglas de deduplicacion (ver decisiones D-03..D-07):

 D-03  Secciones 3 (medidor de activa principal) y 4 (medidor de reactiva
       principal) describen EL MISMO equipo fisico. Prueba: en el proyecto de
       muestra 3.3 y 4.3 traen la misma serie (88866569), y la seccion 5
       (respaldo) ya describe activa y reactiva en UN solo bloque. Se colapsan
       en MEDIDOR_PRINCIPAL; de la seccion 4 solo sobreviven los campos que de
       verdad difieren (indice de clase, constante, unidad y canales reactivos).

 D-04  Un parametro se define UNA vez y se instancia por equipo. "Numero de
       serie" no es un parametro del medidor principal y otro del de respaldo:
       es un unico parametro `medidor.numero_de_serie` que aplica a los dos
       tipos de equipo (campo `equipo_tipos`). La unicidad real vive en la BD:
       (proyecto_id, clave, equipo_tipo, equipo_posicion).

 D-05  Los sellos son una tabla de 7 filas x 7 columnas repetida en cada
       seccion (49 "campos" del formato). No son 49 parametros: se colapsan en
       un unico parametro de tipo TABLA con sus columnas declaradas.

 D-06  Las secciones 13, 14 y 15 son bitacoras historicas (registros de acceso
       y de novedades), no parametros del proyecto. Se omiten.

 D-07  Los bloques de certificado se nombran por su semantica, no por su
       numeral: 3.38.1 no es "cert_38_numero" sino
       `medidor.cert_conformidad_numero`. El mapa BLOQUES lo hace explicito.
"""

from __future__ import annotations

import re
import sys
import unicodedata
from collections import OrderedDict

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Falta openpyxl: pip install openpyxl")


# ---------------------------------------------------------------------------
# Seccion -> (grupo, ambito, equipo_tipo, etiqueta del grupo)
#   ambito PROYECTO: un valor por proyecto.
#   ambito EQUIPO:   un valor por (equipo_tipo, posicion).
# ---------------------------------------------------------------------------
SECCIONES = {
    "1":  ("novedad",     "PROYECTO", "",                  "Registro de novedades"),
    "2":  ("frontera",    "PROYECTO", "",                  "Informacion general de la frontera"),
    "3":  ("medidor",     "EQUIPO",   "MEDIDOR_PRINCIPAL", "Medidores"),
    "4":  ("medidor",     "EQUIPO",   "MEDIDOR_PRINCIPAL", "Medidores"),
    "5":  ("medidor",     "EQUIPO",   "MEDIDOR_RESPALDO",  "Medidores"),
    "6":  ("tc",          "EQUIPO",   "TC",                "Transformadores de corriente"),
    "7":  ("tp",          "EQUIPO",   "TP",                "Transformadores de tension"),
    "8":  ("conductor",   "EQUIPO",   "CONDUCTOR",         "Conductores"),
    "9":  ("celda",       "EQUIPO",   "CELDA",             "Paneles o cajas de seguridad"),
    "10": ("bornera",     "EQUIPO",   "BORNERA",           "Bornera de prueba"),
    "11": ("modem",       "EQUIPO",   "MODEM_PRINCIPAL",   "Comunicaciones"),
    "12": ("modem",       "EQUIPO",   "MODEM_RESPALDO",    "Comunicaciones"),
    "16": ("frontera",    "PROYECTO", "",                  "Informacion general de la frontera"),
    "18": ("responsable", "PROYECTO", "",                  "Persona designada por el RF"),
}

SECCIONES_OMITIDAS = {"13", "14", "15", "17"}   # D-06

# (seccion, sub) -> nombre semantico del bloque. D-07.
BLOQUES = {
    ("3", "38"): "cert_conformidad", ("3", "40"): "cert_calibracion",
    ("4", "38"): "cert_conformidad", ("4", "40"): "cert_calibracion",
    ("5", "43"): "cert_conformidad", ("5", "45"): "cert_calibracion",
    ("6", "20"): "cert_calibracion", ("6", "21"): "cert_pruebas_rutina",
    ("7", "16"): "cert_calibracion", ("7", "17"): "cert_pruebas_rutina",
    ("10", "3"): "cert_conformidad",
}

# (seccion, sub) -> (equipo_tipo, etiqueta). Sub-bloques que son equipos distintos.
SUBGRUPOS = {
    ("8", "1"): ("CONDUCTOR_CORRIENTE", "Conductores de senal de corriente"),
    ("8", "2"): ("CONDUCTOR_TENSION",   "Conductores de senal de tension"),
    ("9", "1"): ("CELDA_MEDIDOR",       "Celda o caja del medidor"),
    ("9", "2"): ("CELDA_TRAFOS",        "Celda o caja de transformadores de medida"),
    ("9", "3"): ("CELDA_OTRA",          "Otras celdas del sistema de medida"),
}
# Dentro de esos sub-bloques, el bloque de certificado va un nivel mas abajo.
SUBGRUPOS_BLOQUES = {("8", "1", "7"): "cert_conformidad", ("8", "2", "7"): "cert_conformidad",
                     ("9", "1", "1"): "cert_conformidad", ("9", "2", "1"): "cert_conformidad",
                     ("9", "3", "1"): "cert_conformidad"}

# Prefijo del bloque de sellos en cada seccion. D-05.
SELLOS = {("3", "27"), ("4", "27"), ("5", "32"), ("6", "19"), ("7", "15"), ("10", "2")}
SELLOS_COLUMNAS = ["ubicacion", "serie", "tipo", "color",
                   "fecha_instalacion", "fecha_retiro", "propiedad"]

# Renombres explicitos: campos cuyo titulo no basta para nombrarlos bien, o que
# colisionarian entre si. Clave = (seccion, sub) -> (nombre, titulo para la UI).
RENOMBRES = {
    # Latitud y longitud comparten el titulo "Coordenadas (...)": hay que separarlas.
    ("2", "6"):  ("latitud",  "Coordenadas - latitud"),
    ("2", "7"):  ("longitud", "Coordenadas - longitud"),
    # Medidor: activa (sec 3) y reactiva (sec 4) conviven en el mismo equipo. D-03.
    ("3", "18"): ("indice_clase_activa",    "Indice de clase - activa (%)"),
    ("3", "19"): ("constante_activa",       "Constante - activa"),
    ("3", "20"): ("unidad_constante_activa", "Unidad de la constante - activa"),
    ("3", "25"): ("canal_impor_activa",     "Canal de importacion - activa"),
    ("3", "26"): ("canal_expor_activa",     "Canal de exportacion - activa"),
    ("4", "18"): ("indice_clase_reactiva",  "Indice de clase - reactiva (%)"),
    ("4", "19"): ("constante_reactiva",     "Constante - reactiva"),
    ("4", "20"): ("unidad_constante_reactiva", "Unidad de la constante - reactiva"),
    ("4", "25"): ("canal_impor_reactiva",   "Canal de importacion - reactiva"),
    ("4", "26"): ("canal_expor_reactiva",   "Canal de exportacion - reactiva"),
    # Seccion 5 ya trae ambas: se alinean con los nombres de arriba para que
    # principal y respaldo compartan definicion.
    ("5", "18"): ("indice_clase_activa",    "Indice de clase - activa (%)"),
    ("5", "19"): ("constante_activa",       "Constante - activa"),
    ("5", "20"): ("unidad_constante_activa", "Unidad de la constante - activa"),
    ("5", "21"): ("indice_clase_reactiva",  "Indice de clase - reactiva (%)"),
    ("5", "22"): ("constante_reactiva",     "Constante - reactiva"),
    ("5", "23"): ("unidad_constante_reactiva", "Unidad de la constante - reactiva"),
    ("5", "28"): ("canal_impor_activa",     "Canal de importacion - activa"),
    ("5", "29"): ("canal_expor_activa",     "Canal de exportacion - activa"),
    ("5", "30"): ("canal_impor_reactiva",   "Canal de importacion - reactiva"),
    ("5", "31"): ("canal_expor_reactiva",   "Canal de exportacion - reactiva"),
    # La seccion 5 abrevia dos titulos que la 3 escribe completos. Sin esto
    # quedaban dos parametros distintos para el mismo dato.
    ("5", "8"):  ("proveedor_o_representante", "Proveedor o representante"),
    ("5", "37"): ("cap_mem_mb", "Capacidad de memoria (MB)"),
    ("4", "8"):  ("proveedor_o_representante", "Proveedor o representante"),
    # TC: dos numeros de certificado, uno por relacion de transformacion.
    ("6", "20"): (None, None),   # el bloque se resuelve por BLOQUES
}

# De la seccion 4 SOLO sobreviven estos: el resto es el mismo medidor. D-03.
SEC4_SUPERVIVIENTES = {"18", "19", "20", "25", "26"}

# Instancias por tipo de equipo. El formato declara los TC una vez por fase
# (R/S/T) y los TP una sola vez, aunque la frontera lleva tres. Ver decision D-09.
INSTANCIAS = {"TC": 3, "TP": 3}
ETIQUETAS_DEFECTO = {"TC": ["Fase R", "Fase S", "Fase T"],
                     "TP": ["Fase R", "Fase S", "Fase T"]}

TIPOS = {"Texto": "TEXTO", "Número": "NUMERO", "Numero": "NUMERO", "Fecha": "FECHA",
         "Categoría": "CATEGORIA", "Categoria": "CATEGORIA", "Lista": "LISTA"}


def slug(texto: str) -> str:
    """'Índ. Clase (%)' -> 'ind_clase'. Sin tildes; se quitan unidades entre parentesis."""
    t = re.sub(r"\([^)]*\)", " ", texto)
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = t.lower().replace("ñ", "n")
    t = re.sub(r"[^a-z0-9]+", "_", t).strip("_")
    return re.sub(r"_+", "_", t) or "campo"


# El formato numera "18. Persona designada por el RF" pero sus tres campos
# quedaron rotulados 17.1/17.2/17.3, y el 17 es "Anexo diagrama unifilar" (sin
# campos). Se corrige el numeral al leer. Ver decision D-08.
CORRECCION_IDS = {"17.1": "18.1", "17.2": "18.2", "17.3": "18.3"}


def _etiqueta_instancia(titulo: str, base: str) -> str:
    """'Número de Serie (Fase R)' + base 'Número de Serie' -> 'Fase R'."""
    sufijo = titulo[len(base):].strip() if titulo.startswith(base) else ""
    m = re.match(r"^\((.+)\)$", sufijo)
    return m.group(1) if m else (sufijo or titulo)


def leer_diccionario(ruta: str) -> list[dict]:
    """Lee el diccionario de campos respetando su layout de dos niveles.

    Los transformadores de corriente no declaran un campo por fase: declaran el
    campo una vez (fila con numeral y bandera de bloque, sin tipo de dato) y
    debajo una fila SIN numeral por cada fase R/S/T con el tipo. Esas filas hijas
    son instancias del MISMO parametro, no parametros distintos -- leerlas como
    campos sueltos duplicaba el catalogo, e ignorarlas perdia la seccion entera.
    """
    ws = openpyxl.load_workbook(ruta, data_only=True)["Definición de campos"]

    def celda(fila: int, col: int) -> str:
        v = ws.cell(fila, col).value
        return "" if v is None else " ".join(str(v).split())

    filas: list[dict] = []
    padre: dict | None = None

    for r in range(2, ws.max_row + 1):
        ident = CORRECCION_IDS.get(celda(r, 1).rstrip("."), celda(r, 1).rstrip("."))
        titulo, bloque, tipo = celda(r, 2), celda(r, 3), celda(r, 4)

        # Fila hija: sin numeral propio y con tipo de dato -> instancia del padre.
        if not ident and tipo and padre is not None:
            padre["instancias_etiquetas"].append(_etiqueta_instancia(titulo, padre["titulo"]))
            padre["tipo"] = padre["tipo"] or TIPOS.get(tipo, "TEXTO")
            padre["requerido"] = padre["requerido"] or celda(r, 5) == "1"
            padre["descripcion"] = padre["descripcion"] or celda(r, 6)
            continue

        if padre is not None:                      # se cierra el padre anterior
            if padre["tipo"]:
                filas.append(padre)
            padre = None

        if not ident:
            continue

        # Cabecera de bloque sin tipo: puede ser una seccion o un padre con hijas.
        if bloque and not tipo:
            if titulo.isupper():                   # cabecera de seccion: se ignora
                continue
            padre = {"id": ident, "titulo": titulo, "tipo": "", "requerido": False,
                     "descripcion": "", "instancias_etiquetas": []}
            continue

        if not tipo:
            continue

        filas.append({"id": ident, "titulo": titulo, "tipo": TIPOS.get(tipo, "TEXTO"),
                      "requerido": celda(r, 5) == "1", "descripcion": celda(r, 6),
                      "instancias_etiquetas": []})

    if padre is not None and padre["tipo"]:
        filas.append(padre)
    return filas


def resolver(f: dict) -> dict | None:
    """Traduce una fila del formato a una definicion de parametro (o None si se descarta)."""
    partes = f["id"].split(".")
    sec = partes[0]
    if sec in SECCIONES_OMITIDAS or sec not in SECCIONES:
        return None

    grupo, ambito, equipo_tipo, etiqueta = SECCIONES[sec]
    sub = partes[1] if len(partes) > 1 else ""
    sub2 = partes[2] if len(partes) > 2 else ""
    titulo, requerido = f["titulo"], f["requerido"]

    # D-03: de la seccion 4 solo pasan los campos realmente distintos.
    if sec == "4" and sub not in SEC4_SUPERVIVIENTES:
        return None

    # D-05: todo el bloque de sellos colapsa en un unico parametro TABLA.
    if (sec, sub) in SELLOS:
        return {"clave": f"{grupo}.sellos", "titulo": "Sellos instalados", "tipo": "TABLA",
                "requerido": False, "ambito": ambito, "equipo_tipos": [equipo_tipo],
                "grupo": grupo, "etiqueta_grupo": etiqueta,
                "instancias": INSTANCIAS.get(equipo_tipo, 1),
                "etiquetas": ETIQUETAS_DEFECTO.get(equipo_tipo, []),
                "origen_hv": [f"{sec}.{sub}"],
                "columnas": SELLOS_COLUMNAS,
                "descripcion": "Tabla de sellos del equipo (una fila por sello)."}

    # Sub-bloques que en realidad son equipos distintos (conductores, celdas).
    if (sec, sub) in SUBGRUPOS:
        equipo_tipo, etiqueta = SUBGRUPOS[(sec, sub)]
        bloque = SUBGRUPOS_BLOQUES.get((sec, sub, sub2))
        nombre = f"{bloque}_{slug(titulo)}" if bloque else slug(titulo)
    else:
        renombre = RENOMBRES.get((sec, sub))
        if renombre and renombre[0]:
            nombre, titulo = renombre
        elif (sec, sub) in BLOQUES:
            nombre = f"{BLOQUES[(sec, sub)]}_{slug(titulo)}"
        else:
            nombre = slug(titulo)

    etiquetas = f.get("instancias_etiquetas") or ETIQUETAS_DEFECTO.get(equipo_tipo, [])
    return {"clave": f"{grupo}.{nombre}", "titulo": titulo, "tipo": f["tipo"],
            "requerido": requerido, "ambito": ambito, "equipo_tipos": [equipo_tipo],
            "grupo": grupo, "etiqueta_grupo": etiqueta,
            "instancias": max(INSTANCIAS.get(equipo_tipo, 1), len(etiquetas) or 1),
            "etiquetas": etiquetas, "origen_hv": [f["id"]],
            "descripcion": f["descripcion"]}


def construir_catalogo(filas: list[dict]) -> "OrderedDict[str, dict]":
    cat: "OrderedDict[str, dict]" = OrderedDict()
    for f in filas:
        p = resolver(f)
        if p is None:
            continue
        existente = cat.get(p["clave"])
        if existente is None:
            cat[p["clave"]] = p
            continue
        # Misma clave => es el mismo dato. Se une el alcance en vez de duplicar. D-04.
        for et in p["equipo_tipos"]:
            if et not in existente["equipo_tipos"]:
                existente["equipo_tipos"].append(et)
        for hv in p["origen_hv"]:
            if hv not in existente["origen_hv"]:
                existente["origen_hv"].append(hv)
        existente["requerido"] = existente["requerido"] or p["requerido"]
        existente["instancias"] = max(existente["instancias"], p["instancias"])
        if not existente.get("etiquetas") and p.get("etiquetas"):
            existente["etiquetas"] = p["etiquetas"]
    return cat


CABECERA = '''"""Catalogo de parametros unicos del sistema de medida (proceso SIC/ASIC).

GENERADO por scripts/generar_catalogo_parametros.py a partir de la hoja
"Definicion de campos" del formato oficial de Hoja de Vida (CREG 038/2014).
No editar a mano: ajustar el generador y regenerar.

Cada entrada es un dato UNICO del proyecto. Los documentos que lo contienen no
lo repiten: lo referencian por su clave (ver mapa_documentos.py). Ese es el
principio rector del modulo -- un dato, una fila, una fuente de verdad.

  clave         identificador estable "<grupo>.<nombre>", unico en el catalogo
  titulo        etiqueta para la UI
  tipo          TEXTO | NUMERO | FECHA | CATEGORIA | LISTA | TABLA
  requerido     lo exige el formato oficial
  ambito        PROYECTO (un valor por proyecto) | EQUIPO (uno por equipo y posicion)
  equipo_tipos  tipos de equipo a los que aplica la definicion (ambito EQUIPO)
  instancias    cuantos equipos de ese tipo admite la frontera (TC y TP: 3)
  etiquetas     nombre de cada instancia cuando aplica (p. ej. Fase R/S/T)
  origen_hv     numerales de la Hoja de Vida de donde sale (trazabilidad)
  columnas      solo para tipo TABLA: columnas de la sub-tabla
"""

from __future__ import annotations

'''


def emitir(cat: "OrderedDict[str, dict]", destino: str) -> None:
    partes = [CABECERA, "PARAMETROS: list[dict] = [\n"]
    grupo_actual = None
    for p in cat.values():
        if p["etiqueta_grupo"] != grupo_actual:
            grupo_actual = p["etiqueta_grupo"]
            partes.append(f"\n    # --- {grupo_actual} " + "-" * max(4, 62 - len(grupo_actual)) + "\n")
        campos = [f'"clave": "{p["clave"]}"', f'"titulo": {p["titulo"]!r}',
                  f'"tipo": "{p["tipo"]}"', f'"requerido": {p["requerido"]}',
                  f'"ambito": "{p["ambito"]}"', f'"equipo_tipos": {p["equipo_tipos"]!r}',
                  f'"grupo": "{p["grupo"]}"', f'"instancias": {p["instancias"]}',
                  f'"origen_hv": {p["origen_hv"]!r}']
        if p.get("etiquetas"):
            campos.append(f'"etiquetas": {p["etiquetas"]!r}')
        if p.get("columnas"):
            campos.append(f'"columnas": {p["columnas"]!r}')
        partes.append("    {" + ", ".join(campos) + "},\n")
    partes.append("]\n\nPARAMETROS_POR_CLAVE: dict[str, dict] = {p['clave']: p for p in PARAMETROS}\n")
    with open(destino, "w", encoding="utf-8") as fh:
        fh.writelines(partes)


def main() -> None:
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    filas = leer_diccionario(sys.argv[1])
    cat = construir_catalogo(filas)
    emitir(cat, sys.argv[2])
    equipos = {e for p in cat.values() for e in p["equipo_tipos"] if e}
    print(f"{len(filas)} campos del formato -> {len(cat)} parametros unicos "
          f"({len(equipos)} tipos de equipo) -> {sys.argv[2]}")
    avisar_casi_duplicados(cat)


def avisar_casi_duplicados(cat: "OrderedDict[str, dict]") -> None:
    """Avisa de claves del mismo grupo donde una es prefijo de la otra.

    Suele significar que el formato escribio el mismo campo con dos titulos
    ("Proveedor" y "Proveedor o representante") y quedaron dos parametros para
    un solo dato -- justo lo que este catalogo existe para evitar.
    """
    claves = sorted(cat)
    sospechosas = [
        (a, b) for i, a in enumerate(claves) for b in claves[i + 1:]
        if a.split(".")[0] == b.split(".")[0] and b.startswith(a + "_")
    ]
    if sospechosas:
        print()
        print("  AVISO: posibles duplicados (revisar y renombrar si son el mismo dato):")
        for a, b in sospechosas:
            print(f"    {a}  ~  {b}")


if __name__ == "__main__":
    main()

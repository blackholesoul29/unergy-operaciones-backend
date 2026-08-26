"""El ER que recibe el inversionista.

Es un entregable, no un volcado: va con fórmulas vivas para que se pueda auditar,
sin cuadrícula, con la marca, y con el resumen arriba -- que es lo que se mira
primero.
"""
import io
import types

from openpyxl import load_workbook

from app.services.er_export import generar_er_xlsx


def _linea(grupo, concepto, valor, inv="QUANTUM ENERGY S.A.S", pct=100.0):
    return types.SimpleNamespace(grupo=grupo, concepto=concepto, valor_cop=valor,
                                 inversionista_nombre=inv, porcentaje=pct, orden=0)


PANEL = types.SimpleNamespace(
    periodo="2026-07",
    comercializador="UNERGY ENERGIA DIGITAL S.A.S ESP",
    lineas=[
        _linea("ingresos", "Venta en bolsa", 118_673_860.5),
        _linea("comercializacion", "Energía en Bolsa (Gen)", -603_083.64),
        _linea("comercializacion", "Arranque y parada", -171_129.13),
        _linea("costos", "Servicio de Internet", -64_706.31),
        _linea("facturas", "CGM", -1_011_610.02),
    ],
)

DIARIO = [
    {"fecha": "2026-07-01", "generacion_kwh": 5629.63, "importacion_kwh": 22.52,
     "venta_kwh": 5629.63, "venta_cop": 4_491_783.26},
    {"fecha": "2026-07-02", "generacion_kwh": 3597.46, "importacion_kwh": 23.62,
     "venta_kwh": 3597.46, "venta_cop": 2_839_856.69},
]


def _hoja(contenido):
    return load_workbook(io.BytesIO(contenido)).active


def _celdas(contenido):
    return [c.value for f in _hoja(contenido).iter_rows() for c in f]


# ── Es un entregable, no un volcado ──────────────────────────────────────────

def test_no_muestra_la_cuadricula():
    """Una cuadrícula de fondo hace que parezca una hoja de trabajo."""
    assert _hoja(generar_er_xlsx(PANEL, "X", DIARIO)).sheet_view.showGridLines is False


def test_lleva_el_titulo_y_el_proyecto():
    valores = _celdas(generar_er_xlsx(PANEL, "GD Agustín 3", DIARIO))
    assert "ESTADO DE RESULTADOS" in valores
    assert "GD Agustín 3" in valores
    assert any(isinstance(v, str) and "2026-07" in v for v in valores)


def test_el_resumen_va_arriba_del_detalle():
    """Es lo que mira el inversionista primero."""
    hoja = _hoja(generar_er_xlsx(PANEL, "X", DIARIO))
    pos = {}
    for f in hoja.iter_rows():
        for c in f:
            if c.value in ("RESUMEN DEL PERÍODO", "INGRESOS", "DETALLE DIARIO"):
                pos.setdefault(c.value, c.row)
    assert pos["RESUMEN DEL PERÍODO"] < pos["INGRESOS"] < pos["DETALLE DIARIO"]


def test_el_resumen_trae_utilidad_y_tarifa():
    valores = _celdas(generar_er_xlsx(PANEL, "X", DIARIO))
    for etiqueta in ("Energía generada", "Ingresos brutos", "UTILIDAD DEL PERÍODO",
                     "Tarifa neta"):
        assert etiqueta in valores, etiqueta


# ── Fórmulas vivas ───────────────────────────────────────────────────────────

def test_los_totales_son_formulas_no_valores():
    """Pegar el número impide auditar de dónde sale."""
    formulas = [v for v in _celdas(generar_er_xlsx(PANEL, "X", DIARIO))
                if isinstance(v, str) and v.startswith("=")]
    assert any(f.startswith("=SUM(") for f in formulas)


def test_las_funciones_van_en_ingles_para_verse_en_espanol():
    """Un .xlsx guarda las funciones SIEMPRE en inglés y Excel las localiza:
    escribir SUMA daría #¿NOMBRE? en cualquier idioma."""
    formulas = [v for v in _celdas(generar_er_xlsx(PANEL, "X", DIARIO))
                if isinstance(v, str) and v.startswith("=")]
    assert not any("SUMA(" in f for f in formulas)


def test_la_utilidad_suma_las_lineas_del_resumen():
    formulas = [v for v in _celdas(generar_er_xlsx(PANEL, "X", DIARIO))
                if isinstance(v, str) and v.startswith("=SUM(D")]
    assert formulas, "la utilidad debe salir de sumar el resumen"


def test_la_tarifa_no_divide_por_cero():
    """Sin energía, la fórmula tiene que protegerse sola dentro del Excel."""
    formulas = [v for v in _celdas(generar_er_xlsx(PANEL, "X", diario=[]))
                if isinstance(v, str) and v.startswith("=IF(")]
    assert formulas, "la tarifa debe llevar guarda de división por cero"


def test_el_total_diario_es_formula():
    hoja = _hoja(generar_er_xlsx(PANEL, "X", DIARIO))
    fs = [c.value for f in hoja.iter_rows() for c in f
          if isinstance(c.value, str) and c.value.startswith("=SUM(C")]
    assert fs


# ── Formato ──────────────────────────────────────────────────────────────────

def test_la_participacion_va_en_porcentaje():
    """El 1.0 crudo se leía como "1", no como 100%."""
    hoja = _hoja(generar_er_xlsx(PANEL, "X", DIARIO, inversionista="QUANTUM ENERGY S.A.S"))
    pcts = [c for f in hoja.iter_rows() for c in f if c.number_format == "0.0%"]
    assert pcts and pcts[0].value == 1.0


def test_los_valores_van_en_moneda():
    hoja = _hoja(generar_er_xlsx(PANEL, "X", DIARIO))
    assert any('"$"' in (c.number_format or "") for f in hoja.iter_rows() for c in f)


def test_la_energia_lleva_su_unidad():
    hoja = _hoja(generar_er_xlsx(PANEL, "X", DIARIO))
    assert any("kWh" in (c.number_format or "") for f in hoja.iter_rows() for c in f)


# ── Contenido ────────────────────────────────────────────────────────────────

def test_trae_los_tres_bloques():
    valores = _celdas(generar_er_xlsx(PANEL, "X", DIARIO))
    for t in ("INGRESOS", "COMERCIALIZACIÓN (XM)", "COSTOS OPERATIVOS", "DETALLE DIARIO"):
        assert t in valores, t


def test_la_tabla_diaria_trae_la_importacion():
    """El consumo es una columna del ER, no un dato aparte."""
    valores = _celdas(generar_er_xlsx(PANEL, "X", DIARIO))
    assert any(isinstance(v, str) and "Importación" in v for v in valores)
    assert 22.52 in valores


def test_el_encabezado_de_venta_nombra_al_comercializador():
    valores = _celdas(generar_er_xlsx(PANEL, "X", DIARIO))
    assert any(isinstance(v, str) and "UNERGY ENERGIA DIGITAL" in v for v in valores)


def test_un_bloque_vacio_lo_dice_en_vez_de_quedar_en_blanco():
    """Agustín 3 no tiene costos cargados: hay que verlo, no adivinarlo."""
    panel = types.SimpleNamespace(periodo="2026-07", comercializador="X", lineas=[
        _linea("ingresos", "Venta", 100.0)])
    assert "Sin movimientos en el período" in _celdas(generar_er_xlsx(panel, "X", []))


def test_sin_datos_diarios_no_revienta():
    assert _celdas(generar_er_xlsx(PANEL, "X", diario=[]))


# ── Filtro por inversionista ─────────────────────────────────────────────────

def test_filtrar_por_inversionista_deja_solo_lo_suyo():
    panel = types.SimpleNamespace(periodo="2026-07", comercializador="X", lineas=[
        _linea("ingresos", "Venta", 100.0, inv="ACME", pct=60.0),
        _linea("ingresos", "Venta", 40.0, inv="OTRA", pct=40.0),
    ])
    valores = _celdas(generar_er_xlsx(panel, "X", [], inversionista="ACME"))
    assert 100.0 in valores and 40.0 not in valores


def test_el_encabezado_nombra_al_inversionista():
    panel = types.SimpleNamespace(periodo="2026-07", comercializador="X", lineas=[
        _linea("ingresos", "Venta", 100.0, inv="ACME", pct=60.0)])
    valores = _celdas(generar_er_xlsx(panel, "X", [], inversionista="ACME"))
    assert any(isinstance(v, str) and "ACME" in v for v in valores)

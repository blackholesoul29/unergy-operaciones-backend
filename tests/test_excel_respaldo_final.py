"""generar_excel_dia() (excel.py) -- columnas Consumo_Respaldo/
Generación_Respaldo.

Hasta ahora eran una fórmula de Excel independiente (=C*RAND()), que
ignoraba curva_respaldo_final -- el Excel mostraba un número inventado
distinto cada vez que se recalculaba, sin relación con el dato real del
medidor cuando aplicaba, ni con lo que /enviar realmente manda a Quoia
(pedido de Sara 2026-08-26: "que el excel refleje todo lo del envío").
"""
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

from app.models.base import Base
from app.models.fronteras import Frontera, TipoFronteraEnum
from app.models.reporte_energia import ReporteEnergiaGeneracion, ReporteEnergiaConsumo
from app.services.reporte_energia.excel import generar_excel_dia


@compiles(JSONB, "sqlite")
def _jsonb_as_text(element, compiler, **kw):
    return "TEXT"


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine, tables=[
        Frontera.__table__, ReporteEnergiaGeneracion.__table__, ReporteEnergiaConsumo.__table__,
    ])
    s = sessionmaker(bind=engine)()
    yield s
    s.close()


FECHA = date(2026, 8, 25)


def _leer_filas(xlsx_bytes):
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    return filas


def test_respaldo_gen_usa_curva_respaldo_final_no_formula(db):
    front = Frontera(id=1, nombre_frontera="Test Gen", tipo_frontera=TipoFronteraEnum.generacion)
    db.add(front)
    respaldo = [5.0] * 24
    db.add(ReporteEnergiaGeneracion(
        id=1, frontera_id=1, fecha=FECHA, caso=2, medidor_usado="principal",
        curva_final=[100.0] * 24, curva_respaldo_final=respaldo, respaldo_final_origen="medidor",
    ))
    db.commit()

    filas = _leer_filas(generar_excel_dia(db, FECHA))

    assert len(filas) == 24
    for fila in filas:
        # columnas: nombre_proyecto, Hora, Consumo_Principal, Generación_Principal, Consumo_Respaldo, Generación_Respaldo
        assert fila[5] == 5.0
        assert not str(fila[5]).startswith("=")


def test_respaldo_gen_en_blanco_cuando_cgm_ya_valido(db):
    """Si medidor_usado == 'cgm', /enviar no manda nada -- el Excel tampoco
    debe inventar un Backup."""
    front = Frontera(id=1, nombre_frontera="Test Gen", tipo_frontera=TipoFronteraEnum.generacion)
    db.add(front)
    db.add(ReporteEnergiaGeneracion(
        id=1, frontera_id=1, fecha=FECHA, caso=1, medidor_usado="cgm",
        curva_final=[100.0] * 24, curva_respaldo_final=[999.0] * 24, respaldo_final_origen="estimado",
    ))
    db.commit()

    filas = _leer_filas(generar_excel_dia(db, FECHA))

    assert all(fila[5] is None for fila in filas)


def test_respaldo_gen_sin_curva_final_cae_al_calculo_en_vivo(db):
    """Fila vieja, sin curva_respaldo_final persistida -- se calcula al
    vuelo con el mismo criterio que /enviar (curva_respaldo_a_reportar)."""
    front = Frontera(id=1, nombre_frontera="Test Gen", tipo_frontera=TipoFronteraEnum.generacion)
    db.add(front)
    db.add(ReporteEnergiaGeneracion(
        id=1, frontera_id=1, fecha=FECHA, caso=2, medidor_usado="principal",
        curva_final=[100.0] * 24, curva_respaldo_final=None,
    ))
    db.commit()

    filas = _leer_filas(generar_excel_dia(db, FECHA))

    # Sin medidor de respaldo con qué comparar -- cae a estimado ±1% de 100
    assert all(99.0 <= fila[5] <= 101.0 for fila in filas)


def test_respaldo_con_usa_curva_respaldo_final(db):
    front = Frontera(id=1, nombre_frontera="Test", tipo_frontera=TipoFronteraEnum.generacion, proyecto_id=10)
    db.add(front)
    db.add(ReporteEnergiaGeneracion(
        id=1, frontera_id=1, fecha=FECHA, caso=2, medidor_usado="principal", curva_final=[100.0] * 24,
        curva_respaldo_final=[1.0] * 24,
    ))
    front_con = Frontera(id=2, nombre_frontera="Test Consumo", tipo_frontera=TipoFronteraEnum.consumo_auxiliar, proyecto_id=10)
    db.add(front_con)
    respaldo_con = [3.0] * 24
    db.add(ReporteEnergiaConsumo(
        id=1, frontera_id=2, fecha=FECHA, caso="Medidor", medidor_usado="principal",
        curva_final=[10.0] * 24, curva_respaldo_final=respaldo_con, respaldo_final_origen="medidor",
    ))
    db.commit()

    filas = _leer_filas(generar_excel_dia(db, FECHA))

    assert len(filas) == 24
    for fila in filas:
        assert fila[4] == 3.0  # Consumo_Respaldo


def test_respaldo_con_en_blanco_cuando_caso_cgm(db):
    front = Frontera(id=1, nombre_frontera="Test", tipo_frontera=TipoFronteraEnum.generacion, proyecto_id=10)
    db.add(front)
    db.add(ReporteEnergiaGeneracion(
        id=1, frontera_id=1, fecha=FECHA, caso=2, medidor_usado="principal", curva_final=[100.0] * 24,
        curva_respaldo_final=[1.0] * 24,
    ))
    front_con = Frontera(id=2, nombre_frontera="Test Consumo", tipo_frontera=TipoFronteraEnum.consumo_auxiliar, proyecto_id=10)
    db.add(front_con)
    db.add(ReporteEnergiaConsumo(
        id=1, frontera_id=2, fecha=FECHA, caso="CGM", medidor_usado="cgm",
        curva_final=[10.0] * 24, curva_respaldo_final=[999.0] * 24,
    ))
    db.commit()

    filas = _leer_filas(generar_excel_dia(db, FECHA))

    assert all(fila[4] is None for fila in filas)
